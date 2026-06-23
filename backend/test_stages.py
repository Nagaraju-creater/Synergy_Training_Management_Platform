import asyncio
import openpyxl
import io
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule

import app.models.registry  # noqa: F401
from app.database import AsyncSessionLocal
from app.analytics.team_analytics_service import TeamAnalyticsService
from app.enrollments.models import Enrollment, EnrollmentStatus
from app.trainings.models import Training, TrainingStatus
from app.trainings.categories import TrainingCategory
from app.departments.models import Department
from app.effectiveness.models import Effectiveness, EffectivenessStatus
from app.attendance.models import AttendanceRecord, AttendanceStatus
from app.attendance.service import AttendanceService
from app.employees.models import Employee
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload


async def generate_stage(stage_num: int, include_charts: bool, filename: str):
    print(f"Generating Stage {stage_num} (include_charts={include_charts}) -> {filename}...")
    async with AsyncSessionLocal() as db:
        # We will replicate the generate_kpi_excel logic but prune sheets based on stage_num
        financial_year = "2025-2026"
        fy_start, fy_end, fy_label = TeamAnalyticsService.financial_year_bounds(financial_year)
        months = TeamAnalyticsService.fy_months(fy_start)
        
        # Metadata queries
        depts_stmt = select(Department).where(Department.deleted_at == None).order_by(Department.name.asc())
        depts_res = (await db.execute(depts_stmt)).scalars().all()
        dept_names = [d.name for d in depts_res]
        
        cats_stmt = select(TrainingCategory).order_by(TrainingCategory.name.asc())
        cats_res = (await db.execute(cats_stmt)).scalars().all()
        cat_names = [c.name for c in cats_res]
        
        mgrs_stmt = select(Employee).where(Employee.id.in_(select(Employee.manager_id).where(Employee.manager_id != None))).order_by(Employee.first_name.asc())
        mgrs_res = (await db.execute(mgrs_stmt)).scalars().all()
        mgr_names = [f"{m.first_name} {m.last_name}" for m in mgrs_res]
        
        training_types = ["INTERNAL", "EXTERNAL", "ONLINE", "WORKSHOP", "CERTIFICATION"]
        
        all_employees_stmt = select(Employee).options(selectinload(Employee.department), selectinload(Employee.manager)).where(Employee.deleted_at == None).order_by(Employee.employee_code.asc())
        all_employees = (await db.execute(all_employees_stmt)).scalars().all()
        
        all_attendance_stmt = (
            select(AttendanceRecord, Training, Employee, Enrollment)
            .join(Training, AttendanceRecord.training_id == Training.id)
            .join(Employee, AttendanceRecord.employee_id == Employee.id)
            .join(Enrollment, and_(
                Enrollment.employee_id == AttendanceRecord.employee_id,
                Enrollment.training_id == AttendanceRecord.training_id,
                Enrollment.deleted_at == None,
            ))
            .options(
                selectinload(Employee.department),
                selectinload(Employee.manager),
                selectinload(Training.category)
            )
            .where(and_(
                Employee.deleted_at == None,
                AttendanceService.roster_record_filter(),
                Training.status != TrainingStatus.CANCELLED,
                Training.start_date >= fy_start,
                Training.start_date <= fy_end,
            ))
            .order_by(Training.start_date.asc())
        )
        all_attendance_results = (await db.execute(all_attendance_stmt)).all()
        
        contribution_map = {emp.id: {} for emp in all_employees}
        attendance_rows = []
        for record, training, employee, enrollment in all_attendance_results:
            key = (training.start_date.year, training.start_date.month) if training.start_date else None
            att_status_val = record.status.value if hasattr(record.status, "value") else str(record.status)
            contribution = {
                "employee_id": employee.id,
                "employee_name": f"{employee.first_name} {employee.last_name}",
                "employee_code": employee.employee_code,
                "department": employee.department.name if employee.department else "Unassigned",
                "manager": f"{employee.manager.first_name} {employee.manager.last_name}" if employee.manager else "",
                "training_id": training.id,
                "training_title": training.title,
                "training_date": training.start_date,
                "hours": float(training.duration_hours or 0),
                "max_hours": float(training.max_hours_allowed or training.duration_hours or 0),
                "status": att_status_val,
                "marked_at": record.marked_at,
                "category": training.category.name if training.category else "",
                "trainer": training.trainer_name or "",
                "duration": float(training.duration_hours or 0),
            }
            if att_status_val in ("PRESENT", "LATE") and enrollment.status == EnrollmentStatus.COMPLETED:
                contribution_map.setdefault(employee.id, {}).setdefault(key, []).append(contribution)
            attendance_rows.append(contribution)
            
        wb = openpyxl.Workbook()
        
        # Sheet 1: Employee KPI Report (always present)
        emp_report_ws = wb.active
        emp_report_ws.title = "Employee KPI Report"
        emp_report_ws.views.sheetView[0].showGridLines = True
        
        # Populate Sheet 1
        blue = "4F81BD"
        deep_blue = "1F4E79"
        light_blue = "D9EAF7"
        green = "E2F0D9"
        yellow = "FFF2CC"
        red = "FCE4D6"
        white = "FFFFFF"
        thin = Side(style="thin", color="9EADCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        double_bottom_side = Side(style="double", color="1F4E79")
        double_bottom_border = Border(left=thin, right=thin, top=thin, bottom=double_bottom_side)
        
        def style_cell(cell, fill=white, font_color="1F2937", bold=False, align="center", size=9):
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Calibri", size=size, bold=bold, color=font_color)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = border
            
        total_cols = 4 + len(months) * 3 + 8
        emp_report_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        emp_report_ws.cell(1, 1, "Department KPI Dashboard - Employee Learning KPI Report")
        style_cell(emp_report_ws.cell(1, 1), fill=deep_blue, font_color=white, bold=True, size=14)
        emp_report_ws.row_dimensions[1].height = 30
        
        emp_report_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        filter_summary = f"FY: Apr {fy_start.year} - Mar {fy_end.year} | Date Range: {fy_start:%d-%b-%Y} to {fy_end:%d-%b-%Y}"
        emp_report_ws.cell(2, 1, filter_summary)
        style_cell(emp_report_ws.cell(2, 1), fill=light_blue, bold=True, align="left", size=10)
        emp_report_ws.row_dimensions[2].height = 24
        
        base_headers = ["S.No", "Employee ID", "Employee Name", "Date of Joining"]
        for idx, header in enumerate(base_headers, 1):
            emp_report_ws.merge_cells(start_row=4, start_column=idx, end_row=5, end_column=idx)
            emp_report_ws.cell(4, idx, header)
            style_cell(emp_report_ws.cell(4, idx), fill=blue, font_color=white, bold=True, size=9)
            
        col_idx = 5
        for _, _, month_name in months:
            emp_report_ws.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx + 2)
            emp_report_ws.cell(4, col_idx, month_name.upper())
            style_cell(emp_report_ws.cell(4, col_idx), fill=blue, font_color=white, bold=True, size=9)
            for sub in ["Date", "Actual", "Max Hours"]:
                emp_report_ws.cell(5, col_idx, sub)
                style_cell(emp_report_ws.cell(5, col_idx), fill=blue, font_color=white, bold=True, size=9)
                col_idx += 1
                
        final_headers = ["Total Hours", "Target Hours", "Balance", "KPI %", "Learning Compliance", "Department", "Manager", "Attendance %"]
        for header in final_headers:
            emp_report_ws.merge_cells(start_row=4, start_column=col_idx, end_row=5, end_column=col_idx)
            emp_report_ws.cell(4, col_idx, header)
            style_cell(emp_report_ws.cell(4, col_idx), fill=blue, font_color=white, bold=True, size=9)
            col_idx += 1
            
        emp_report_ws.row_dimensions[4].height = 20
        emp_report_ws.row_dimensions[5].height = 20
        
        sorted_employees = sorted(all_employees, key=lambda e: e.department.name if e.department else "Unassigned")
        from itertools import groupby
        def get_dept_name(e):
            return e.department.name if e.department else "Unassigned"
            
        current_row = 6
        subtotal_rows = []
        s_no_counter = 1
        
        for dept_name, dept_group in groupby(sorted_employees, key=get_dept_name):
            dept_employees = list(dept_group)
            d_start_row = current_row
            for emp in dept_employees:
                r = current_row
                row_values = [
                    s_no_counter,
                    emp.employee_code,
                    f"{emp.first_name} {emp.last_name}",
                    emp.date_of_joining.strftime("%d %m %Y") if emp.date_of_joining else "",
                ]
                s_no_counter += 1
                total_hours = 0.0
                total_attendance_rows = 0
                for y, m, _ in months:
                    entries = contribution_map.get(emp.id, {}).get((y, m), [])
                    total_attendance_rows += len(entries)
                    month_hours = sum(item["hours"] for item in entries)
                    max_hours = sum(item["max_hours"] for item in entries)
                    total_hours += month_hours
                    dates = "\n".join(f"{item['training_date']:%d.%m.%Y} - {item['training_title']}" for item in entries)
                    row_values.extend([dates, month_hours if month_hours else "", max_hours if max_hours else ""])
                    
                target = 16.0
                balance = max(target - total_hours, 0)
                kpi_pct = round((total_hours / target) * 100, 1) if target else 0
                department = emp.department.name if emp.department else "Unassigned"
                manager = f"{emp.manager.first_name} {emp.manager.last_name}" if emp.manager else ""
                attendance_pct = 100.0 if total_attendance_rows else 0.0
                
                row_values.extend([
                    round(total_hours, 2),
                    target,
                    f"{round(balance, 2)} Hours Remaining" if balance else "0",
                    f"{kpi_pct}%",
                    TeamAnalyticsService.compliance_status(kpi_pct),
                    department,
                    manager,
                    f"{attendance_pct:.1f}%",
                ])
                emp_report_ws.append(row_values)
                compliance_fill = green if kpi_pct >= 75 else yellow if kpi_pct >= 25 else red
                for c in range(1, total_cols + 1):
                    fill = compliance_fill if c in (total_cols - 3, total_cols - 4) else white
                    style_cell(emp_report_ws.cell(r, c), fill=fill, align="left" if c in (3, 5) or c > 4 and (c - 5) % 3 == 0 else "center")
                emp_report_ws.row_dimensions[r].height = 36
                current_row += 1
                
            d_end_row = current_row - 1
            r_sub = current_row
            subtotal_rows.append(r_sub)
            emp_report_ws.cell(r_sub, 1, "")
            emp_report_ws.cell(r_sub, 2, "")
            emp_report_ws.cell(r_sub, 3, f"{dept_name} Subtotal")
            emp_report_ws.cell(r_sub, 4, "")
            col_c = 5
            for _ in months:
                emp_report_ws.cell(r_sub, col_c, "")
                col_c += 1
                act_letter = get_column_letter(col_c)
                emp_report_ws.cell(r_sub, col_c, f"=SUM({act_letter}{d_start_row}:{act_letter}{d_end_row})")
                col_c += 1
                max_letter = get_column_letter(col_c)
                emp_report_ws.cell(r_sub, col_c, f"=SUM({max_letter}{d_start_row}:{max_letter}{d_end_row})")
                col_c += 1
            tot_letter = get_column_letter(col_c)
            emp_report_ws.cell(r_sub, col_c, f"=SUM({tot_letter}{d_start_row}:{tot_letter}{d_end_row})")
            col_c += 1
            targ_letter = get_column_letter(col_c)
            emp_report_ws.cell(r_sub, col_c, f"=SUM({targ_letter}{d_start_row}:{targ_letter}{d_end_row})")
            col_c += 1
            bal_letter = get_column_letter(col_c)
            emp_report_ws.cell(r_sub, col_c, f"=SUM({bal_letter}{d_start_row}:{bal_letter}{d_end_row})")
            col_c += 1
            kpi_letter = get_column_letter(col_c)
            emp_report_ws.cell(r_sub, col_c, f"=IF({targ_letter}{r_sub}>0, {tot_letter}{r_sub}/{targ_letter}{r_sub}, 0)")
            col_c += 1
            emp_report_ws.cell(r_sub, col_c, f"=IF({kpi_letter}{r_sub}>=1.0, \"Achieved\", IF({kpi_letter}{r_sub}>=0.75, \"Good\", IF({kpi_letter}{r_sub}>=0.5, \"Moderate\", IF({kpi_letter}{r_sub}>=0.25, \"Low\", \"Critical\"))))")
            col_c += 1
            emp_report_ws.cell(r_sub, col_c, dept_name)
            col_c += 1
            emp_report_ws.cell(r_sub, col_c, "")
            col_c += 1
            att_letter = get_column_letter(col_c)
            emp_report_ws.cell(r_sub, col_c, f"=AVERAGE({att_letter}{d_start_row}:{att_letter}{d_end_row})")
            for c in range(1, total_cols + 1):
                cell = emp_report_ws.cell(r_sub, c)
                style_cell(cell, fill=light_blue, bold=True, align="left" if c in (3, 5) or c > 4 and (c - 5) % 3 == 0 else "center")
                if c in (col_c - 4, col_c):
                    cell.number_format = "0.0%"
                elif c in (col_c - 8, col_c - 7, col_c - 6):
                    cell.number_format = "#,##0.0"
                elif c > 4 and (c - 5) % 3 != 0:
                    cell.number_format = "#,##0.0"
            emp_report_ws.row_dimensions[r_sub].height = 24
            current_row += 1
            
        r_master = current_row
        emp_report_ws.cell(r_master, 1, "")
        emp_report_ws.cell(r_master, 2, "")
        emp_report_ws.cell(r_master, 3, "Master Total")
        emp_report_ws.cell(r_master, 4, "")
        col_c = 5
        for _ in months:
            emp_report_ws.cell(r_master, col_c, "")
            col_c += 1
            act_letter = get_column_letter(col_c)
            sub_sum_str = "+".join(f"{act_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
            emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
            col_c += 1
            max_letter = get_column_letter(col_c)
            sub_sum_str = "+".join(f"{max_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
            emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
            col_c += 1
        tot_letter = get_column_letter(col_c)
        sub_sum_str = "+".join(f"{tot_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
        emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
        col_c += 1
        targ_letter = get_column_letter(col_c)
        sub_sum_str = "+".join(f"{targ_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
        emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
        col_c += 1
        bal_letter = get_column_letter(col_c)
        sub_sum_str = "+".join(f"{bal_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
        emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
        col_c += 1
        kpi_letter = get_column_letter(col_c)
        emp_report_ws.cell(r_master, col_c, f"=IF({targ_letter}{r_master}>0, {tot_letter}{r_master}/{targ_letter}{r_master}, 0)")
        col_c += 1
        emp_report_ws.cell(r_master, col_c, f"=IF({kpi_letter}{r_master}>=1.0, \"Achieved\", IF({kpi_letter}{r_master}>=0.75, \"Good\", IF({kpi_letter}{r_master}>=0.5, \"Moderate\", IF({kpi_letter}{r_master}>=0.25, \"Low\", \"Critical\"))))")
        col_c += 1
        emp_report_ws.cell(r_master, col_c, "")
        col_c += 1
        emp_report_ws.cell(r_master, col_c, "")
        col_c += 1
        att_letter = get_column_letter(col_c)
        sub_avg_str = ",".join(f"{att_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
        emp_report_ws.cell(r_master, col_c, f"=AVERAGE({sub_avg_str})")
        for c in range(1, total_cols + 1):
            cell = emp_report_ws.cell(r_master, c)
            style_cell(cell, fill=light_blue, bold=True, align="left" if c in (3, 5) or c > 4 and (c - 5) % 3 == 0 else "center")
            cell.border = double_bottom_border
            if c in (col_c - 4, col_c):
                cell.number_format = "0.0%"
            elif c in (col_c - 8, col_c - 7, col_c - 6):
                cell.number_format = "#,##0.0"
            elif c > 4 and (c - 5) % 3 != 0:
                cell.number_format = "#,##0.0"
        emp_report_ws.row_dimensions[r_master].height = 26
        widths = [8, 14, 24, 16] + [28, 12, 12] * len(months) + [12, 12, 18, 10, 20, 18, 22, 14]
        for i, w in enumerate(widths, 1):
            emp_report_ws.column_dimensions[get_column_letter(i)].width = w
        emp_report_ws.auto_filter.ref = f"A5:{get_column_letter(total_cols)}{r_master - 1}"
        
        # Stage 2 sheets
        if stage_num >= 2:
            month_ws = wb.create_sheet("Monthly Learning Summary")
            month_ws.views.sheetView[0].showGridLines = True
            month_ws.merge_cells("A1:G1")
            month_ws["A1"] = "Monthly Learning Performance Summary"
            style_cell(month_ws["A1"], fill=deep_blue, font_color=white, bold=True, size=14)
            month_ws.row_dimensions[1].height = 30
            for c_idx, h in enumerate(["Month", "Target Hours", "Actual Hours", "Achievement %", "Attendance %", "Training Count", "Monthly Trends"], 1):
                style_cell(month_ws.cell(3, c_idx, h), fill=blue, font_color=white, bold=True)
            month_ws.row_dimensions[3].height = 24
            for idx, (y, m, month_name) in enumerate(months):
                r = 4 + idx
                month_records = [item for item in attendance_rows if item["training_date"].year == y and item["training_date"].month == m]
                attended_present = sum(1 for item in month_records if item["status"] in ("PRESENT", "LATE"))
                total_enrolled = len(month_records)
                attendance_pct = (attended_present / total_enrolled) if total_enrolled > 0 else 1.0
                unique_trainings = len(set(item["training_id"] for item in month_records))
                month_hours = sum(item["hours"] for item in month_records)
                target_hours_val = len(all_employees) * 1.33
                month_ws.cell(r, 1, month_name)
                month_ws.cell(r, 2, round(target_hours_val, 1))
                month_ws.cell(r, 3, round(month_hours, 1))
                month_ws.cell(r, 4, f"=IF(B{r}>0, C{r}/B{r}, 0)")
                month_ws.cell(r, 5, attendance_pct)
                month_ws.cell(r, 6, unique_trainings)
                if idx == 0:
                    month_ws.cell(r, 7, "—")
                else:
                    month_ws.cell(r, 7, f"=IF(C{r-1}>0, (C{r}-C{r-1})/C{r-1}, 0)")
                for col_i in range(1, 8):
                    cell = month_ws.cell(r, col_i)
                    style_cell(cell, align="left" if col_i == 1 else "center")
                    if col_i in (2, 3):
                        cell.number_format = "#,##0.0"
                    elif col_i in (4, 5, 7):
                        cell.number_format = "0.0%"
                    elif col_i == 6:
                        cell.number_format = "#,##0"
                month_ws.row_dimensions[r].height = 20
            r_tot = 4 + len(months)
            month_ws.cell(r_tot, 1, "Total")
            month_ws.cell(r_tot, 2, f"=SUM(B4:B{r_tot-1})")
            month_ws.cell(r_tot, 3, f"=SUM(C4:C{r_tot-1})")
            month_ws.cell(r_tot, 4, f"=IF(B{r_tot}>0, C{r_tot}/B{r_tot}, 0)")
            month_ws.cell(r_tot, 5, f"=AVERAGE(E4:E{r_tot-1})")
            month_ws.cell(r_tot, 6, f"=SUM(F4:F{r_tot-1})")
            month_ws.cell(r_tot, 7, "")
            for col_i in range(1, 8):
                cell = month_ws.cell(r_tot, col_i)
                style_cell(cell, fill=light_blue, bold=True)
                if col_i in (2, 3):
                    cell.number_format = "#,##0.0"
                elif col_i in (4, 5):
                    cell.number_format = "0.0%"
                elif col_i == 6:
                    cell.number_format = "#,##0"
            month_ws.row_dimensions[r_tot].height = 22
            for i, w in enumerate([16, 16, 22, 16, 18, 16, 18], 1):
                month_ws.column_dimensions[get_column_letter(i)].width = w
                
        # Stage 3 sheets
        if stage_num >= 3:
            att_ws = wb.create_sheet("Training Attendance Summary")
            att_ws.views.sheetView[0].showGridLines = True
            att_ws.merge_cells("A1:J1")
            att_ws["A1"] = "Training Session Attendance Summary"
            style_cell(att_ws["A1"], fill=deep_blue, font_color=white, bold=True, size=14)
            att_ws.row_dimensions[1].height = 30
            for c_idx, h in enumerate(["Training Name", "Department", "Training Date", "Enrolled Count", "Present Count", "Absent Count", "Late Count", "Attendance %", "Trainer", "Duration (Hours)"], 1):
                style_cell(att_ws.cell(3, c_idx, h), fill=blue, font_color=white, bold=True)
            att_ws.row_dimensions[3].height = 24
            training_groups = {}
            for item in attendance_rows:
                training_groups.setdefault(item["training_id"], []).append(item)
            sorted_trainings = sorted(training_groups.values(), key=lambda items: items[0]["training_date"] or date.min)
            for idx, items in enumerate(sorted_trainings):
                r = 4 + idx
                first_item = items[0]
                depts = sorted(list(set(item["department"] for item in items if item["department"])))
                dept_str = ", ".join(depts) if depts else "Unassigned"
                enrolled = len(items)
                present = sum(1 for item in items if item["status"] == "PRESENT")
                absent = sum(1 for item in items if item["status"] == "ABSENT")
                late = sum(1 for item in items if item["status"] == "LATE")
                att_pct = ((present + late) / enrolled) if enrolled > 0 else 1.0
                att_ws.cell(r, 1, first_item["training_title"])
                att_ws.cell(r, 2, dept_str)
                att_ws.cell(r, 3, first_item["training_date"].strftime("%d-%b-%Y") if first_item["training_date"] else "")
                att_ws.cell(r, 4, enrolled)
                att_ws.cell(r, 5, present)
                att_ws.cell(r, 6, absent)
                att_ws.cell(r, 7, late)
                att_ws.cell(r, 8, att_pct)
                att_ws.cell(r, 9, first_item["trainer"])
                att_ws.cell(r, 10, first_item["duration"])
                for col_i in range(1, 11):
                    cell = att_ws.cell(r, col_i)
                    style_cell(cell, align="left" if col_i in (1, 2, 9) else "center")
                    if col_i in (4, 5, 6, 7):
                        cell.number_format = "#,##0"
                    elif col_i == 8:
                        cell.number_format = "0.0%"
                    elif col_i == 10:
                        cell.number_format = "#,##0.0"
                att_ws.row_dimensions[r].height = 20
            r_tot = 4 + len(sorted_trainings)
            att_ws.cell(r_tot, 1, "Total / Average")
            att_ws.cell(r_tot, 2, "")
            att_ws.cell(r_tot, 3, "")
            att_ws.cell(r_tot, 4, f"=SUM(D4:D{r_tot-1})")
            att_ws.cell(r_tot, 5, f"=SUM(E4:E{r_tot-1})")
            att_ws.cell(r_tot, 6, f"=SUM(F4:F{r_tot-1})")
            att_ws.cell(r_tot, 7, f"=SUM(G4:G{r_tot-1})")
            att_ws.cell(r_tot, 8, f"=IF(D{r_tot}>0, (E{r_tot}+G{r_tot})/D{r_tot}, 1.0)")
            att_ws.cell(r_tot, 9, "")
            att_ws.cell(r_tot, 10, f"=SUM(J4:J{r_tot-1})")
            for col_i in range(1, 11):
                cell = att_ws.cell(r_tot, col_i)
                style_cell(cell, fill=light_blue, bold=True)
                if col_i in (4, 5, 6, 7):
                    cell.number_format = "#,##0"
                elif col_i == 8:
                    cell.number_format = "0.0%"
                elif col_i == 10:
                    cell.number_format = "#,##0.0"
            att_ws.row_dimensions[r_tot].height = 22
            for i, w in enumerate([32, 24, 16, 14, 14, 14, 14, 14, 20, 16], 1):
                att_ws.column_dimensions[get_column_letter(i)].width = w
                
        # Stage 4 sheets (The Dashboard and supporting hidden sheets)
        if stage_num >= 4:
            dept_ws = wb.create_sheet("Department KPI Analytics")
            dept_ws.views.sheetView[0].showGridLines = True
            
            raw_ws = wb.create_sheet("Raw_Attendance_Data")
            raw_ws.views.sheetView[0].showGridLines = True
            raw_ws.sheet_state = "hidden"
            
            emp_summary_ws = wb.create_sheet("Employee_KPI_Summary")
            emp_summary_ws.views.sheetView[0].showGridLines = True
            emp_summary_ws.sheet_state = "hidden"
            
            list_ws = wb.create_sheet("Filter_Lists")
            list_ws.sheet_state = "hidden"
            
            # Populate hidden Filter_Lists
            list_ws.cell(1, 1, "Financial Years")
            list_ws.cell(1, 2, "Quarters")
            list_ws.cell(1, 3, "Months")
            list_ws.cell(1, 4, "Departments")
            list_ws.cell(1, 5, "Categories")
            list_ws.cell(1, 6, "Managers")
            list_ws.cell(1, 7, "Training Types")
            quarters_list = ["All", "Q1", "Q2", "Q3", "Q4"]
            months_list = ["All", "April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February", "March"]
            for i, val in enumerate(["All", fy_label], 2):
                list_ws.cell(i, 1, val)
            for i, val in enumerate(quarters_list, 2):
                list_ws.cell(i, 2, val)
            for i, val in enumerate(months_list, 2):
                list_ws.cell(i, 3, val)
            for i, val in enumerate(["All"] + dept_names, 2):
                list_ws.cell(i, 4, val)
            for i, val in enumerate(["All"] + cat_names, 2):
                list_ws.cell(i, 5, val)
            for i, val in enumerate(["All"] + mgr_names, 2):
                list_ws.cell(i, 6, val)
            for i, val in enumerate(["All"] + training_types, 2):
                list_ws.cell(i, 7, val)
                
            # Populate Raw_Attendance_Data
            raw_row_count = len(all_attendance_results) + 1
            if raw_row_count < 2:
                raw_row_count = 2
            raw_headers = ["Matches_Filter", "Employee Code", "Employee Name", "Department", "Manager", "Training Date", "Financial Year", "Quarter", "Month Name", "Training Title", "Training Category", "Training Type", "Attendance Status", "Actual Hours", "Max Hours", "Is_Unique_Training", "Is_Attended_Present", "Is_Completed", "Actual Hours Credit"]
            for c_idx, h in enumerate(raw_headers, 1):
                raw_ws.cell(1, c_idx, h)
            for idx, (record, training, employee, enrollment) in enumerate(all_attendance_results):
                r = idx + 2
                t_date = training.start_date
                fy_str = ""
                q_str = ""
                m_str = ""
                if t_date:
                    y = t_date.year
                    m_val = t_date.month
                    start_yr = y if m_val >= 4 else y - 1
                    fy_str = f"{start_yr}-{start_yr + 1}"
                    q_str = "Q1" if m_val in (4,5,6) else "Q2" if m_val in (7,8,9) else "Q3" if m_val in (10,11,12) else "Q4"
                    m_str = t_date.strftime("%B")
                raw_ws.cell(r, 1, f"=IF(AND(OR('Department KPI Analytics'!$B$6=\"All\", G{r}='Department KPI Analytics'!$B$6), OR('Department KPI Analytics'!$D$6=\"All\", H{r}='Department KPI Analytics'!$D$6), OR('Department KPI Analytics'!$E$6=\"All\", I{r}='Department KPI Analytics'!$E$6), OR('Department KPI Analytics'!$G$6=\"All\", D{r}='Department KPI Analytics'!$G$6), OR('Department KPI Analytics'!$J$6=\"All\", K{r}='Department KPI Analytics'!$J$6)), 1, 0)")
                raw_ws.cell(r, 2, employee.employee_code)
                raw_ws.cell(r, 3, f"{employee.first_name} {employee.last_name}")
                raw_ws.cell(r, 4, employee.department.name if employee.department else "Unassigned")
                raw_ws.cell(r, 5, f"{employee.manager.first_name} {employee.manager.last_name}" if employee.manager else "")
                raw_ws.cell(r, 6, t_date.strftime("%Y-%m-%d") if t_date else "")
                raw_ws.cell(r, 7, fy_str)
                raw_ws.cell(r, 8, q_str)
                raw_ws.cell(r, 9, m_str)
                raw_ws.cell(r, 10, training.title)
                raw_ws.cell(r, 11, training.category.name if training.category else "")
                raw_ws.cell(r, 12, training.training_type.value if hasattr(training.training_type, "value") else str(training.training_type))
                raw_ws.cell(r, 13, record.status.value if hasattr(record.status, "value") else str(record.status))
                raw_ws.cell(r, 14, float(training.duration_hours or 0))
                raw_ws.cell(r, 15, float(training.max_hours_allowed or training.duration_hours or 0))
                raw_ws.cell(r, 16, f"=IF(AND(A{r}=1, COUNTIFS($J$2:J{r}, J{r}, $A$2:A{r}, 1)=1), 1, 0)")
                raw_ws.cell(r, 17, f"=IF(AND(A{r}=1, OR(M{r}=\"PRESENT\", M{r}=\"LATE\")), 1, 0)")
                raw_ws.cell(r, 18, 1 if enrollment.status == EnrollmentStatus.COMPLETED else 0)
                raw_ws.cell(r, 19, f"=IF(AND(Q{r}=1, R{r}=1), N{r}, 0)")
                
            # Populate Employee_KPI_Summary
            emp_headers = ["S.No", "Employee Code", "Employee Name", "Department", "Manager", "Target Hours", "Actual Hours", "Remaining Hours", "Achievement %", "Attended Count", "Total Assigned", "Attendance %"]
            for c_idx, h in enumerate(emp_headers, 1):
                emp_summary_ws.cell(1, c_idx, h)
            emp_summary_ws.merge_cells("A1:L1")
            emp_summary_ws.cell(1, 1, "Detailed Employee KPI Metrics & Compliance Status").font = Font(name="Calibri", size=12, bold=True)
            emp_summary_ws.cell(2, 1, "Do not edit this sheet directly. It contains formulas linking Employee metrics to the Dashboard filters.")
            emp_start_row = 6
            emp_row_count = len(all_employees) + emp_start_row - 1
            for idx, emp in enumerate(all_employees):
                r = emp_start_row + idx
                emp_summary_ws.cell(r, 1, idx + 1)
                emp_summary_ws.cell(r, 2, emp.employee_code)
                emp_summary_ws.cell(r, 3, f"{emp.first_name} {emp.last_name}")
                emp_summary_ws.cell(r, 4, emp.department.name if emp.department else "Unassigned")
                emp_summary_ws.cell(r, 5, f"{emp.manager.first_name} {emp.manager.last_name}" if emp.manager else "")
                emp_summary_ws.cell(r, 6, f"=IF(OR('Department KPI Analytics'!$G$6=\"All\", D{r}='Department KPI Analytics'!$G$6), 16, 0)")
                emp_summary_ws.cell(r, 7, f"=SUMIFS(Raw_Attendance_Data!$S$2:$S${raw_row_count}, Raw_Attendance_Data!$B$2:$B${raw_row_count}, B{r}, Raw_Attendance_Data!$A$2:$A${raw_row_count}, 1)")
                emp_summary_ws.cell(r, 8, f"=MAX(F{r}-G{r}, 0)")
                emp_summary_ws.cell(r, 9, f"=IF(F{r}>0, G{r}/F{r}, 0)")
                emp_summary_ws.cell(r, 10, f"=COUNTIFS(Raw_Attendance_Data!$B$2:$B${raw_row_count}, B{r}, Raw_Attendance_Data!$A$2:$A${raw_row_count}, 1, Raw_Attendance_Data!$Q$2:$Q${raw_row_count}, 1)")
                emp_summary_ws.cell(r, 11, f"=COUNTIFS(Raw_Attendance_Data!$B$2:$B${raw_row_count}, B{r}, Raw_Attendance_Data!$A$2:$A${raw_row_count}, 1)")
                emp_summary_ws.cell(r, 12, f"=IF(K{r}>0, J{r}/K{r}, 1.0)")
                
            # Populate Department KPI Analytics
            dash_blue = "1F4E79"
            dash_light_blue = "D9EAF7"
            dash_light_gray = "F2F4F7"
            dash_white = "FFFFFF"
            dash_yellow = "FFF2CC"
            dash_font = "Segoe UI"
            thin_side = Side(style="thin", color="D9D9D9")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            card_border = Border(left=Side(style="thin", color="B0C4DE"), right=Side(style="thin", color="B0C4DE"), top=Side(style="thin", color="B0C4DE"), bottom=Side(style="thin", color="B0C4DE"))
            
            def create_filter_field(label_range, value_range, label_text, default_val):
                dept_ws.merge_cells(label_range)
                dept_ws.merge_cells(value_range)
                l_cell = dept_ws[label_range.split(":")[0]]
                v_cell = dept_ws[value_range.split(":")[0]]
                l_cell.value = label_text
                l_cell.font = Font(name=dash_font, size=9, bold=True, color="5A5A5A")
                l_cell.alignment = Alignment(horizontal="center", vertical="center")
                l_cell.fill = PatternFill("solid", fgColor="E9ECEF")
                v_cell.value = default_val
                v_cell.font = Font(name=dash_font, size=10, bold=True, color=dash_blue)
                v_cell.alignment = Alignment(horizontal="center", vertical="center")
                v_cell.fill = PatternFill("solid", fgColor=dash_yellow)
                from openpyxl.utils.cell import range_boundaries
                for r_str in [label_range, value_range]:
                    min_col, min_row, max_col, max_row = range_boundaries(r_str)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            dept_ws.cell(row, col).border = thin_border
                            
            def create_kpi_card(label_range, value_range, label_text, formula_val, num_format=None):
                dept_ws.merge_cells(label_range)
                dept_ws.merge_cells(value_range)
                l_cell = dept_ws[label_range.split(":")[0]]
                v_cell = dept_ws[value_range.split(":")[0]]
                l_cell.value = label_text
                l_cell.font = Font(name=dash_font, size=8, bold=True, color="5A5A5A")
                l_cell.alignment = Alignment(horizontal="center", vertical="center")
                l_cell.fill = PatternFill("solid", fgColor=dash_light_gray)
                v_cell.value = formula_val
                v_cell.font = Font(name=dash_font, size=13, bold=True, color=dash_blue)
                v_cell.alignment = Alignment(horizontal="center", vertical="center")
                v_cell.fill = PatternFill("solid", fgColor=dash_light_gray)
                if num_format:
                    v_cell.number_format = num_format
                from openpyxl.utils.cell import range_boundaries
                for r_str in [label_range, value_range]:
                    min_col, min_row, max_col, max_row = range_boundaries(r_str)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            dept_ws.cell(row, col).border = card_border
                            dept_ws.cell(row, col).fill = PatternFill("solid", fgColor=dash_light_gray)
                            
            dept_ws.merge_cells("A1:L1")
            dept_ws["A1"] = "ORGANIZATION LEARNING & KPI ANALYTICS DASHBOARD"
            dept_ws["A1"].font = Font(name=dash_font, size=16, bold=True, color=dash_white)
            dept_ws["A1"].fill = PatternFill("solid", fgColor=dash_blue)
            dept_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            dept_ws.row_dimensions[1].height = 36
            
            dept_ws.merge_cells("A2:L2")
            dept_ws["A2"] = "Executive Overview & Department Contribution Analysis"
            dept_ws["A2"].font = Font(name=dash_font, size=10, italic=True, color="D9EAF7")
            dept_ws["A2"].fill = PatternFill("solid", fgColor=dash_blue)
            dept_ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
            dept_ws.row_dimensions[2].height = 24
            
            dept_ws.merge_cells("A4:L4")
            dept_ws["A4"] = "Interactive Filters (Select values to dynamically update the dashboard)"
            dept_ws["A4"].font = Font(name=dash_font, size=10, bold=True, color=dash_blue)
            dept_ws["A4"].fill = PatternFill("solid", fgColor=dash_light_blue)
            dept_ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
            dept_ws.row_dimensions[4].height = 24
            
            create_filter_field("B5:C5", "B6:C6", "Financial Year", "2025-2026")
            create_filter_field("D5", "D6", "Quarter", "All")
            create_filter_field("E5:F5", "E6:F6", "Month", "All")
            create_filter_field("G5:I5", "G6:I6", "Department", "All")
            create_filter_field("J5:L5", "J6:L6", "Training Category", "All")
            dept_ws.row_dimensions[5].height = 20
            dept_ws.row_dimensions[6].height = 24
            
            dv_fy = DataValidation(type="list", formula1="Filter_Lists!$A$2:$A$20", allow_blank=True)
            dv_q = DataValidation(type="list", formula1="Filter_Lists!$B$2:$B$6", allow_blank=True)
            dv_m = DataValidation(type="list", formula1="Filter_Lists!$C$2:$C$14", allow_blank=True)
            dv_dept = DataValidation(type="list", formula1=f"Filter_Lists!$D$2:$D${len(dept_names)+2}", allow_blank=True)
            dv_cat = DataValidation(type="list", formula1=f"Filter_Lists!$E$2:$E${len(cat_names)+2}", allow_blank=True)
            
            dept_ws.add_data_validation(dv_fy)
            dept_ws.add_data_validation(dv_q)
            dept_ws.add_data_validation(dv_m)
            dept_ws.add_data_validation(dv_dept)
            dept_ws.add_data_validation(dv_cat)
            
            dv_fy.add(dept_ws["B6"])
            dv_q.add(dept_ws["D6"])
            dv_m.add(dept_ws["E6"])
            dv_dept.add(dept_ws["G6"])
            dv_cat.add(dept_ws["J6"])
            
            create_kpi_card("A10:B10", "A11:B11", "TOTAL EMPLOYEES", f"=COUNTIFS(Employee_KPI_Summary!$F$6:$F${emp_row_count}, \">0\")", "#,##0")
            create_kpi_card("C10:D10", "C11:D11", "TOTAL TARGET HOURS", f"=SUM(Employee_KPI_Summary!$F$6:$F${emp_row_count})", "#,##0.0")
            create_kpi_card("E10:F10", "E11:F11", "TOTAL ACTUAL HOURS", f"=SUM(Employee_KPI_Summary!$G$6:$G${emp_row_count})", "#,##0.0")
            create_kpi_card("G10:H10", "G11:H11", "TOTAL REMAINING HOURS", f"=SUM(Employee_KPI_Summary!$H$6:$H${emp_row_count})", "#,##0.0")
            create_kpi_card("I10:L10", "I11:L11", "OVERALL ACHIEVEMENT %", "=IF(C11>0, E11/C11, 0)", "0.0%")
            create_kpi_card("A12:B12", "A13:B13", "LEARNING COMPLIANCE %", f"=IF(COUNTIFS(Employee_KPI_Summary!$F$6:$F${emp_row_count}, \">0\")>0, COUNTIFS(Employee_KPI_Summary!$F$6:$F${emp_row_count}, \">0\", Employee_KPI_Summary!$I$6:$I${emp_row_count}, \">=1.0\")/COUNTIFS(Employee_KPI_Summary!$F$6:$F${emp_row_count}, \">0\"), 1.0)", "0.0%")
            create_kpi_card("C12:D12", "C13:D13", "TRAININGS CONDUCTED", f"=SUM(Raw_Attendance_Data!$P$2:$P${raw_row_count})", "#,##0")
            create_kpi_card("E12:F12", "E13:F13", "TOTAL ATTENDANCE %", f"=IF(SUM(Employee_KPI_Summary!$K$6:$K${emp_row_count})>0, SUM(Employee_KPI_Summary!$J$6:$J${emp_row_count})/SUM(Employee_KPI_Summary!$K$6:$K${emp_row_count}), 1.0)", "0.0%")
            
            depts_to_write = dept_names if dept_names else ["Unassigned"]
            N_DEPTS = len(depts_to_write)
            depts_start_row = 17
            depts_end_row = depts_start_row + N_DEPTS - 1
            totals_row = depts_end_row + 1
            
            create_kpi_card("G12:H12", "G13:H13", "TOP PERFORMING DEPT", f"=INDEX(B17:B{depts_end_row}, MATCH(MAX(G17:G{depts_end_row}), G17:G{depts_end_row}, 0))")
            create_kpi_card("I12:L12", "I13:L13", "LOWEST PERFORMING DEPT", f"=INDEX(B17:B{depts_end_row}, MATCH(MIN(K17:K{depts_end_row}), K17:K{depts_end_row}, 0))")
            
            dept_ws.merge_cells("A15:K15")
            dept_ws["A15"] = "DEPARTMENT CONTRIBUTION & KPI SUMMARY"
            dept_ws["A15"].font = Font(name=dash_font, size=11, bold=True, color=dash_white)
            dept_ws["A15"].fill = PatternFill("solid", fgColor=dash_blue)
            dept_ws["A15"].alignment = Alignment(horizontal="center", vertical="center")
            dept_ws.row_dimensions[15].height = 24
            
            headers = ["S.No", "Department", "Employees", "Target Hours", "Actual Hours", "Remaining Hours", "Achievement %", "Attendance %", "KPI Status", "Rank Score", "Filtered Achievement"]
            for c_idx, h in enumerate(headers, 1):
                cell = dept_ws.cell(16, c_idx, h)
                cell.font = Font(name=dash_font, size=9, bold=True, color=dash_blue)
                cell.fill = PatternFill("solid", fgColor=dash_light_blue)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            dept_ws.row_dimensions[16].height = 20
            
            for idx, dept in enumerate(depts_to_write):
                r = depts_start_row + idx
                dept_ws.cell(r, 1, f"=ROW() - 16")
                dept_ws.cell(r, 2, dept)
                dept_ws.cell(r, 3, f"=COUNTIFS(Employee_KPI_Summary!$D$6:$D${emp_row_count}, B{r}, Employee_KPI_Summary!$F$6:$F${emp_row_count}, \">0\")")
                dept_ws.cell(r, 4, f"=SUMIFS(Employee_KPI_Summary!$F$6:$F${emp_row_count}, Employee_KPI_Summary!$D$6:$D${emp_row_count}, B{r})")
                dept_ws.cell(r, 5, f"=SUMIFS(Employee_KPI_Summary!$G$6:$G${emp_row_count}, Employee_KPI_Summary!$D$6:$D${emp_row_count}, B{r})")
                dept_ws.cell(r, 6, f"=SUMIFS(Employee_KPI_Summary!$H$6:$H${emp_row_count}, Employee_KPI_Summary!$D$6:$D${emp_row_count}, B{r})")
                dept_ws.cell(r, 7, f"=IF(D{r}>0, E{r}/D{r}, 0)")
                dept_ws.cell(r, 8, f"=IF(SUMIFS(Employee_KPI_Summary!$K$6:$K${emp_row_count}, Employee_KPI_Summary!$D$6:$D${emp_row_count}, B{r})>0, SUMIFS(Employee_KPI_Summary!$J$6:$J${emp_row_count}, Employee_KPI_Summary!$D$6:$D${emp_row_count}, B{r})/SUMIFS(Employee_KPI_Summary!$K$6:$K${emp_row_count}, Employee_KPI_Summary!$D$6:$D${emp_row_count}, B{r}), 1.0)")
                dept_ws.cell(r, 9, f"=IF(G{r}>=1.0, \"Achieved\", IF(G{r}>=0.75, \"Good\", IF(G{r}>=0.5, \"Moderate\", IF(G{r}>=0.25, \"Low\", \"Critical\"))))")
                dept_ws.cell(r, 10, f"=G{r}+(1000-ROW())/1000000")
                dept_ws.cell(r, 11, f"=IF(C{r}>0, G{r}+ROW()/1000000, 9.99)")
                for c_idx in range(1, 12):
                    cell = dept_ws.cell(r, c_idx)
                    cell.font = Font(name=dash_font, size=9, color="1F2937")
                    cell.border = thin_border
                    if c_idx == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif c_idx == 9:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    if c_idx in (3, 10, 11):
                        cell.number_format = "#,##0"
                    elif c_idx in (4, 5, 6):
                        cell.number_format = "#,##0.0"
                    elif c_idx in (7, 8):
                        cell.number_format = "0.0%"
                dept_ws.row_dimensions[r].height = 20
                
            dept_ws.cell(totals_row, 2, "Total")
            dept_ws.cell(totals_row, 3, f"=SUM(C17:C{depts_end_row})")
            dept_ws.cell(totals_row, 4, f"=SUM(D17:D{depts_end_row})")
            dept_ws.cell(totals_row, 5, f"=SUM(E17:E{depts_end_row})")
            dept_ws.cell(totals_row, 6, f"=SUM(F17:F{depts_end_row})")
            dept_ws.cell(totals_row, 7, f"=IF(D{totals_row}>0, E{totals_row}/D{totals_row}, 0)")
            dept_ws.cell(totals_row, 8, f"=IF(SUM(Employee_KPI_Summary!$K$6:$K${emp_row_count})>0, SUM(Employee_KPI_Summary!$J$6:$J${emp_row_count})/SUM(Employee_KPI_Summary!$K$6:$K${emp_row_count}), 1.0)")
            dept_ws.cell(totals_row, 9, f"=IF(G{totals_row}>=1.0, \"Achieved\", IF(G{totals_row}>=0.75, \"Good\", IF(G{totals_row}>=0.5, \"Moderate\", IF(G{totals_row}>=0.25, \"Low\", \"Critical\"))))")
            for c_idx in range(1, 12):
                cell = dept_ws.cell(totals_row, c_idx)
                cell.font = Font(name=dash_font, size=9, bold=True, color=dash_blue)
                cell.fill = PatternFill("solid", fgColor=dash_light_blue)
                cell.border = double_bottom_border
                if c_idx == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 9:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if c_idx == 3:
                    cell.number_format = "#,##0"
                elif c_idx in (4, 5, 6):
                    cell.number_format = "#,##0.0"
                elif c_idx in (7, 8):
                    cell.number_format = "0.0%"
            dept_ws.row_dimensions[totals_row].height = 22
            
            # Rankings
            rank_start_row = totals_row + 3
            dept_ws.merge_cells(f"A{rank_start_row}:E{rank_start_row}")
            dept_ws[f"A{rank_start_row}"] = "Top Performing Departments"
            dept_ws[f"A{rank_start_row}"].font = Font(name=dash_font, size=10, bold=True, color=dash_white)
            dept_ws[f"A{rank_start_row}"].fill = PatternFill("solid", fgColor="10B981")
            dept_ws[f"A{rank_start_row}"].alignment = Alignment(horizontal="center", vertical="center")
            
            dept_ws.merge_cells(f"G{rank_start_row}:J{rank_start_row}")
            dept_ws[f"G{rank_start_row}"] = "Departments Requiring Attention"
            dept_ws[f"G{rank_start_row}"].font = Font(name=dash_font, size=10, bold=True, color=dash_white)
            dept_ws[f"G{rank_start_row}"].fill = PatternFill("solid", fgColor="EF4444")
            dept_ws[f"G{rank_start_row}"].alignment = Alignment(horizontal="center", vertical="center")
            dept_ws.row_dimensions[rank_start_row].height = 22
            
            top_headers = ["Rank", "Department", "Achievement %", "Actual Hours", "Attendance %"]
            for c_idx, h in enumerate(top_headers, 1):
                cell = dept_ws.cell(rank_start_row + 1, c_idx, h)
                cell.font = Font(name=dash_font, size=9, bold=True, color="5A5A5A")
                cell.fill = PatternFill("solid", fgColor="F2F4F7")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            bot_headers = ["Department", "Remaining Hours", "Compliance %", "Risk Indicator"]
            for c_idx, h in enumerate(bot_headers, 7):
                cell = dept_ws.cell(rank_start_row + 1, c_idx, h)
                cell.font = Font(name=dash_font, size=9, bold=True, color="5A5A5A")
                cell.fill = PatternFill("solid", fgColor="F2F4F7")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            dept_ws.row_dimensions[rank_start_row + 1].height = 20
            
            num_rank_rows = min(5, N_DEPTS)
            for i in range(1, num_rank_rows + 1):
                r_idx = rank_start_row + 1 + i
                dept_ws.cell(r_idx, 1, i)
                dept_ws.cell(r_idx, 2, f"=INDEX($B$17:$B${depts_end_row}, MATCH(LARGE($J$17:$J${depts_end_row}, A{r_idx}), $J$17:$J${depts_end_row}, 0))")
                dept_ws.cell(r_idx, 3, f"=INDEX($G$17:$G${depts_end_row}, MATCH(LARGE($J$17:$J${depts_end_row}, A{r_idx}), $J$17:$J${depts_end_row}, 0))")
                dept_ws.cell(r_idx, 4, f"=INDEX($E$17:$E${depts_end_row}, MATCH(LARGE($J$17:$J${depts_end_row}, A{r_idx}), $J$17:$J${depts_end_row}, 0))")
                dept_ws.cell(r_idx, 5, f"=INDEX($H$17:$H${depts_end_row}, MATCH(LARGE($J$17:$J${depts_end_row}, A{r_idx}), $J$17:$J${depts_end_row}, 0))")
                
                dept_ws.cell(r_idx, 7, f"=INDEX($B$17:$B${depts_end_row}, MATCH(SMALL($K$17:$K${depts_end_row}, {i}), $K$17:$K${depts_end_row}, 0))")
                dept_ws.cell(r_idx, 8, f"=INDEX($F$17:$F${depts_end_row}, MATCH(SMALL($K$17:$K${depts_end_row}, {i}), $K$17:$K${depts_end_row}, 0))")
                dept_ws.cell(r_idx, 9, f"=INDEX($G$17:$G${depts_end_row}, MATCH(SMALL($K$17:$K${depts_end_row}, {i}), $K$17:$K${depts_end_row}, 0))")
                dept_ws.cell(r_idx, 10, f"=IF(I{r_idx}<0.5, \"🔴 Critical Risk\", IF(I{r_idx}<0.75, \"🟡 Moderate Risk\", \"🟢 Low Risk\"))")
                for c in range(1, 6):
                    cell = dept_ws.cell(r_idx, c)
                    cell.font = Font(name=dash_font, size=9, color="1F2937")
                    cell.border = thin_border
                    if c == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif c == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    if c in (3, 5):
                        cell.number_format = "0.0%"
                    elif c == 4:
                        cell.number_format = "#,##0.0"
                for c in range(7, 11):
                    cell = dept_ws.cell(r_idx, c)
                    cell.font = Font(name=dash_font, size=9, color="1F2937")
                    cell.border = thin_border
                    if c == 7:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif c == 10:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    if c == 8:
                        cell.number_format = "#,##0.0"
                    elif c == 9:
                        cell.number_format = "0.0%"
                dept_ws.row_dimensions[r_idx].height = 20
                
            # Monthly run rate table + Insights
            rank_end_row = rank_start_row + 1 + num_rank_rows
            insights_start_row = rank_end_row + 3
            dept_ws.merge_cells(f"A{insights_start_row}:E{insights_start_row}")
            dept_ws[f"A{insights_start_row}"] = "Monthly Learning Trends (FY)"
            dept_ws[f"A{insights_start_row}"].font = Font(name=dash_font, size=10, bold=True, color=dash_white)
            dept_ws[f"A{insights_start_row}"].fill = PatternFill("solid", fgColor="8B5CF6")
            dept_ws[f"A{insights_start_row}"].alignment = Alignment(horizontal="center", vertical="center")
            
            dept_ws.merge_cells(f"G{insights_start_row}:K{insights_start_row}")
            dept_ws[f"G{insights_start_row}"] = "Executive KPI Insights"
            dept_ws[f"G{insights_start_row}"].font = Font(name=dash_font, size=10, bold=True, color=dash_white)
            dept_ws[f"G{insights_start_row}"].fill = PatternFill("solid", fgColor="1F4E79")
            dept_ws[f"G{insights_start_row}"].alignment = Alignment(horizontal="center", vertical="center")
            dept_ws.row_dimensions[insights_start_row].height = 22
            
            monthly_headers = ["Month", "Target Hours", "Actual Hours", "Achievement %"]
            for c_idx, h in enumerate(monthly_headers, 1):
                cell = dept_ws.cell(insights_start_row + 1, c_idx, h)
                cell.font = Font(name=dash_font, size=9, bold=True, color="5A5A5A")
                cell.fill = PatternFill("solid", fgColor="F2F4F7")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            dept_ws.row_dimensions[insights_start_row + 1].height = 20
            
            months_names = ["April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February", "March"]
            for idx, month_name in enumerate(months_names):
                curr_row = insights_start_row + 2 + idx
                dept_ws.cell(curr_row, 1, month_name)
                dept_ws.cell(curr_row, 2, "=$A$11 * 1.33")
                dept_ws.cell(curr_row, 3, f"=SUMIFS(Raw_Attendance_Data!$S$2:$S${raw_row_count}, Raw_Attendance_Data!$I$2:$I${raw_row_count}, A{curr_row}, Raw_Attendance_Data!$A$2:$A${raw_row_count}, 1)")
                dept_ws.cell(curr_row, 4, f"=IF(B{curr_row}>0, C{curr_row}/B{curr_row}, 0)")
                for c in range(1, 5):
                    cell = dept_ws.cell(curr_row, c)
                    cell.font = Font(name=dash_font, size=9, color="1F2937")
                    cell.border = thin_border
                    if c == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    if c in (2, 3):
                        cell.number_format = "#,##0.0"
                    elif c == 4:
                        cell.number_format = "0.0%"
                dept_ws.row_dimensions[curr_row].height = 18
                
            insights_formulas = [
                f"=\"• \" & INDEX(B17:B{depts_end_row}, MATCH(MAX(E17:E{depts_end_row}), E17:E{depts_end_row}, 0)) & \" contributed the highest learning hours (\" & TEXT(MAX(E17:E{depts_end_row}), \"#,##0\") & \" hours).\"",
                f"=\"• \" & INDEX(B17:B{depts_end_row}, MATCH(MIN(K17:K{depts_end_row}), K17:K{depts_end_row}, 0)) & \" is \" & TEXT(MAX(0, 1-MIN(K17:K{depts_end_row})), \"0%\") & \" below annual learning target.\"",
                f"=\"• \" & INDEX(B17:B{depts_end_row}, MATCH(MAX(G17:G{depts_end_row}), G17:G{depts_end_row}, 0)) & \" achieved \" & TEXT(MAX(G17:G{depts_end_row}), \"0.0%\") & \" learning compliance.\"",
                "=\"• Organization-wide learning compliance stands at \" & TEXT(A13, \"0.0%\") & \".\""
            ]
            for idx, formula in enumerate(insights_formulas):
                curr_row = insights_start_row + 2 + idx
                dept_ws.merge_cells(start_row=curr_row, start_column=7, end_row=curr_row, end_column=11)
                cell = dept_ws.cell(curr_row, 7, formula)
                cell.font = Font(name=dash_font, size=9.5, italic=True, color="1F4E79")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                for col in range(7, 12):
                    c_cell = dept_ws.cell(curr_row, col)
                    c_cell.fill = PatternFill("solid", fgColor="F8F9FA")
                    c_cell.border = thin_border
                dept_ws.row_dimensions[curr_row].height = 24
                
            column_widths = {"A": 8, "B": 24, "C": 14, "D": 16, "E": 16, "F": 16, "G": 18, "H": 18, "I": 16, "J": 14, "K": 18, "L": 4}
            for col_letter, w in column_widths.items():
                dept_ws.column_dimensions[col_letter].width = w
                
            # Embed Charts if enabled
            if include_charts:
                chart1 = PieChart()
                chart1.title = "Organization Learning Contribution by Department"
                chart1.style = 10
                chart1.width = 16
                chart1.height = 10
                data1 = Reference(dept_ws, min_col=5, min_row=16, max_row=depts_end_row)
                cats1 = Reference(dept_ws, min_col=2, min_row=17, max_row=depts_end_row)
                chart1.add_data(data1, titles_from_data=True)
                chart1.set_categories(cats1)
                dept_ws.add_chart(chart1, "N15")
                
                chart2 = BarChart()
                chart2.type = "col"
                chart2.style = 11
                chart2.title = "Target Hours vs Actual Hours by Department"
                chart2.y_axis.title = "Hours"
                chart2.x_axis.title = "Department"
                chart2.width = 16
                chart2.height = 10
                data2 = Reference(dept_ws, min_col=4, max_col=5, min_row=16, max_row=depts_end_row)
                cats2 = Reference(dept_ws, min_col=2, min_row=17, max_row=depts_end_row)
                chart2.add_data(data2, titles_from_data=True)
                chart2.set_categories(cats2)
                dept_ws.add_chart(chart2, "N30")
                
                chart3 = BarChart()
                chart3.type = "col"
                chart3.style = 12
                chart3.title = "Department Compliance %"
                chart3.y_axis.title = "Compliance %"
                chart3.x_axis.title = "Department"
                chart3.width = 16
                chart3.height = 10
                data3 = Reference(dept_ws, min_col=3, min_row=rank_start_row + 1, max_row=rank_start_row + 1 + num_rank_rows)
                cats3 = Reference(dept_ws, min_col=2, min_row=rank_start_row + 2, max_row=rank_start_row + 1 + num_rank_rows)
                chart3.add_data(data3, titles_from_data=True)
                chart3.set_categories(cats3)
                dept_ws.add_chart(chart3, "N45")
                
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        with open(filename, "wb") as f:
            f.write(output.getvalue())
        print("Done.")

async def run_all_stages():
    await generate_stage(1, False, "stage1.xlsx")
    await generate_stage(2, False, "stage2.xlsx")
    await generate_stage(3, False, "stage3.xlsx")
    await generate_stage(4, False, "stage4_no_charts.xlsx")
    await generate_stage(4, True, "stage4_with_charts.xlsx")

if __name__ == "__main__":
    asyncio.run(run_all_stages())
