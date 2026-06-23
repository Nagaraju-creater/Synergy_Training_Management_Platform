from typing import List, Optional, Tuple, Any
from uuid import UUID
from datetime import date, datetime, time, timedelta
from sqlalchemy import select, or_, and_, update, func
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import HTTPException
import io
import re

from app.trainings.models import Training, TrainingType, TrainingStatus, DeliveryMode, TrainingImportHistory
from app.trainings.categories import TrainingCategory
from app.trainings.documents import TrainingDocument
from app.trainings.schemas import TrainingCreate, TrainingUpdate, TrainingCategoryCreate
from app.utils.exceptions import NotFoundException, BadRequestException
from app.utils.pagination import paginate

def parse_excel_date(val: Any) -> date:
    if not val:
        raise ValueError("Date is empty.")
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) else val.date()
    val_str = str(val).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date format: '{val_str}'. Expected DD-MM-YYYY.")

def parse_excel_time(val: Any) -> str:
    if not val:
        raise ValueError("Time is empty.")
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    if isinstance(val, time):
        return val.strftime("%H:%M")
    from datetime import time as dt_time
    if isinstance(val, dt_time):
        return val.strftime("%H:%M")
    
    val_str = str(val).strip()
    val_str = re.sub(r'\s+', ' ', val_str)
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p", "%I:%M%p", "%I:%M:%S%p"):
        try:
            t = datetime.strptime(val_str, fmt).time()
            return t.strftime("%H:%M")
        except ValueError:
            continue
    raise ValueError(f"Invalid time format: '{val_str}'. Expected HH:MM.")

def format_time_for_db(time_str: str) -> str:
    t = datetime.strptime(time_str, "%H:%M").time()
    return t.strftime("%I:%M %p")


def _compute_status(training: Training) -> TrainingStatus:
    """Auto-compute training status based on real-time datetime.
    Only transitions non-DRAFT, non-CANCELLED trainings."""
    if training.status in (TrainingStatus.DRAFT, TrainingStatus.CANCELLED):
        return training.status
    
    from datetime import datetime, time, timedelta
    now = datetime.now()
    
    # Combine start date and time
    start_dt = None
    if training.start_date:
        try:
            t = datetime.strptime(training.start_time, "%I:%M %p").time() if training.start_time else time(0, 0)
            start_dt = datetime.combine(training.start_date, t)
        except (ValueError, TypeError):
            start_dt = datetime.combine(training.start_date, time(0, 0))
    
    # Combine end date and time.
    # duration_hours is the source of truth for when a session actually ends.
    # end_date alone is not enough — hardcoding 23:59:59 means a 30-min class
    # that starts at 4:30 PM wouldn't show COMPLETED until midnight.
    end_dt = None
    if start_dt:
        duration = training.duration_hours or 2.0
        end_dt = start_dt + timedelta(hours=duration)
        # For multi-day trainings (end_date is a later calendar date), extend
        # to end of that day so the training is never prematurely completed.
        if training.end_date and training.start_date and training.end_date > training.start_date:
            end_of_end_date = datetime.combine(training.end_date, time(23, 59, 59))
            if end_of_end_date > end_dt:
                end_dt = end_of_end_date
    elif training.end_date:
        # Fallback: no start_dt at all — use end of end_date
        end_dt = datetime.combine(training.end_date, time(23, 59, 59))
    
    if start_dt and end_dt:
        if now < start_dt:
            return TrainingStatus.SCHEDULED
        elif start_dt <= now <= end_dt:
            return TrainingStatus.ONGOING
        else:
            return TrainingStatus.COMPLETED
    elif start_dt:
        if now < start_dt:
            return TrainingStatus.SCHEDULED
        else:
            return TrainingStatus.ONGOING
            
    return training.status


class TrainingService:
    @staticmethod
    async def get_all(
        db: AsyncSession, 
        page: int = 1, 
        per_page: int = 20,
        category_id: Optional[UUID] = None,
        department_id: Optional[UUID] = None,
        status: Optional[str] = None,
        training_type: Optional[str] = None,
        search: Optional[str] = None,
        include_archived: bool = False,
        is_admin: bool = False,
    ) -> Tuple[List[Training], int, dict]:
        # FIRST: sync all active statuses (non-DRAFT, non-CANCELLED) based on date/time logic
        sync_stmt = select(Training).where(Training.status.notin_([TrainingStatus.DRAFT, TrainingStatus.CANCELLED]))
        sync_res = await db.execute(sync_stmt)
        all_active_trainings = sync_res.scalars().all()
        updated_any = False
        newly_completed: list = []
        for t in all_active_trainings:
            new_status = _compute_status(t)
            if new_status != t.status:
                if new_status == TrainingStatus.COMPLETED:
                    newly_completed.append(t)
                t.status = new_status
                updated_any = True
        if updated_any:
            await db.flush()
        # Trigger effectiveness assignment for newly completed trainings
        if newly_completed:
            from app.effectiveness.service import EffectivenessService
            for t in newly_completed:
                try:
                    await EffectivenessService.assign_training_effectiveness(db, t)
                except Exception:
                    pass

        # SECOND: Calculate status counts based on filters, excluding status filter itself
        count_filters = []
        if not include_archived:
            count_filters.append(Training.is_archived == False)
        if not is_admin:
            count_filters.append(Training.status != TrainingStatus.DRAFT)
        if category_id:
            count_filters.append(Training.category_id == category_id)
        if department_id:
            count_filters.append(
                or_(
                    Training.is_global == True,
                    Training.departments.any(id=department_id)
                )
            )
        if training_type and training_type.strip():
            count_filters.append(Training.training_type == training_type.upper())
        if search and search.strip():
            count_filters.append(Training.title.ilike(f"%{search}%"))

        count_stmt = select(Training.status, func.count(Training.id)).group_by(Training.status)
        if count_filters:
            count_stmt = count_stmt.where(and_(*count_filters))
        
        count_res = await db.execute(count_stmt)
        counts_dict = {
            (status_val.upper() if isinstance(status_val, str) else status_val.value.upper()): count_val 
            for status_val, count_val in count_res.all() if status_val
        }
        
        status_counts = {
            "scheduled": counts_dict.get("SCHEDULED", 0),
            "ongoing": counts_dict.get("ONGOING", 0),
            "completed": counts_dict.get("COMPLETED", 0),
            "all": sum(counts_dict.values())
        }

        # THIRD: Fetch the actual page of data
        stmt = select(Training).options(
            selectinload(Training.category),
            selectinload(Training.departments),
            selectinload(Training.documents)
        )
        
        filters = []
        if not include_archived:
            filters.append(Training.is_archived == False)
        
        # Non-admins never see DRAFT trainings
        if not is_admin:
            filters.append(Training.status != TrainingStatus.DRAFT)
        
        if category_id:
            filters.append(Training.category_id == category_id)
        if department_id:
            filters.append(
                or_(
                    Training.is_global == True,
                    Training.departments.any(id=department_id)
                )
            )
        if status and status.strip():
            filters.append(Training.status == status.upper())
        if training_type and training_type.strip():
            filters.append(Training.training_type == training_type.upper())
        if search and search.strip():
            filters.append(Training.title.ilike(f"%{search}%"))
            
        if filters:
            stmt = stmt.where(and_(*filters))
            
        from app.utils.pagination import paginate_query
        trainings, total = await paginate_query(db, stmt, page, per_page)
        
        # Populate enrolled_count dynamically
        if trainings:
            from app.enrollments.models import Enrollment, EnrollmentStatus
            training_ids = [t.id for t in trainings]
            count_stmt = select(
                Enrollment.training_id, 
                func.count(Enrollment.id)
            ).where(
                Enrollment.training_id.in_(training_ids),
                Enrollment.status.in_([EnrollmentStatus.APPROVED, EnrollmentStatus.COMPLETED])
            ).group_by(Enrollment.training_id)
            
            res = await db.execute(count_stmt)
            counts = dict(res.all())
            for t in trainings:
                setattr(t, "enrolled_count", counts.get(t.id, 0))
        
        return trainings, total, status_counts

    @staticmethod
    async def get_by_id(db: AsyncSession, training_id: UUID) -> Training:
        stmt = select(Training).options(
            selectinload(Training.category),
            selectinload(Training.departments),
            selectinload(Training.documents)
        ).where(Training.id == training_id)
        
        result = await db.execute(stmt)
        t = result.scalar_one_or_none()
        if not t:
            raise NotFoundException("Training")
            
        # Sync status on fetch
        new_status = _compute_status(t)
        if new_status != t.status:
            old_status = t.status
            t.status = new_status
            await db.flush()
            # Trigger effectiveness assignment if newly completed
            if new_status == TrainingStatus.COMPLETED:
                from app.effectiveness.service import EffectivenessService
                try:
                    await EffectivenessService.assign_training_effectiveness(db, t)
                except Exception:
                    pass
            
        from app.enrollments.models import Enrollment, EnrollmentStatus
        count_stmt = select(func.count(Enrollment.id)).where(
            Enrollment.training_id == training_id,
            Enrollment.status.in_([EnrollmentStatus.APPROVED, EnrollmentStatus.COMPLETED])
        )
        res = await db.execute(count_stmt)
        setattr(t, "enrolled_count", res.scalar() or 0)

        # Attach linked learning module id
        from app.learning_hub.models import LearningModule
        mod_res = await db.execute(
            select(LearningModule.id).where(LearningModule.training_id == training_id)
        )
        setattr(t, "learning_module_id", mod_res.scalar_one_or_none())

        return t

    @staticmethod
    async def validate_training(payload: TrainingCreate):
        if payload.delivery_mode == "online" and not payload.meeting_link:
            raise BadRequestException("Meeting link is required for online trainings")
        if payload.delivery_mode == "in_person" and not payload.venue:
            raise BadRequestException("Venue is required for in-person trainings")

    @staticmethod
    def _combine_deadline(deadline_date: Optional[Any], deadline_time: Optional[str]) -> Optional[datetime]:
        """
        Combine a date and a time string into a full datetime.
        Supports both 24-hour (HH:MM) and 12-hour (HH:MM AM/PM) formats.
        This is critical: the frontend sends enrollment_deadline_time in HH:MM (24h)
        from an <input type="time">, so we must try that format first.
        """
        import logging
        logger = logging.getLogger(__name__)

        if not deadline_date:
            return None

        # If it's already a datetime (e.g. from a combined picker), extract its date
        if isinstance(deadline_date, datetime):
            target_date = deadline_date.date()
            if deadline_time:
                # Explicit time provided — always use it
                target_time_str = deadline_time
            elif deadline_date.hour == 0 and deadline_date.minute == 0 and deadline_date.second == 0:
                # Midnight means Pydantic parsed a date-only string ("YYYY-MM-DD");
                # no meaningful time was set, so fall back to 23:59.
                target_time_str = "23:59"
            else:
                # Non-midnight time in the datetime — use it (e.g. from a datetime picker)
                target_time_str = deadline_date.strftime("%H:%M")
        else:
            target_date = deadline_date
            target_time_str = deadline_time or "23:59"

        # Try 24-hour format first (HH:MM or HH:MM:SS) — used by <input type="time">
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                t = datetime.strptime(target_time_str.strip(), fmt).time()
                combined = datetime.combine(target_date, t)
                logger.info(
                    "[deadline] Parsed '%s' with fmt='%s' -> %s (date=%s)",
                    target_time_str, fmt, combined, target_date
                )
                return combined
            except (ValueError, TypeError):
                pass

        # Try 12-hour AM/PM format — normalize to uppercase first since Python's
        # %p directive is case-sensitive (requires "PM" not "pm").
        # date-fns format(_, "hh:mm a") returns lowercase "pm", so .upper() is critical.
        time_upper = target_time_str.strip().upper()
        for fmt in ("%I:%M %p", "%I:%M:%S %p"):
            try:
                t = datetime.strptime(time_upper, fmt).time()
                combined = datetime.combine(target_date, t)
                logger.info(
                    "[deadline] Parsed '%s' (as '%s') with fmt='%s' -> %s (date=%s)",
                    target_time_str, time_upper, fmt, combined, target_date
                )
                return combined
            except (ValueError, TypeError):
                pass

        # Final fallback — log a warning so the bug is visible
        from datetime import time as dt_time
        logger.warning(
            "[deadline] Could not parse time string '%s'; defaulting to 23:59 for date=%s",
            target_time_str, target_date
        )
        return datetime.combine(target_date, dt_time(23, 59))

    @staticmethod
    async def create(db: AsyncSession, payload: TrainingCreate, created_by: UUID) -> Training:
        try:
            await TrainingService.validate_training(payload)
            
            data = payload.model_dump()
            dept_ids = data.pop("department_ids", [])
            
            # Combine enrollment deadline
            deadline_date = data.get("enrollment_deadline")
            deadline_time = data.get("enrollment_deadline_time")
            if deadline_date:
                data["enrollment_deadline"] = TrainingService._combine_deadline(deadline_date, deadline_time)
            
            # Initialize available seats
            data["available_seats"] = data.get("max_participants", 20)
            
            # Clean up temporary fields
            data.pop("enrollment_deadline_time", None)
            data.pop("end_time", None)
            
            training = Training(**data, created_by=created_by)
            
            if dept_ids:
                from app.departments.models import Department
                result = await db.execute(select(Department).where(Department.id.in_(dept_ids)))
                training.departments = list(result.scalars().all())
            
            db.add(training)
            await db.flush()
            await TrainingService._ensure_learning_module(db, training)
            return await TrainingService.get_by_id(db, training.id)
        except Exception as e:
            if isinstance(e, (BadRequestException, NotFoundException)):
                raise e
            raise BadRequestException(f"Failed to create training: {str(e)}")

    @staticmethod
    async def update(db: AsyncSession, training_id: UUID, payload: TrainingUpdate) -> Training:
        try:
            t = await TrainingService.get_by_id(db, training_id)
            if not t:
                raise NotFoundException("Training not found")
            
            if t.status == TrainingStatus.COMPLETED or _compute_status(t) == TrainingStatus.COMPLETED:
                raise BadRequestException("Completed trainings are locked and cannot be modified.")
            
            data = payload.model_dump(exclude_unset=True)
            dept_ids = data.pop("department_ids", None)
            
            # Handle seat count differences if max_participants changes
            if "max_participants" in data and data["max_participants"] is not None:
                diff = data["max_participants"] - (t.max_participants or 0)
                t.available_seats = (t.available_seats or 0) + diff
            
            # Combine enrollment deadline if date/time provided
            if "enrollment_deadline" in data or "enrollment_deadline_time" in data:
                deadline_date = data.get("enrollment_deadline", t.enrollment_deadline)
                deadline_time = data.get("enrollment_deadline_time")
                data["enrollment_deadline"] = TrainingService._combine_deadline(deadline_date, deadline_time)
            
            # Clean up temporary fields
            data.pop("enrollment_deadline_time", None)
            data.pop("end_time", None)
            
            for key, value in data.items():
                setattr(t, key, value)
            
            if dept_ids is not None:
                from app.departments.models import Department
                result = await db.execute(select(Department).where(Department.id.in_(dept_ids)))
                t.departments = list(result.scalars().all())
            
            await db.flush()
            return await TrainingService.get_by_id(db, t.id)
        except Exception as e:
            if isinstance(e, (BadRequestException, NotFoundException)):
                raise e
            raise BadRequestException(f"Failed to update training: {str(e)}")

    @staticmethod
    async def archive(db: AsyncSession, training_id: UUID) -> Training:
        t = await TrainingService.get_by_id(db, training_id)
        if not t:
            raise NotFoundException("Training not found")
        if t.status == TrainingStatus.COMPLETED or _compute_status(t) == TrainingStatus.COMPLETED:
            raise BadRequestException("Completed trainings are locked and cannot be archived.")
        t.is_archived = True
        await db.flush()
        await db.refresh(t, ["updated_at"])
        return t

    @staticmethod
    async def delete(db: AsyncSession, training_id: UUID) -> None:
        try:
            t = await TrainingService.get_by_id(db, training_id)
            if not t:
                raise NotFoundException("Training not found")
            
            await db.delete(t)
            await db.commit()
        except Exception as e:
            await db.rollback()
            if isinstance(e, (BadRequestException, NotFoundException)):
                raise e
            raise BadRequestException(f"Failed to delete training. It might have active enrollments or attendance data that prevents deletion: {str(e)}")

    @staticmethod
    async def _ensure_learning_module(db: AsyncSession, training) -> None:
        """Auto-create a LearningModule linked to a Training if one doesn't already exist."""
        from app.learning_hub.models import LearningModule
        res = await db.execute(
            select(LearningModule).where(LearningModule.training_id == training.id)
        )
        if res.scalar_one_or_none():
            return  # Already exists
        module = LearningModule(
            title=training.title,
            description=training.description,
            category_id=training.category_id,
            department_id=None,
            training_id=training.id,
            created_by=training.created_by,
        )
        db.add(module)
        await db.flush()

    # --- Categories ---
    @staticmethod
    async def list_categories(db: AsyncSession) -> List[TrainingCategory]:
        result = await db.execute(select(TrainingCategory).order_by(TrainingCategory.name.asc()))
        return result.scalars().all()

    @staticmethod
    async def create_category(db: AsyncSession, payload: TrainingCategoryCreate) -> TrainingCategory:
        cat = TrainingCategory(**payload.model_dump())
        db.add(cat)
        await db.flush()
        await db.refresh(cat, ["id", "created_at", "updated_at"])
        return cat

    # --- Documents ---
    @staticmethod
    async def upload_document(
        db: AsyncSession, 
        training_id: UUID, 
        title: str, 
        file_path: str
    ) -> TrainingDocument:
        doc = TrainingDocument(
            title=title,
            file_path=file_path,
            training_id=training_id
        )
        db.add(doc)
        await db.flush()
        await db.refresh(doc, ["id", "created_at", "updated_at"])
        return doc

    # --- Import Center ---
    @staticmethod
    async def generate_import_template(db: AsyncSession) -> io.BytesIO:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)
            
        ws = wb.create_sheet(title="Training Import Template")
        ws.views.sheetView[0].showGridLines = True
        
        headers = [
            "Training Name", "Description", "Department", "Training Category",
            "Trainer Name", "Training Date", "Start Time", "End Time",
            "Duration Hours", "Mode", "Venue", "Max Seats",
            "Enrollment Deadline Date", "Enrollment Deadline Time",
            "Training Type", "Status"
        ]
        
        sample_row = [
            "Advanced Excel Analytics",
            "Learn advanced formula design, pivot tables, and dashboard integration.",
            "Accounts",
            "Technical",
            "Jane Doe",
            "25-06-2026",
            "10:00",
            "12:00",
            2.0,
            "Online",
            "Zoom Meeting",
            50,
            "24-06-2026",
            "18:00",
            "Mandatory",
            "Scheduled"
        ]
        
        ws.append(headers)
        ws.append(sample_row)
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )
        
        ws.row_dimensions[1].height = 28
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = 22
            
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            
        for r in range(2, 100):
            ws.cell(row=r, column=6).number_format = '@'
            ws.cell(row=r, column=7).number_format = '@'
            ws.cell(row=r, column=8).number_format = '@'
            ws.cell(row=r, column=13).number_format = '@'
            ws.cell(row=r, column=14).number_format = '@'
            ws.cell(row=r, column=9).number_format = '0.00'
            ws.cell(row=r, column=12).number_format = '0'
            
        from app.departments.models import Department
        dept_res = await db.execute(select(Department.name))
        dept_names = [d[0] for d in dept_res.all()]
        
        dv_mode = DataValidation(type="list", formula1='"Online,In-Person,Hybrid"', allow_blank=True)
        dv_mode.error = 'Value must be Online, In-Person, or Hybrid'
        dv_mode.errorTitle = 'Invalid Delivery Mode'
        ws.add_data_validation(dv_mode)
        dv_mode.add("J2:J100")
        
        dv_type = DataValidation(type="list", formula1='"Internal,External,Mandatory,Optional"', allow_blank=True)
        dv_type.error = 'Value must be Internal, External, Mandatory, or Optional'
        dv_type.errorTitle = 'Invalid Training Type'
        ws.add_data_validation(dv_type)
        dv_type.add("O2:O100")
        
        dv_status = DataValidation(type="list", formula1='"Scheduled,Draft"', allow_blank=True)
        dv_status.error = 'Value must be Scheduled or Draft'
        dv_status.errorTitle = 'Invalid Status'
        ws.add_data_validation(dv_status)
        dv_status.add("P2:P100")
        
        if dept_names:
            dept_formula = f'"{",".join(dept_names)}"'
            if len(dept_formula) <= 255:
                dv_dept = DataValidation(type="list", formula1=dept_formula, allow_blank=True)
                dv_dept.error = 'Please select a valid department'
                dv_dept.errorTitle = 'Invalid Department'
                ws.add_data_validation(dv_dept)
                dv_dept.add("C2:C100")
                
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    @staticmethod
    async def parse_import_file(db: AsyncSession, file_content: bytes) -> dict:
        from openpyxl import load_workbook
        
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        sheet_name = "Training Import Template"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            
        records = []
        
        header_row = None
        for r in range(1, 10):
            row_vals = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, 17)]
            if "training name" in row_vals or "training title" in row_vals:
                header_row = r
                break
                
        if not header_row:
            header_row = 1
            
        from app.departments.models import Department
        from app.trainings.categories import TrainingCategory
        
        dept_stmt = select(Department)
        all_depts = (await db.execute(dept_stmt)).scalars().all()
        
        records_found = 0
        valid_records = 0
        warnings_count = 0
        failed_records = 0
        
        for r in range(header_row + 1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 17)]
            if not any(v is not None and str(v).strip() != "" for v in row_vals):
                continue
                
            records_found += 1
            idx = r - header_row
            
            title = str(row_vals[0] or "").strip()
            desc = str(row_vals[1] or "").strip() if row_vals[1] is not None else None
            dept_str = str(row_vals[2] or "").strip()
            cat_str = str(row_vals[3] or "").strip() if row_vals[3] is not None else None
            trainer = str(row_vals[4] or "").strip() if row_vals[4] is not None else None
            date_raw = row_vals[5]
            start_time_raw = row_vals[6]
            end_time_raw = row_vals[7]
            duration_raw = row_vals[8]
            mode_str = str(row_vals[9] or "").strip()
            venue = str(row_vals[10] or "").strip() if row_vals[10] is not None else None
            max_seats_raw = row_vals[11]
            deadline_date_raw = row_vals[12]
            deadline_time_raw = row_vals[13]
            type_str = str(row_vals[14] or "").strip()
            status_str = str(row_vals[15] or "").strip()
            
            errors = []
            warnings = []
            
            if not title:
                errors.append("Training Name is mandatory.")
                
            mode = "ONLINE"
            if not mode_str:
                errors.append("Mode is mandatory.")
            else:
                m_norm = mode_str.lower().replace(" ", "").replace("_", "")
                if m_norm in ("online",):
                    mode = "ONLINE"
                elif m_norm in ("inperson", "in-person"):
                    mode = "IN_PERSON"
                elif m_norm in ("hybrid",):
                    mode = "HYBRID"
                else:
                    errors.append(f"Invalid Mode: '{mode_str}'. Must be In-Person, Online, or Hybrid.")
                    
            if mode == "IN_PERSON" and not venue:
                errors.append("Venue is mandatory for In-Person training.")
                
            is_global = False
            dept_ids = []
            if not dept_str:
                errors.append("Department is mandatory.")
            elif dept_str.lower() in ("all", "global", "all departments"):
                is_global = True
            else:
                depts_split = [d.strip() for d in re.split(r'[,;]', dept_str) if d.strip()]
                for d_name in depts_split:
                    aliases = {
                        "it": "IT Support",
                        "finance": "Accounts"
                    }
                    d_norm = d_name.lower()
                    mapped_name = aliases.get(d_norm, d_name)
                    
                    matched_dept = None
                    for d in all_depts:
                        if d.name.lower() == mapped_name.lower() or d.code.strip().lower() == mapped_name.lower():
                            matched_dept = d
                            break
                    if matched_dept:
                        dept_ids.append(str(matched_dept.id))
                    else:
                        errors.append(f"Department '{d_name}' not found.")
                        
            parsed_date = None
            if not date_raw:
                errors.append("Training Date is mandatory.")
            else:
                try:
                    parsed_date = parse_excel_date(date_raw)
                except ValueError as e:
                    errors.append(str(e))
                    
            parsed_start_time = None
            if not start_time_raw:
                errors.append("Start Time is mandatory.")
            else:
                try:
                    parsed_start_time = parse_excel_time(start_time_raw)
                except ValueError as e:
                    errors.append(str(e))
                    
            parsed_end_time = None
            if end_time_raw:
                try:
                    parsed_end_time = parse_excel_time(end_time_raw)
                except ValueError as e:
                    errors.append(f"End Time: {str(e)}")
                    
            duration_hours = 2.0
            if duration_raw is not None and duration_raw != "":
                try:
                    duration_hours = float(duration_raw)
                    if duration_hours <= 0:
                        errors.append("Duration Hours must be greater than 0.")
                except (ValueError, TypeError):
                    errors.append("Duration Hours must be a number.")
            else:
                errors.append("Duration Hours is mandatory.")
                
            max_seats = 20
            if max_seats_raw is not None and max_seats_raw != "":
                try:
                    max_seats = int(max_seats_raw)
                    if max_seats <= 0:
                        errors.append("Max Seats must be greater than 0.")
                except (ValueError, TypeError):
                    errors.append("Max Seats must be an integer.")
                    
            is_mandatory = False
            training_type = "INTERNAL"
            if not type_str:
                errors.append("Training Type is mandatory.")
            else:
                t_norm = type_str.lower()
                if t_norm == "mandatory":
                    is_mandatory = True
                    training_type = "INTERNAL"
                elif t_norm == "optional":
                    is_mandatory = False
                    training_type = "INTERNAL"
                elif t_norm == "internal":
                    training_type = "INTERNAL"
                elif t_norm == "external":
                    training_type = "EXTERNAL"
                elif t_norm == "workshop":
                    training_type = "WORKSHOP"
                elif t_norm == "online":
                    training_type = "ONLINE"
                elif t_norm == "certification":
                    training_type = "CERTIFICATION"
                else:
                    errors.append(f"Invalid Training Type: '{type_str}'. Must be Internal, External, Mandatory, or Optional.")
                    
            status = "DRAFT"
            if not status_str:
                errors.append("Status is mandatory.")
            else:
                s_norm = status_str.lower()
                if s_norm == "draft":
                    status = "DRAFT"
                elif s_norm == "scheduled":
                    status = "SCHEDULED"
                else:
                    errors.append(f"Invalid Status: '{status_str}'. Must be Scheduled or Draft.")
                    
            parsed_deadline_date = None
            if deadline_date_raw:
                try:
                    parsed_deadline_date = parse_excel_date(deadline_date_raw)
                except ValueError as e:
                    errors.append(f"Enrollment Deadline Date: {str(e)}")
            elif parsed_date:
                parsed_deadline_date = parsed_date - timedelta(days=1)
                
            parsed_deadline_time = "23:59"
            if deadline_time_raw:
                try:
                    parsed_deadline_time = parse_excel_time(deadline_time_raw)
                except ValueError as e:
                    errors.append(f"Enrollment Deadline Time: {str(e)}")
                    
            is_duplicate = False
            if not errors and parsed_date and parsed_start_time:
                db_start_time = format_time_for_db(parsed_start_time)
                dup_stmt = select(Training).options(selectinload(Training.departments)).where(
                    Training.title == title,
                    Training.start_date == parsed_date,
                    Training.start_time == db_start_time
                )
                dup_res = await db.execute(dup_stmt)
                dup_trainings = dup_res.scalars().all()
                
                for dt in dup_trainings:
                    if is_global and dt.is_global:
                        is_duplicate = True
                        break
                    elif not is_global and not dt.is_global:
                        dt_dept_ids = {str(dep.id) for dep in dt.departments}
                        if set(dept_ids) & dt_dept_ids:
                            is_duplicate = True
                            break
                            
            if is_duplicate:
                warnings.append("This training record already exists in the system (duplicate name, date, start time, and department).")
                
            is_valid = len(errors) == 0
            if is_valid:
                valid_records += 1
            else:
                failed_records += 1
                
            if warnings:
                warnings_count += 1
                
            records.append({
                "index": idx,
                "title": title,
                "description": desc,
                "department": dept_str,
                "training_category": cat_str,
                "trainer_name": trainer,
                "training_date": parsed_date.strftime("%d-%m-%Y") if parsed_date else str(date_raw or ""),
                "start_time": parsed_start_time if parsed_start_time else str(start_time_raw or ""),
                "end_time": parsed_end_time if parsed_end_time else (str(end_time_raw or "") if end_time_raw else None),
                "duration_hours": duration_hours,
                "mode": mode,
                "venue": venue,
                "max_seats": max_seats,
                "enrollment_deadline_date": parsed_deadline_date.strftime("%d-%m-%Y") if parsed_deadline_date else str(deadline_date_raw or ""),
                "enrollment_deadline_time": parsed_deadline_time,
                "training_type": training_type,
                "status": status,
                "is_valid": is_valid,
                "errors": errors,
                "warnings": warnings,
                "is_duplicate": is_duplicate,
                "is_mandatory": is_mandatory,
                "is_global": is_global,
                "dept_ids": dept_ids
            })
            
        summary = {
            "records_found": records_found,
            "valid_records": valid_records,
            "warnings": warnings_count,
            "failed_records": failed_records
        }
        
        return {
            "summary": summary,
            "records": records
        }

    @staticmethod
    async def confirm_import(
        db: AsyncSession,
        records: list[dict],
        duplicate_strategy: str,
        created_by: UUID,
        source_file: Optional[str] = None
    ) -> dict:
        from app.trainings.models import Training, TrainingStatus, TrainingType, DeliveryMode
        from app.departments.models import Department
        from app.trainings.categories import TrainingCategory
        from app.attendance.service import AttendanceService
        from app.effectiveness.service import EffectivenessService
        from app.enrollments.models import Enrollment, EnrollmentStatus
        from app.attendance.models import AttendanceRecord, AttendanceStatus
        
        successfully_imported = 0
        skipped_duplicates = 0
        failed_records = 0
        
        try:
            for rec in records:
                if not rec.get("is_valid", True):
                    failed_records += 1
                    continue
                    
                title = rec["title"]
                desc = rec.get("description")
                dept_str = rec["department"]
                cat_str = rec.get("training_category")
                trainer = rec.get("trainer_name")
                date_str = rec["training_date"]
                start_time_str = rec["start_time"]
                end_time_str = rec.get("end_time")
                duration_hours = rec["duration_hours"]
                mode = rec["mode"]
                venue = rec.get("venue")
                max_seats = rec.get("max_seats", 20)
                deadline_date_str = rec.get("enrollment_deadline_date")
                deadline_time_str = rec.get("enrollment_deadline_time")
                training_type = rec["training_type"]
                status = rec["status"]
                is_mandatory = rec.get("is_mandatory", False)
                is_global = rec.get("is_global", False)
                dept_ids = rec.get("dept_ids", [])
                
                parsed_date = datetime.strptime(date_str, "%d-%m-%Y").date()
                parsed_start_time = datetime.strptime(start_time_str, "%H:%M").time()
                db_start_time = parsed_start_time.strftime("%I:%M %p")
                
                category_id = None
                if cat_str:
                    cat_stmt = select(TrainingCategory).where(func.lower(TrainingCategory.name) == cat_str.lower())
                    cat_res = await db.execute(cat_stmt)
                    cat_obj = cat_res.scalar_one_or_none()
                    if not cat_obj:
                        cat_obj = TrainingCategory(name=cat_str, description=f"Autocreated from import")
                        db.add(cat_obj)
                        await db.flush()
                    category_id = cat_obj.id
                    
                today = date.today()
                if parsed_date < today:
                    db_status = TrainingStatus.COMPLETED
                else:
                    db_status = TrainingStatus[status]
                    
                dup_stmt = select(Training).options(selectinload(Training.departments)).where(
                    Training.title == title,
                    Training.start_date == parsed_date,
                    Training.start_time == db_start_time
                )
                dup_res = await db.execute(dup_stmt)
                dup_trainings = dup_res.scalars().all()
                
                existing_training = None
                for dt in dup_trainings:
                    if is_global and dt.is_global:
                        existing_training = dt
                        break
                    elif not is_global and not dt.is_global:
                        dt_dept_ids = {str(dep.id) for dep in dt.departments}
                        if set(dept_ids) & dt_dept_ids:
                            existing_training = dt
                            break
                            
                is_dup = existing_training is not None
                
                if is_dup:
                    if duplicate_strategy == "skip":
                        skipped_duplicates += 1
                        continue
                    elif duplicate_strategy == "replace":
                        await db.delete(existing_training)
                        await db.flush()
                        existing_training = None
                    elif duplicate_strategy == "update":
                        existing_training.description = desc
                        existing_training.training_type = TrainingType[training_type]
                        existing_training.delivery_mode = DeliveryMode[mode]
                        existing_training.status = db_status
                        existing_training.duration_hours = duration_hours
                        existing_training.max_hours_allowed = duration_hours
                        existing_training.max_participants = max_seats
                        existing_training.venue = venue
                        existing_training.trainer_name = trainer
                        existing_training.is_mandatory = is_mandatory
                        existing_training.is_global = is_global
                        existing_training.category_id = category_id
                        
                        if deadline_date_str:
                            d_date = datetime.strptime(deadline_date_str, "%d-%m-%Y").date()
                            d_time = datetime.strptime(deadline_time_str or "23:59", "%H:%M").time()
                            existing_training.enrollment_deadline = datetime.combine(d_date, d_time)
                        else:
                            existing_training.enrollment_deadline = None
                            
                        if is_global:
                            existing_training.departments = []
                        else:
                            dept_stmt = select(Department).where(Department.id.in_(dept_ids))
                            dept_res = await db.execute(dept_stmt)
                            existing_training.departments = list(dept_res.scalars().all())
                            
                        await db.flush()
                        training = existing_training
                        await TrainingService._ensure_learning_module(db, training)
                        successfully_imported += 1
                        
                if not is_dup or duplicate_strategy == "replace":
                    training = Training(
                        title=title,
                        description=desc,
                        training_type=TrainingType[training_type],
                        delivery_mode=DeliveryMode[mode],
                        status=db_status,
                        start_date=parsed_date,
                        start_time=db_start_time,
                        duration_hours=duration_hours,
                        max_hours_allowed=duration_hours,
                        max_participants=max_seats,
                        available_seats=max_seats,
                        venue=venue,
                        trainer_name=trainer,
                        is_mandatory=is_mandatory,
                        is_global=is_global,
                        category_id=category_id,
                        created_by=created_by
                    )
                    
                    if deadline_date_str:
                        d_date = datetime.strptime(deadline_date_str, "%d-%m-%Y").date()
                        d_time = datetime.strptime(deadline_time_str or "23:59", "%H:%M").time()
                        training.enrollment_deadline = datetime.combine(d_date, d_time)
                        
                    if not is_global and dept_ids:
                        dept_stmt = select(Department).where(Department.id.in_(dept_ids))
                        dept_res = await db.execute(dept_stmt)
                        training.departments = list(dept_res.scalars().all())
                        
                    db.add(training)
                    await db.flush()
                    await TrainingService._ensure_learning_module(db, training)
                    successfully_imported += 1
                    
                from app.employees.models import Employee, EmploymentStatus
                emp_stmt = select(Employee).where(Employee.status == EmploymentStatus.ACTIVE)
                if not is_global and dept_ids:
                    emp_stmt = emp_stmt.where(Employee.department_id.in_(dept_ids))
                emp_res = await db.execute(emp_stmt)
                active_employees = emp_res.scalars().all()
                
                if db_status == TrainingStatus.COMPLETED:
                    session = await AttendanceService.ensure_session(db, training.id)
                    start_dt = datetime.combine(parsed_date, parsed_start_time)
                    
                    for emp in active_employees:
                        enr_stmt = select(Enrollment).where(
                            Enrollment.employee_id == emp.id,
                            Enrollment.training_id == training.id
                        )
                        enrollment = (await db.execute(enr_stmt)).scalar_one_or_none()
                        if not enrollment:
                            enrollment = Enrollment(
                                employee_id=emp.id,
                                training_id=training.id,
                                status=EnrollmentStatus.APPROVED,
                                progress=0.0
                            )
                            db.add(enrollment)
                            await db.flush()
                            
                        att_stmt = select(AttendanceRecord).where(
                            AttendanceRecord.employee_id == emp.id,
                            AttendanceRecord.training_id == training.id
                        )
                        att_rec = (await db.execute(att_stmt)).scalar_one_or_none()
                        if not att_rec:
                            att_rec = AttendanceRecord(
                                employee_id=emp.id,
                                training_id=training.id,
                                session_id=session.id,
                                status=AttendanceStatus.PRESENT,
                                marked_at=start_dt,
                                attendance_open_time=session.opens_at,
                                attendance_close_time=session.closes_at
                            )
                            db.add(att_rec)
                            await db.flush()
                        else:
                            att_rec.status = AttendanceStatus.PRESENT
                            att_rec.marked_at = start_dt
                            att_rec.session_id = session.id
                            await db.flush()
                            
                    await EffectivenessService.assign_training_effectiveness(db, training)
                    
                elif db_status == TrainingStatus.SCHEDULED and is_mandatory:
                    for emp in active_employees:
                        enr_stmt = select(Enrollment).where(
                            Enrollment.employee_id == emp.id,
                            Enrollment.training_id == training.id
                        )
                        enrollment = (await db.execute(enr_stmt)).scalar_one_or_none()
                        if not enrollment:
                            enrollment = Enrollment(
                                employee_id=emp.id,
                                training_id=training.id,
                                status=EnrollmentStatus.APPROVED,
                                progress=0.0
                            )
                            db.add(enrollment)
                            
                    await db.flush()
                    
            from app.trainings.models import TrainingImportHistory
            history = TrainingImportHistory(
                created_by=created_by,
                records_imported=successfully_imported,
                records_failed=failed_records,
                records_skipped=skipped_duplicates,
                source_file=source_file
            )
            db.add(history)
            await db.commit()
            
            return {
                "successfully_imported": successfully_imported,
                "failed_records": failed_records,
                "skipped_duplicates": skipped_duplicates
            }
        except Exception as e:
            await db.rollback()
            raise e

    @staticmethod
    async def get_import_history(db: AsyncSession) -> list[dict]:
        from sqlalchemy.orm import selectinload
        stmt = select(TrainingImportHistory).options(selectinload(TrainingImportHistory.imported_by)).order_by(TrainingImportHistory.created_at.desc())
        res = await db.execute(stmt)
        records = res.scalars().all()
        
        results = []
        for r in records:
            results.append({
                "id": r.id,
                "created_at": r.created_at,
                "records_imported": r.records_imported,
                "records_failed": r.records_failed,
                "records_skipped": r.records_skipped,
                "source_file": r.source_file,
                "imported_by_name": r.imported_by.full_name if r.imported_by else "System"
            })
        return results
