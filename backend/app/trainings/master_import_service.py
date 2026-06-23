import io
import re
from datetime import date, datetime, time, timedelta
from typing import Any, Optional, Dict, List
from uuid import UUID

from sqlalchemy import select, or_, and_, func
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.trainings.models import Training, TrainingType, TrainingStatus, DeliveryMode, TrainingImportHistory
from app.trainings.categories import TrainingCategory
from app.departments.models import Department
from app.employees.models import Employee, EmploymentStatus
from app.users.models import User
from app.roles.models import Role
from app.users.service import UserService
from app.enrollments.models import Enrollment, EnrollmentStatus
from app.attendance.models import AttendanceSession, AttendanceRecord, AttendanceStatus
from app.attendance.service import AttendanceService
from app.effectiveness.models import Effectiveness, EffectivenessLevel, EffectivenessStatus
from app.effectiveness.service import EffectivenessService

from app.trainings.service import parse_excel_date, parse_excel_time, format_time_for_db


def financial_year_of_date(d: date) -> str:
    """Returns financial year for a date (e.g. 15-06-2026 -> '2026-2027')."""
    if d.month >= 4:
        return f"{d.year}-{d.year + 1}"
    else:
        return f"{d.year - 1}-{d.year}"


class MasterImportService:

    @staticmethod
    async def generate_master_template(db: AsyncSession) -> io.BytesIO:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        # Remove default active sheet
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        # 1. Create Lookup_Master (Hidden) Sheet first to reference it
        ws_lookup = wb.create_sheet(title="Lookup_Master")
        ws_lookup.views.sheetView[0].showGridLines = True
        ws_lookup.sheet_state = "hidden"

        # Populate Lookup Options
        # Col A: Departments
        dept_res = await db.execute(select(Department.name).order_by(Department.name.asc()))
        dept_names = [d[0] for d in dept_res.all()]
        if not dept_names:
            dept_names = ["Management", "Accounts", "QA", "HR", "IT Support"]
        
        # Col B: Categories
        ws_lookup.cell(row=1, column=1, value="Departments")
        for i, val in enumerate(dept_names):
            ws_lookup.cell(row=i + 2, column=1, value=val)

        ws_lookup.cell(row=1, column=2, value="Categories")
        categories = ["Technical", "Soft Skills", "Leadership", "Compliance", "Onboarding"]
        for i, val in enumerate(categories):
            ws_lookup.cell(row=i + 2, column=2, value=val)

        # Col C: Training Types
        ws_lookup.cell(row=1, column=3, value="Training Types")
        types = ["Internal", "External", "Mandatory", "Optional"]
        for i, val in enumerate(types):
            ws_lookup.cell(row=i + 2, column=3, value=val)

        # Col D: Attendance Statuses
        ws_lookup.cell(row=1, column=4, value="Attendance Statuses")
        statuses = ["Present", "Absent", "Late", "Excused"]
        for i, val in enumerate(statuses):
            ws_lookup.cell(row=i + 2, column=4, value=val)

        # Col E: Ratings
        ws_lookup.cell(row=1, column=5, value="Ratings")
        ratings = [1, 2, 3, 4, 5]
        for i, val in enumerate(ratings):
            ws_lookup.cell(row=i + 2, column=5, value=val)

        # Col F: Modes
        ws_lookup.cell(row=1, column=6, value="Modes")
        modes = ["Online", "In-Person", "Hybrid"]
        for i, val in enumerate(modes):
            ws_lookup.cell(row=i + 2, column=6, value=val)

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        def style_sheet(ws, headers, sample=None):
            ws.views.sheetView[0].showGridLines = True
            ws.append(headers)
            if sample:
                ws.append(sample)
            ws.row_dimensions[1].height = 28
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
                col_letter = cell.column_letter
                ws.column_dimensions[col_letter].width = 22
            if sample:
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=2, column=col_idx)
                    cell.font = Font(name="Calibri", size=11)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = thin_border

        # ── Sheet 1: Trainings ──
        ws_train = wb.create_sheet(title="Trainings")
        style_sheet(ws_train, [
            "Training Name", "Description", "Department", "Category",
            "Trainer Name", "Training Date", "Start Time", "End Time",
            "Duration Hours", "Mode", "Venue", "Max Seats",
            "Enrollment Deadline Date", "Enrollment Deadline Time",
            "Training Type", "Status", "Material Title", "Material Link or Path"
        ], [
            "Advanced Excel Analytics", "Learn advanced formula design and KPI sheets.", "Accounts", "Technical",
            "Jane Doe", "25-06-2026", "10:00", "12:00", 2.0, "Online", "Zoom Meeting", 50,
            "24-06-2026", "18:00", "Mandatory", "Scheduled", "Excel Shortcuts Guide", "https://example.com/shortcuts.pdf"
        ])

        # ── Sheet 2: Enrollments ──
        ws_enroll = wb.create_sheet(title="Enrollments")
        style_sheet(ws_enroll, [
            "Employee ID", "Employee Name", "Department", "Training Name",
            "Training Date", "Enrollment Date", "Enrollment Status"
        ], [
            "EMP1001", "John Smith", "Accounts", "Advanced Excel Analytics",
            "25-06-2026", "10-06-2026", "enrolled"
        ])

        # ── Sheet 3: Attendance ──
        ws_attend = wb.create_sheet(title="Attendance")
        style_sheet(ws_attend, [
            "Employee ID", "Employee Name", "Department", "Training Name",
            "Training Date", "Attendance Status", "Hours Earned", "Attendance Marked Date"
        ], [
            "EMP1001", "John Smith", "Accounts", "Advanced Excel Analytics",
            "25-06-2026", "Present", 2.0, "25-06-2026"
        ])

        # ── Sheet 4: Feedback ──
        ws_feedback = wb.create_sheet(title="Feedback")
        style_sheet(ws_feedback, [
            "Employee ID", "Employee Name", "Training Name", "Training Date",
            "Overall Rating", "Trainer Rating", "Content Rating", "Feedback Comments"
        ], [
            "EMP1001", "John Smith", "Advanced Excel Analytics", "25-06-2026",
            5, 5, 4, "Excellent session, very practical examples."
        ])

        # ── Sheet 5: Learning_Hours ──
        ws_hours = wb.create_sheet(title="Learning_Hours")
        style_sheet(ws_hours, [
            "Employee ID", "Employee Name", "Department", "Training Name",
            "Training Date", "Learning Hours Earned", "Financial Year"
        ], [
            "EMP1001", "John Smith", "Accounts", "Advanced Excel Analytics",
            "25-06-2026", 2.0, "2026-2027"
        ])

        # ── Sheet 6: Employees ──
        ws_emp = wb.create_sheet(title="Employees")
        style_sheet(ws_emp, [
            "Employee ID", "Employee Name", "Department", "Designation",
            "Manager ID", "Date of Joining", "Status"
        ], [
            "EMP1001", "John Smith", "Accounts", "Senior Accountant",
            "EMP1000", "01-01-2024", "active"
        ])

        # ── Sheet 7: Instructions ──
        ws_inst = wb.create_sheet(title="Instructions")
        ws_inst.views.sheetView[0].showGridLines = True
        ws_inst.column_dimensions["A"].width = 25
        ws_inst.column_dimensions["B"].width = 80
        
        ws_inst.append(["Import Section", "Guidelines & Validation Rules"])
        ws_inst.cell(row=1, column=1).font = header_font
        ws_inst.cell(row=1, column=1).fill = header_fill
        ws_inst.cell(row=1, column=2).font = header_font
        ws_inst.cell(row=1, column=2).fill = header_fill

        instructions = [
            ("Dates", "Must be formatted as DD-MM-YYYY (e.g. 25-06-2026)."),
            ("Times", "Must be formatted as HH:MM in 24-hour style (e.g. 14:30) or standard 12-hour style (e.g. 02:30 PM)."),
            ("Trainings Sheet", "Training Name, Training Date, Duration Hours, Mode, Training Type, and Status are mandatory. If Mode is In-Person, Venue is mandatory."),
            ("Enrollments Sheet", "Employee ID, Training Name, Training Date, and Enrollment Status are mandatory. Valid statuses: enrolled, completed, withdrawn, pending, rejected."),
            ("Attendance Sheet", "Employee ID, Training Name, Training Date, and Attendance Status are mandatory. Valid statuses: Present, Absent, Late, Excused."),
            ("Feedback Sheet", "Employee ID, Training Name, Training Date, and Ratings (1-5) are mandatory. Comments are optional."),
            ("Learning_Hours Sheet", "Employee ID, Training Name, Training Date, and Learning Hours Earned (numeric) are mandatory. Financial Year is optional and auto-computed if omitted."),
            ("Employees Sheet", "Employee ID, Employee Name, Department, Date of Joining, and Status are mandatory. Valid statuses: active, on_leave, terminated.")
        ]
        for row_idx, (section, text) in enumerate(instructions):
            ws_inst.cell(row=row_idx + 2, column=1, value=section).font = Font(name="Calibri", size=11, bold=True)
            ws_inst.cell(row=row_idx + 2, column=1).alignment = Alignment(vertical="top")
            ws_inst.cell(row=row_idx + 2, column=2, value=text).alignment = Alignment(wrap_text=True)
            ws_inst.cell(row=row_idx + 2, column=1).border = thin_border
            ws_inst.cell(row=row_idx + 2, column=2).border = thin_border

        # ── Setup Dropdown Validations ──
        dv_dept = DataValidation(type="list", formula1=f"Lookup_Master!$A$2:$A${len(dept_names)+1}", allow_blank=True)
        dv_cat = DataValidation(type="list", formula1=f"Lookup_Master!$B$2:$B${len(categories)+1}", allow_blank=True)
        dv_type = DataValidation(type="list", formula1=f"Lookup_Master!$C$2:$C${len(types)+1}", allow_blank=True)
        dv_attend = DataValidation(type="list", formula1=f"Lookup_Master!$D$2:$D${len(statuses)+1}", allow_blank=True)
        dv_rating = DataValidation(type="list", formula1=f"Lookup_Master!$E$2:$E${len(ratings)+1}", allow_blank=True)
        dv_mode = DataValidation(type="list", formula1=f"Lookup_Master!$F$2:$F${len(modes)+1}", allow_blank=True)

        dv_enroll_status = DataValidation(type="list", formula1='"enrolled,completed,withdrawn,pending,rejected"', allow_blank=True)
        dv_emp_status = DataValidation(type="list", formula1='"active,on_leave,terminated"', allow_blank=True)
        dv_train_status = DataValidation(type="list", formula1='"Scheduled,Draft"', allow_blank=True)

        # Apply to sheets
        # Sheet 1: Trainings
        ws_train.add_data_validation(dv_dept)
        dv_dept.add("C2:C100")
        ws_train.add_data_validation(dv_cat)
        dv_cat.add("D2:D100")
        ws_train.add_data_validation(dv_mode)
        dv_mode.add("J2:J100")
        ws_train.add_data_validation(dv_type)
        dv_type.add("O2:O100")
        ws_train.add_data_validation(dv_train_status)
        dv_train_status.add("P2:P100")

        # Sheet 2: Enrollments
        ws_enroll.add_data_validation(dv_dept)
        dv_dept.add("C2:C100")
        ws_enroll.add_data_validation(dv_enroll_status)
        dv_enroll_status.add("G2:G100")

        # Sheet 3: Attendance
        ws_attend.add_data_validation(dv_dept)
        dv_dept.add("C2:C100")
        ws_attend.add_data_validation(dv_attend)
        dv_attend.add("F2:F100")

        # Sheet 4: Feedback
        ws_feedback.add_data_validation(dv_rating)
        dv_rating.add("E2:E100")
        dv_rating.add("F2:F100")
        dv_rating.add("G2:G100")

        # Sheet 5: Learning Hours
        ws_hours.add_data_validation(dv_dept)
        dv_dept.add("C2:C100")

        # Sheet 6: Employees
        ws_emp.add_data_validation(dv_dept)
        dv_dept.add("C2:C100")
        ws_emp.add_data_validation(dv_emp_status)
        dv_emp_status.add("G2:G100")

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    @staticmethod
    async def parse_master_file(db: AsyncSession, file_content: bytes) -> dict:
        from openpyxl import load_workbook
        
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        
        parsed_sheets = {
            "trainings": [],
            "enrollments": [],
            "attendance": [],
            "feedback": [],
            "learning_hours": [],
            "employees": []
        }

        # Cache reference lists to reduce DB calls
        from app.departments.models import Department
        dept_stmt = select(Department)
        all_depts = (await db.execute(dept_stmt)).scalars().all()
        
        # Check active employees in DB
        from app.employees.models import Employee
        emp_stmt = select(Employee.employee_code)
        existing_emp_codes = set((await db.execute(emp_stmt)).scalars().all())

        # Check existing trainings in DB (title + start_date)
        train_stmt = select(Training.title, Training.start_date)
        existing_train_keys = set((await db.execute(train_stmt)).all())

        # Total counts
        records_found = 0
        valid_records = 0
        warnings_count = 0
        failed_records = 0

        # Helper to read cell safely
        def cell_val(ws, r, c) -> Any:
            return ws.cell(row=r, column=c).value

        # ── Parse Employees ──
        if "Employees" in wb.sheetnames:
            ws = wb["Employees"]
            for r in range(2, ws.max_row + 1):
                row_vals = [cell_val(ws, r, c) for c in range(1, 8)]
                if not any(v is not None and str(v).strip() != "" for v in row_vals):
                    continue
                records_found += 1
                idx = r - 1
                
                emp_id = str(row_vals[0] or "").strip()
                emp_name = str(row_vals[1] or "").strip()
                dept_str = str(row_vals[2] or "").strip()
                designation = str(row_vals[3] or "").strip() if row_vals[3] is not None else None
                manager_str = str(row_vals[4] or "").strip() if row_vals[4] is not None else None
                doj_raw = row_vals[5]
                status_str = str(row_vals[6] or "").strip()

                errors = []
                warnings = []
                is_duplicate = False

                if not emp_id:
                    errors.append("Employee ID is mandatory.")
                if not emp_name:
                    errors.append("Employee Name is mandatory.")
                if not dept_str:
                    errors.append("Department is mandatory.")
                else:
                    # Validate dept
                    aliases = {"it": "IT Support", "finance": "Accounts"}
                    mapped_name = aliases.get(dept_str.lower(), dept_str)
                    matched_dept = None
                    for d in all_depts:
                        if d.name.lower() == mapped_name.lower() or d.code.strip().lower() == mapped_name.lower():
                            matched_dept = d
                            break
                    if not matched_dept:
                        errors.append(f"Department '{dept_str}' not found.")

                parsed_doj = None
                if not doj_raw:
                    errors.append("Date of Joining is mandatory.")
                else:
                    try:
                        parsed_doj = parse_excel_date(doj_raw)
                    except ValueError as e:
                        errors.append(str(e))

                status = "active"
                if not status_str:
                    errors.append("Status is mandatory.")
                else:
                    s_norm = status_str.lower()
                    if s_norm in ("active", "on_leave", "terminated"):
                        status = s_norm
                    else:
                        errors.append(f"Invalid Status: '{status_str}'. Must be active, on_leave, or terminated.")

                if emp_id in existing_emp_codes:
                    is_duplicate = True
                    warnings.append(f"Employee with code '{emp_id}' already exists in database.")

                is_valid = len(errors) == 0
                if is_valid:
                    valid_records += 1
                else:
                    failed_records += 1
                if warnings:
                    warnings_count += 1

                parsed_sheets["employees"].append({
                    "index": idx,
                    "is_valid": is_valid,
                    "errors": errors,
                    "warnings": warnings,
                    "is_duplicate": is_duplicate,
                    "data": {
                        "employee_id": emp_id,
                        "employee_name": emp_name,
                        "department": dept_str,
                        "designation": designation,
                        "manager": manager_str,
                        "date_of_joining": parsed_doj.strftime("%d-%m-%Y") if parsed_doj else str(doj_raw or ""),
                        "status": status
                    }
                })

        # ── Parse Trainings ──
        if "Trainings" in wb.sheetnames:
            ws = wb["Trainings"]
            for r in range(2, ws.max_row + 1):
                row_vals = [cell_val(ws, r, c) for c in range(1, 19)]
                if not any(v is not None and str(v).strip() != "" for v in row_vals):
                    continue
                records_found += 1
                idx = r - 1

                title = str(row_vals[0] or "").strip()
                desc = str(row_vals[1] or "").strip() if row_vals[1] is not None else None
                dept_str = str(row_vals[2] or "").strip()
                cat_str = str(row_vals[3] or "").strip() if row_vals[3] is not None else None
                trainer = str(row_vals[4] or "").strip() if row_vals[4] is not None else None
                date_raw = row_vals[5]
                start_time_raw = row_vals[6]
                end_time_raw = row_vals[7]
                duration_raw = row_vals[8]
                mode_str = str(row_vals[9] or "").strip()
                venue = str(row_vals[10] or "").strip() if row_vals[10] is not None else None
                max_seats_raw = row_vals[11]
                deadline_date_raw = row_vals[12]
                deadline_time_raw = row_vals[13]
                type_str = str(row_vals[14] or "").strip()
                status_str = str(row_vals[15] or "").strip()
                material_title = str(row_vals[16] or "").strip() if len(row_vals) > 16 and row_vals[16] is not None else None
                material_link = str(row_vals[17] or "").strip() if len(row_vals) > 17 and row_vals[17] is not None else None

                errors = []
                warnings = []
                is_duplicate = False

                if not title:
                    errors.append("Training Name is mandatory.")
                if not dept_str:
                    errors.append("Department is mandatory.")
                else:
                    is_global = dept_str.lower() in ("all", "global", "all departments")
                    if not is_global:
                        depts_split = [d.strip() for d in re.split(r'[,;]', dept_str) if d.strip()]
                        for d_name in depts_split:
                            aliases = {"it": "IT Support", "finance": "Accounts"}
                            mapped_name = aliases.get(d_name.lower(), d_name)
                            matched_dept = None
                            for d in all_depts:
                                if d.name.lower() == mapped_name.lower() or d.code.strip().lower() == mapped_name.lower():
                                    matched_dept = d
                                    break
                            if not matched_dept:
                                errors.append(f"Department '{d_name}' not found.")

                parsed_date = None
                if not date_raw:
                    errors.append("Training Date is mandatory.")
                else:
                    try:
                        parsed_date = parse_excel_date(date_raw)
                    except ValueError as e:
                        errors.append(str(e))

                parsed_start_time = None
                if not start_time_raw:
                    errors.append("Start Time is mandatory.")
                else:
                    try:
                        parsed_start_time = parse_excel_time(start_time_raw)
                    except ValueError as e:
                        errors.append(str(e))

                parsed_end_time = None
                if end_time_raw:
                    try:
                        parsed_end_time = parse_excel_time(end_time_raw)
                    except ValueError as e:
                        errors.append(f"End Time: {str(e)}")

                duration_hours = 2.0
                if duration_raw is not None and duration_raw != "":
                    try:
                        duration_hours = float(duration_raw)
                        if duration_hours <= 0:
                            errors.append("Duration Hours must be greater than 0.")
                    except (ValueError, TypeError):
                        errors.append("Duration Hours must be a number.")
                else:
                    errors.append("Duration Hours is mandatory.")

                mode = "ONLINE"
                if not mode_str:
                    errors.append("Mode is mandatory.")
                else:
                    m_norm = mode_str.lower().replace(" ", "").replace("_", "")
                    if m_norm in ("online",):
                        mode = "ONLINE"
                    elif m_norm in ("inperson", "in-person"):
                        mode = "IN_PERSON"
                    elif m_norm in ("hybrid",):
                        mode = "HYBRID"
                    else:
                        errors.append(f"Invalid Mode: '{mode_str}'. Must be In-Person, Online, or Hybrid.")

                if mode == "IN_PERSON" and not venue:
                    errors.append("Venue is mandatory for In-Person training.")

                max_seats = 20
                if max_seats_raw is not None and max_seats_raw != "":
                    try:
                        max_seats = int(max_seats_raw)
                        if max_seats <= 0:
                            errors.append("Max Seats must be greater than 0.")
                    except (ValueError, TypeError):
                        errors.append("Max Seats must be an integer.")

                is_mandatory = False
                training_type = "INTERNAL"
                if not type_str:
                    errors.append("Training Type is mandatory.")
                else:
                    t_norm = type_str.lower()
                    if t_norm == "mandatory":
                        is_mandatory = True
                        training_type = "INTERNAL"
                    elif t_norm == "optional":
                        is_mandatory = False
                        training_type = "INTERNAL"
                    elif t_norm in ("internal", "external", "workshop", "online", "certification"):
                        training_type = t_norm.upper()
                    else:
                        errors.append(f"Invalid Training Type: '{type_str}'.")

                status = "DRAFT"
                if not status_str:
                    errors.append("Status is mandatory.")
                else:
                    s_norm = status_str.lower()
                    if s_norm in ("draft", "scheduled"):
                        status = s_norm.upper()
                    else:
                        errors.append(f"Invalid Status: '{status_str}'. Must be Scheduled or Draft.")

                parsed_deadline_date = None
                if deadline_date_raw:
                    try:
                        parsed_deadline_date = parse_excel_date(deadline_date_raw)
                    except ValueError as e:
                        errors.append(f"Enrollment Deadline Date: {str(e)}")
                elif parsed_date:
                    parsed_deadline_date = parsed_date - timedelta(days=1)

                parsed_deadline_time = "23:59"
                if deadline_time_raw:
                    try:
                        parsed_deadline_time = parse_excel_time(deadline_time_raw)
                    except ValueError as e:
                        errors.append(f"Enrollment Deadline Time: {str(e)}")

                if not errors and parsed_date:
                    if (title, parsed_date) in existing_train_keys:
                        is_duplicate = True
                        warnings.append("Training already exists in system with same Name and Date.")

                is_valid = len(errors) == 0
                if is_valid:
                    valid_records += 1
                else:
                    failed_records += 1
                if warnings:
                    warnings_count += 1

                parsed_sheets["trainings"].append({
                    "index": idx,
                    "is_valid": is_valid,
                    "errors": errors,
                    "warnings": warnings,
                    "is_duplicate": is_duplicate,
                    "data": {
                        "title": title,
                        "description": desc,
                        "department": dept_str,
                        "training_category": cat_str,
                        "trainer_name": trainer,
                        "training_date": parsed_date.strftime("%d-%m-%Y") if parsed_date else str(date_raw or ""),
                        "start_time": parsed_start_time if parsed_start_time else str(start_time_raw or ""),
                        "end_time": parsed_end_time if parsed_end_time else (str(end_time_raw or "") if end_time_raw else None),
                        "duration_hours": duration_hours,
                        "mode": mode,
                        "venue": venue,
                        "max_seats": max_seats,
                        "enrollment_deadline_date": parsed_deadline_date.strftime("%d-%m-%Y") if parsed_deadline_date else str(deadline_date_raw or ""),
                        "enrollment_deadline_time": parsed_deadline_time,
                        "training_type": training_type,
                        "status": status,
                        "is_mandatory": is_mandatory,
                        "material_title": material_title,
                        "material_link": material_link
                    }
                })

        # Helper validator for Employee/Training existence inside loop
        imported_emp_ids = {item["data"]["employee_id"] for item in parsed_sheets["employees"]}
        imported_train_keys = {(item["data"]["title"], item["data"]["training_date"]) for item in parsed_sheets["trainings"]}

        def check_employee_exists(emp_id: str) -> bool:
            return emp_id in existing_emp_codes or emp_id in imported_emp_ids

        def check_training_exists(title: str, date_str: str) -> bool:
            try:
                parsed_d = datetime.strptime(date_str, "%d-%m-%Y").date()
                if (title, parsed_d) in existing_train_keys:
                    return True
            except Exception:
                pass
            return (title, date_str) in imported_train_keys

        # ── Parse Enrollments ──
        if "Enrollments" in wb.sheetnames:
            ws = wb["Enrollments"]
            for r in range(2, ws.max_row + 1):
                row_vals = [cell_val(ws, r, c) for c in range(1, 8)]
                if not any(v is not None and str(v).strip() != "" for v in row_vals):
                    continue
                records_found += 1
                idx = r - 1

                emp_id = str(row_vals[0] or "").strip()
                emp_name = str(row_vals[1] or "").strip()
                dept_str = str(row_vals[2] or "").strip()
                title = str(row_vals[3] or "").strip()
                date_raw = row_vals[4]
                enr_date_raw = row_vals[5]
                status_str = str(row_vals[6] or "").strip()

                errors = []
                warnings = []
                is_duplicate = False

                if not emp_id:
                    errors.append("Employee ID is mandatory.")
                elif not check_employee_exists(emp_id):
                    errors.append(f"Employee ID '{emp_id}' not found in Employee sheet or DB.")

                if not title:
                    errors.append("Training Name is mandatory.")

                parsed_date = None
                if not date_raw:
                    errors.append("Training Date is mandatory.")
                else:
                    try:
                        parsed_date = parse_excel_date(date_raw)
                        if title and not check_training_exists(title, parsed_date.strftime("%d-%m-%Y")):
                            errors.append(f"Training '{title}' on {parsed_date.strftime('%d-%m-%Y')} not found in Trainings sheet or DB.")
                    except ValueError as e:
                        errors.append(str(e))

                parsed_enr_date = None
                if enr_date_raw:
                    try:
                        parsed_enr_date = parse_excel_date(enr_date_raw)
                    except ValueError as e:
                        errors.append(f"Enrollment Date: {str(e)}")

                status = "enrolled"
                if not status_str:
                    errors.append("Enrollment Status is mandatory.")
                else:
                    s_norm = status_str.lower()
                    if s_norm in ("enrolled", "approved", "completed", "withdrawn", "pending", "rejected"):
                        status = "enrolled" if s_norm == "approved" else s_norm
                    else:
                        errors.append(f"Invalid Enrollment Status: '{status_str}'.")

                is_valid = len(errors) == 0
                if is_valid:
                    valid_records += 1
                else:
                    failed_records += 1
                if warnings:
                    warnings_count += 1

                parsed_sheets["enrollments"].append({
                    "index": idx,
                    "is_valid": is_valid,
                    "errors": errors,
                    "warnings": warnings,
                    "is_duplicate": is_duplicate,
                    "data": {
                        "employee_id": emp_id,
                        "employee_name": emp_name,
                        "department": dept_str,
                        "training_name": title,
                        "training_date": parsed_date.strftime("%d-%m-%Y") if parsed_date else str(date_raw or ""),
                        "enrollment_date": parsed_enr_date.strftime("%d-%m-%Y") if parsed_enr_date else (parsed_date.strftime("%d-%m-%Y") if parsed_date else ""),
                        "enrollment_status": status
                    }
                })

        # ── Parse Attendance ──
        if "Attendance" in wb.sheetnames:
            ws = wb["Attendance"]
            for r in range(2, ws.max_row + 1):
                row_vals = [cell_val(ws, r, c) for c in range(1, 9)]
                if not any(v is not None and str(v).strip() != "" for v in row_vals):
                    continue
                records_found += 1
                idx = r - 1

                emp_id = str(row_vals[0] or "").strip()
                emp_name = str(row_vals[1] or "").strip()
                dept_str = str(row_vals[2] or "").strip()
                title = str(row_vals[3] or "").strip()
                date_raw = row_vals[4]
                status_str = str(row_vals[5] or "").strip()
                hours_raw = row_vals[6]
                marked_date_raw = row_vals[7]

                errors = []
                warnings = []
                is_duplicate = False

                if not emp_id:
                    errors.append("Employee ID is mandatory.")
                elif not check_employee_exists(emp_id):
                    errors.append(f"Employee ID '{emp_id}' not found in Employee sheet or DB.")

                if not title:
                    errors.append("Training Name is mandatory.")

                parsed_date = None
                if not date_raw:
                    errors.append("Training Date is mandatory.")
                else:
                    try:
                        parsed_date = parse_excel_date(date_raw)
                        if title and not check_training_exists(title, parsed_date.strftime("%d-%m-%Y")):
                            errors.append(f"Training '{title}' on {parsed_date.strftime('%d-%m-%Y')} not found in Trainings sheet or DB.")
                    except ValueError as e:
                        errors.append(str(e))

                status = "PRESENT"
                if not status_str:
                    errors.append("Attendance Status is mandatory.")
                else:
                    s_norm = status_str.lower()
                    if s_norm == "present":
                        status = "PRESENT"
                    elif s_norm == "absent":
                        status = "ABSENT"
                    elif s_norm == "late":
                        status = "LATE"
                    elif s_norm == "excused":
                        status = "ABSENT"
                    else:
                        errors.append(f"Invalid Attendance Status: '{status_str}'. Must be Present, Absent, Late, or Excused.")

                hours_earned = 0.0
                if hours_raw is not None and hours_raw != "":
                    try:
                        hours_earned = float(hours_raw)
                        if hours_earned < 0:
                            errors.append("Hours Earned cannot be negative.")
                    except (ValueError, TypeError):
                        errors.append("Hours Earned must be a number.")

                parsed_marked_date = None
                if marked_date_raw:
                    try:
                        parsed_marked_date = parse_excel_date(marked_date_raw)
                    except ValueError as e:
                        errors.append(f"Attendance Marked Date: {str(e)}")

                is_valid = len(errors) == 0
                if is_valid:
                    valid_records += 1
                else:
                    failed_records += 1
                if warnings:
                    warnings_count += 1

                parsed_sheets["attendance"].append({
                    "index": idx,
                    "is_valid": is_valid,
                    "errors": errors,
                    "warnings": warnings,
                    "is_duplicate": is_duplicate,
                    "data": {
                        "employee_id": emp_id,
                        "employee_name": emp_name,
                        "department": dept_str,
                        "training_name": title,
                        "training_date": parsed_date.strftime("%d-%m-%Y") if parsed_date else str(date_raw or ""),
                        "attendance_status": status_str,
                        "hours_earned": hours_earned,
                        "attendance_marked_date": parsed_marked_date.strftime("%d-%m-%Y") if parsed_marked_date else (parsed_date.strftime("%d-%m-%Y") if parsed_date else "")
                    }
                })

        # ── Parse Feedback ──
        if "Feedback" in wb.sheetnames:
            ws = wb["Feedback"]
            for r in range(2, ws.max_row + 1):
                row_vals = [cell_val(ws, r, c) for c in range(1, 9)]
                if not any(v is not None and str(v).strip() != "" for v in row_vals):
                    continue
                records_found += 1
                idx = r - 1

                emp_id = str(row_vals[0] or "").strip()
                emp_name = str(row_vals[1] or "").strip()
                title = str(row_vals[2] or "").strip()
                date_raw = row_vals[3]
                overall_raw = row_vals[4]
                trainer_raw = row_vals[5]
                content_raw = row_vals[6]
                comments = str(row_vals[7] or "").strip() if row_vals[7] is not None else ""

                errors = []
                warnings = []
                is_duplicate = False

                if not emp_id:
                    errors.append("Employee ID is mandatory.")
                elif not check_employee_exists(emp_id):
                    errors.append(f"Employee ID '{emp_id}' not found in Employee sheet or DB.")

                if not title:
                    errors.append("Training Name is mandatory.")

                parsed_date = None
                if not date_raw:
                    errors.append("Training Date is mandatory.")
                else:
                    try:
                        parsed_date = parse_excel_date(date_raw)
                        if title and not check_training_exists(title, parsed_date.strftime("%d-%m-%Y")):
                            errors.append(f"Training '{title}' on {parsed_date.strftime('%d-%m-%Y')} not found in Trainings sheet or DB.")
                    except ValueError as e:
                        errors.append(str(e))

                overall_rating = 5
                if overall_raw is not None and overall_raw != "":
                    try:
                        overall_rating = int(overall_raw)
                        if overall_rating < 1 or overall_rating > 5:
                            errors.append("Overall Rating must be between 1 and 5.")
                    except (ValueError, TypeError):
                        errors.append("Overall Rating must be an integer.")
                else:
                    errors.append("Overall Rating is mandatory.")

                trainer_rating = 5
                if trainer_raw is not None and trainer_raw != "":
                    try:
                        trainer_rating = int(trainer_raw)
                        if trainer_rating < 1 or trainer_rating > 5:
                            errors.append("Trainer Rating must be between 1 and 5.")
                    except (ValueError, TypeError):
                        errors.append("Trainer Rating must be an integer.")

                content_rating = 5
                if content_raw is not None and content_raw != "":
                    try:
                        content_rating = int(content_raw)
                        if content_rating < 1 or content_rating > 5:
                            errors.append("Content Rating must be between 1 and 5.")
                    except (ValueError, TypeError):
                        errors.append("Content Rating must be an integer.")

                is_valid = len(errors) == 0
                if is_valid:
                    valid_records += 1
                else:
                    failed_records += 1
                if warnings:
                    warnings_count += 1

                parsed_sheets["feedback"].append({
                    "index": idx,
                    "is_valid": is_valid,
                    "errors": errors,
                    "warnings": warnings,
                    "is_duplicate": is_duplicate,
                    "data": {
                        "employee_id": emp_id,
                        "employee_name": emp_name,
                        "training_name": title,
                        "training_date": parsed_date.strftime("%d-%m-%Y") if parsed_date else str(date_raw or ""),
                        "overall_rating": overall_rating,
                        "trainer_rating": trainer_rating,
                        "content_rating": content_rating,
                        "feedback_comments": comments
                    }
                })

        # ── Parse Learning_Hours ──
        if "Learning_Hours" in wb.sheetnames:
            ws = wb["Learning_Hours"]
            for r in range(2, ws.max_row + 1):
                row_vals = [cell_val(ws, r, c) for c in range(1, 8)]
                if not any(v is not None and str(v).strip() != "" for v in row_vals):
                    continue
                records_found += 1
                idx = r - 1

                emp_id = str(row_vals[0] or "").strip()
                emp_name = str(row_vals[1] or "").strip()
                dept_str = str(row_vals[2] or "").strip()
                title = str(row_vals[3] or "").strip()
                date_raw = row_vals[4]
                hours_raw = row_vals[5]
                fy_raw = str(row_vals[6] or "").strip() if row_vals[6] is not None else ""

                errors = []
                warnings = []
                is_duplicate = False

                if not emp_id:
                    errors.append("Employee ID is mandatory.")
                elif not check_employee_exists(emp_id):
                    errors.append(f"Employee ID '{emp_id}' not found in Employee sheet or DB.")

                if not title:
                    errors.append("Training Name is mandatory.")

                parsed_date = None
                if not date_raw:
                    errors.append("Training Date is mandatory.")
                else:
                    try:
                        parsed_date = parse_excel_date(date_raw)
                        if title and not check_training_exists(title, parsed_date.strftime("%d-%m-%Y")):
                            errors.append(f"Training '{title}' on {parsed_date.strftime('%d-%m-%Y')} not found in Trainings sheet or DB.")
                    except ValueError as e:
                        errors.append(str(e))

                hours = 0.0
                if hours_raw is not None and hours_raw != "":
                    try:
                        hours = float(hours_raw)
                        if hours <= 0:
                            errors.append("Learning Hours Earned must be greater than 0.")
                    except (ValueError, TypeError):
                        errors.append("Learning Hours Earned must be a number.")
                else:
                    errors.append("Learning Hours Earned is mandatory.")

                computed_fy = ""
                if parsed_date:
                    computed_fy = financial_year_of_date(parsed_date)

                is_valid = len(errors) == 0
                if is_valid:
                    valid_records += 1
                else:
                    failed_records += 1
                if warnings:
                    warnings_count += 1

                parsed_sheets["learning_hours"].append({
                    "index": idx,
                    "is_valid": is_valid,
                    "errors": errors,
                    "warnings": warnings,
                    "is_duplicate": is_duplicate,
                    "data": {
                        "employee_id": emp_id,
                        "employee_name": emp_name,
                        "department": dept_str,
                        "training_name": title,
                        "training_date": parsed_date.strftime("%d-%m-%Y") if parsed_date else str(date_raw or ""),
                        "learning_hours_earned": hours,
                        "financial_year": fy_raw or computed_fy
                    }
                })

        summary = {
            "records_found": records_found,
            "valid_records": valid_records,
            "warnings": warnings_count,
            "failed_records": failed_records
        }

        return {
            "summary": summary,
            "sheets": parsed_sheets
        }

    @staticmethod
    async def confirm_master_import(
        db: AsyncSession,
        sheets: dict,
        duplicate_strategy: str,
        created_by: UUID,
        source_file: Optional[str] = None
    ) -> dict:
        successfully_imported = 0
        skipped_duplicates = 0
        failed_records = 0

        sheet_breakdown = {
            "employees": {"imported": 0, "failed": 0, "skipped": 0},
            "trainings": {"imported": 0, "failed": 0, "skipped": 0},
            "enrollments": {"imported": 0, "failed": 0, "skipped": 0},
            "attendance": {"imported": 0, "failed": 0, "skipped": 0},
            "feedback": {"imported": 0, "failed": 0, "skipped": 0},
            "learning_hours": {"imported": 0, "failed": 0, "skipped": 0}
        }

        dept_stmt = select(Department)
        all_depts = (await db.execute(dept_stmt)).scalars().all()

        role_stmt = select(Role).where(func.lower(Role.name) == "employee")
        role_res = await db.execute(role_stmt)
        employee_role = role_res.scalar_one_or_none()
        if not employee_role:
            role_stmt_any = select(Role)
            employee_role = (await db.execute(role_stmt_any)).scalars().first()

        try:
            # ─────────────────────────────────────────────────────────────────
            # STEP 1: EMPLOYEES
            # ─────────────────────────────────────────────────────────────────
            emp_records = sheets.get("employees", [])
            for r in emp_records:
                if not r.get("is_valid", True):
                    failed_records += 1
                    sheet_breakdown["employees"]["failed"] += 1
                    continue

                d = r["data"]
                emp_code = d["employee_id"]
                full_name = d["employee_name"]
                dept_str = d["department"]
                designation = d["designation"]
                manager_code = d["manager"]
                doj_str = d["date_of_joining"]
                status = d["status"]

                parsed_doj = datetime.strptime(doj_str, "%d-%m-%Y").date() if doj_str else None
                names = full_name.split(" ", 1)
                first_name = names[0]
                last_name = names[1] if len(names) > 1 else ""

                aliases = {"it": "IT Support", "finance": "Accounts"}
                mapped_dept_name = aliases.get(dept_str.lower(), dept_str)
                matched_dept = None
                for dept_obj in all_depts:
                    if dept_obj.name.lower() == mapped_dept_name.lower() or dept_obj.code.strip().lower() == mapped_dept_name.lower():
                        matched_dept = dept_obj
                        break
                dept_id = matched_dept.id if matched_dept else None

                manager_id = None
                if manager_code:
                    mgr_stmt = select(Employee.id).where(Employee.employee_code == manager_code)
                    manager_id = (await db.execute(mgr_stmt)).scalar_one_or_none()

                existing_emp_stmt = select(Employee).where(Employee.employee_code == emp_code)
                existing_emp = (await db.execute(existing_emp_stmt)).scalar_one_or_none()

                if existing_emp:
                    if duplicate_strategy == "skip":
                        skipped_duplicates += 1
                        sheet_breakdown["employees"]["skipped"] += 1
                        continue
                    elif duplicate_strategy == "replace":
                        if existing_emp.user_id:
                            usr_stmt = select(User).where(User.id == existing_emp.user_id)
                            usr_obj = (await db.execute(usr_stmt)).scalar_one_or_none()
                            if usr_obj:
                                await db.delete(usr_obj)
                        await db.delete(existing_emp)
                        await db.flush()
                        existing_emp = None
                    elif duplicate_strategy == "update":
                        existing_emp.first_name = first_name
                        existing_emp.last_name = last_name
                        existing_emp.department_id = dept_id
                        existing_emp.designation = designation
                        existing_emp.manager_id = manager_id
                        existing_emp.date_of_joining = parsed_doj
                        existing_emp.status = EmploymentStatus(status)
                        await db.flush()
                        successfully_imported += 1
                        sheet_breakdown["employees"]["imported"] += 1

                if not existing_emp:
                    email = f"{emp_code.lower()}@synergyglobal.in"
                    hashed_pw = UserService.hash_password("Password123")
                    
                    user_obj = User(
                        email=email,
                        full_name=full_name,
                        hashed_password=hashed_pw,
                        role_id=employee_role.id if employee_role else None,
                        is_active=True,
                        is_verified=True
                    )
                    db.add(user_obj)
                    await db.flush()

                    new_emp = Employee(
                        employee_code=emp_code,
                        user_id=user_obj.id,
                        department_id=dept_id,
                        first_name=first_name,
                        last_name=last_name,
                        designation=designation,
                        email=email,
                        date_of_joining=parsed_doj,
                        status=EmploymentStatus(status),
                        manager_id=manager_id
                    )
                    db.add(new_emp)
                    await db.flush()
                    successfully_imported += 1
                    sheet_breakdown["employees"]["imported"] += 1

            # ─────────────────────────────────────────────────────────────────
            # STEP 2: TRAININGS
            # ─────────────────────────────────────────────────────────────────
            train_records = sheets.get("trainings", [])
            for r in train_records:
                if not r.get("is_valid", True):
                    failed_records += 1
                    sheet_breakdown["trainings"]["failed"] += 1
                    continue

                d = r["data"]
                title = d["title"]
                desc = d["description"]
                dept_str = d["department"]
                cat_str = d["training_category"]
                trainer = d["trainer_name"]
                date_str = d["training_date"]
                start_time = d["start_time"]
                end_time = d["end_time"]
                duration_hours = d["duration_hours"]
                mode = d["mode"]
                venue = d["venue"]
                max_seats = d["max_seats"]
                deadline_date_str = d["enrollment_deadline_date"]
                deadline_time_str = d["enrollment_deadline_time"]
                training_type = d["training_type"]
                status = d["status"]
                is_mandatory = d["is_mandatory"]
                material_title = d.get("material_title")
                material_link = d.get("material_link")

                parsed_date = datetime.strptime(date_str, "%d-%m-%Y").date()
                parsed_start_time = datetime.strptime(start_time, "%H:%M").time()
                db_start_time = parsed_start_time.strftime("%I:%M %p")

                category_id = None
                if cat_str:
                    cat_stmt = select(TrainingCategory).where(func.lower(TrainingCategory.name) == cat_str.lower())
                    cat_res = await db.execute(cat_stmt)
                    cat_obj = cat_res.scalar_one_or_none()
                    if not cat_obj:
                        cat_obj = TrainingCategory(name=cat_str, description="Autocreated from import")
                        db.add(cat_obj)
                        await db.flush()
                    category_id = cat_obj.id

                today = date.today()
                if parsed_date < today:
                    db_status = TrainingStatus.COMPLETED
                else:
                    db_status = TrainingStatus[status]

                is_global = dept_str.lower() in ("all", "global", "all departments")
                dept_ids = []
                if not is_global:
                    depts_split = [dep.strip() for dep in re.split(r'[,;]', dept_str) if dep.strip()]
                    for d_name in depts_split:
                        aliases = {"it": "IT Support", "finance": "Accounts"}
                        mapped_dept_name = aliases.get(d_name.lower(), d_name)
                        matched_dept = None
                        for dept_obj in all_depts:
                            if dept_obj.name.lower() == mapped_dept_name.lower() or dept_obj.code.strip().lower() == mapped_dept_name.lower():
                                matched_dept = dept_obj
                                break
                        if matched_dept:
                            dept_ids.append(matched_dept.id)

                dup_stmt = select(Training).options(selectinload(Training.departments)).where(
                    Training.title == title,
                    Training.start_date == parsed_date
                )
                dup_res = await db.execute(dup_stmt)
                existing_train = dup_res.scalar_one_or_none()

                if existing_train:
                    if duplicate_strategy == "skip":
                        skipped_duplicates += 1
                        sheet_breakdown["trainings"]["skipped"] += 1
                        continue
                    elif duplicate_strategy == "replace":
                        await db.delete(existing_train)
                        await db.flush()
                        existing_train = None
                    elif duplicate_strategy == "update":
                        existing_train.description = desc
                        existing_train.training_type = TrainingType[training_type]
                        existing_train.delivery_mode = DeliveryMode[mode]
                        existing_train.status = db_status
                        existing_train.duration_hours = duration_hours
                        existing_train.max_hours_allowed = duration_hours
                        existing_train.max_participants = max_seats
                        existing_train.venue = venue
                        existing_train.trainer_name = trainer
                        existing_train.is_mandatory = is_mandatory
                        existing_train.is_global = is_global
                        existing_train.category_id = category_id
                        
                        if deadline_date_str:
                            d_date = datetime.strptime(deadline_date_str, "%d-%m-%Y").date()
                            d_time = datetime.strptime(deadline_time_str or "23:59", "%H:%M").time()
                            existing_train.enrollment_deadline = datetime.combine(d_date, d_time)
                        else:
                            existing_train.enrollment_deadline = None

                        if is_global:
                            existing_train.departments = []
                        else:
                            dept_stmt = select(Department).where(Department.id.in_(dept_ids))
                            existing_train.departments = list((await db.execute(dept_stmt)).scalars().all())

                        await db.flush()
                        successfully_imported += 1
                        sheet_breakdown["trainings"]["imported"] += 1

                if not existing_train:
                    new_train = Training(
                        title=title,
                        description=desc,
                        training_type=TrainingType[training_type],
                        delivery_mode=DeliveryMode[mode],
                        status=db_status,
                        start_date=parsed_date,
                        start_time=db_start_time,
                        duration_hours=duration_hours,
                        max_hours_allowed=duration_hours,
                        max_participants=max_seats,
                        available_seats=max_seats,
                        venue=venue,
                        trainer_name=trainer,
                        is_mandatory=is_mandatory,
                        is_global=is_global,
                        category_id=category_id,
                        created_by=created_by
                    )

                    if deadline_date_str:
                        d_date = datetime.strptime(deadline_date_str, "%d-%m-%Y").date()
                        d_time = datetime.strptime(deadline_time_str or "23:59", "%H:%M").time()
                        new_train.enrollment_deadline = datetime.combine(d_date, d_time)

                    if not is_global and dept_ids:
                        dept_stmt = select(Department).where(Department.id.in_(dept_ids))
                        new_train.departments = list((await db.execute(dept_stmt)).scalars().all())

                    db.add(new_train)
                    await db.flush()
                    successfully_imported += 1
                    sheet_breakdown["trainings"]["imported"] += 1

                # ── E-Learning Module Auto-creation & Material Linking ──
                learning_category = None
                if cat_str:
                    from app.learning_hub.models import LearningCategory, LearningModule, LearningMaterial
                    cat_stmt = select(LearningCategory).where(func.lower(LearningCategory.name) == cat_str.lower())
                    learning_category = (await db.execute(cat_stmt)).scalar_one_or_none()
                    if not learning_category:
                        learning_category = LearningCategory(name=cat_str, description=f"Category for {cat_str} courses")
                        db.add(learning_category)
                        await db.flush()

                # Find or Create LearningModule
                target_train_id = existing_train.id if existing_train else new_train.id
                module_stmt = select(LearningModule).where(
                    or_(
                        LearningModule.training_id == target_train_id,
                        func.lower(LearningModule.title) == title.lower()
                    )
                )
                learning_module = (await db.execute(module_stmt)).scalars().first()

                if not learning_module:
                    from app.learning_hub.models import LearningModule
                    learning_module = LearningModule(
                        title=title,
                        description=desc or f"Learning resources for {title}",
                        category_id=learning_category.id if learning_category else None,
                        department_id=dept_ids[0] if (not is_global and dept_ids) else None,
                        training_id=target_train_id,
                        created_by=created_by
                    )
                    db.add(learning_module)
                    await db.flush()
                else:
                    if not learning_module.training_id:
                        learning_module.training_id = target_train_id
                    if learning_category and not learning_module.category_id:
                        learning_module.category_id = learning_category.id
                    await db.flush()

                # Create LearningMaterial if provided
                if material_title or material_link:
                    from app.learning_hub.models import LearningMaterial
                    mat_title = material_title or "Course Reference Material"
                    mat_stmt = select(LearningMaterial).where(
                        LearningMaterial.module_id == learning_module.id,
                        or_(
                            func.lower(LearningMaterial.title) == mat_title.lower(),
                            LearningMaterial.external_url == material_link
                        )
                    )
                    material = (await db.execute(mat_stmt)).scalars().first()
                    if not material:
                        material = LearningMaterial(
                            module_id=learning_module.id,
                            title=mat_title,
                            description=f"Auto-imported material during training workbook ingestion",
                            external_url=material_link,
                            file_type="link" if material_link else "document",
                            is_approved=True,
                            created_by=created_by
                        )
                        db.add(material)
                        await db.flush()

            # ─────────────────────────────────────────────────────────────────
            # STEP 3: ENROLLMENTS
            # ─────────────────────────────────────────────────────────────────
            enr_records = sheets.get("enrollments", [])
            for r in enr_records:
                if not r.get("is_valid", True):
                    failed_records += 1
                    sheet_breakdown["enrollments"]["failed"] += 1
                    continue

                d = r["data"]
                emp_code = d["employee_id"]
                title = d["training_name"]
                t_date_str = d["training_date"]
                enr_date_str = d["enrollment_date"]
                status = d["enrollment_status"]

                t_date = datetime.strptime(t_date_str, "%d-%m-%Y").date()
                enr_date = datetime.strptime(enr_date_str, "%d-%m-%Y").date() if enr_date_str else None

                emp_stmt = select(Employee).where(Employee.employee_code == emp_code)
                employee = (await db.execute(emp_stmt)).scalar_one_or_none()
                
                train_stmt = select(Training).where(Training.title == title, Training.start_date == t_date)
                training = (await db.execute(train_stmt)).scalar_one_or_none()

                if not employee or not training:
                    failed_records += 1
                    sheet_breakdown["enrollments"]["failed"] += 1
                    continue

                dup_stmt = select(Enrollment).where(
                    Enrollment.employee_id == employee.id,
                    Enrollment.training_id == training.id
                )
                existing_enr = (await db.execute(dup_stmt)).scalar_one_or_none()

                db_status = EnrollmentStatus(status)
                db_progress = 100.0 if db_status == EnrollmentStatus.COMPLETED else 0.0

                if existing_enr:
                    if duplicate_strategy == "skip":
                        skipped_duplicates += 1
                        sheet_breakdown["enrollments"]["skipped"] += 1
                        continue
                    elif duplicate_strategy == "replace":
                        await db.delete(existing_enr)
                        await db.flush()
                        existing_enr = None
                    elif duplicate_strategy == "update":
                        existing_enr.status = db_status
                        existing_enr.progress = db_progress
                        if enr_date:
                            existing_enr.created_at = datetime.combine(enr_date, time(0, 0))
                        await db.flush()
                        successfully_imported += 1
                        sheet_breakdown["enrollments"]["imported"] += 1

                if not existing_enr:
                    new_enr = Enrollment(
                        employee_id=employee.id,
                        training_id=training.id,
                        status=db_status,
                        progress=db_progress
                    )
                    if enr_date:
                        new_enr.created_at = datetime.combine(enr_date, time(0, 0))
                    db.add(new_enr)
                    await db.flush()
                    successfully_imported += 1
                    sheet_breakdown["enrollments"]["imported"] += 1

            # ─────────────────────────────────────────────────────────────────
            # STEP 4: ATTENDANCE
            # ─────────────────────────────────────────────────────────────────
            att_records = sheets.get("attendance", [])
            for r in att_records:
                if not r.get("is_valid", True):
                    failed_records += 1
                    sheet_breakdown["attendance"]["failed"] += 1
                    continue

                d = r["data"]
                emp_code = d["employee_id"]
                title = d["training_name"]
                t_date_str = d["training_date"]
                att_status_excel = d["attendance_status"]
                hours_earned = d["hours_earned"]
                marked_date_str = d["attendance_marked_date"]

                t_date = datetime.strptime(t_date_str, "%d-%m-%Y").date()
                marked_date = datetime.strptime(marked_date_str, "%d-%m-%Y").date() if marked_date_str else t_date

                emp_stmt = select(Employee).where(Employee.employee_code == emp_code)
                employee = (await db.execute(emp_stmt)).scalar_one_or_none()

                train_stmt = select(Training).where(Training.title == title, Training.start_date == t_date)
                training = (await db.execute(train_stmt)).scalar_one_or_none()

                if not employee or not training:
                    failed_records += 1
                    sheet_breakdown["attendance"]["failed"] += 1
                    continue

                session = await AttendanceService.ensure_session(db, training.id)

                norm_status = att_status_excel.lower()
                status = AttendanceStatus.PRESENT
                remarks = None
                if norm_status == "present":
                    status = AttendanceStatus.PRESENT
                elif norm_status == "absent":
                    status = AttendanceStatus.ABSENT
                elif norm_status == "late":
                    status = AttendanceStatus.LATE
                elif norm_status == "excused":
                    status = AttendanceStatus.ABSENT
                    remarks = "Excused absence"

                dup_stmt = select(AttendanceRecord).where(
                    AttendanceRecord.employee_id == employee.id,
                    AttendanceRecord.training_id == training.id
                )
                existing_att = (await db.execute(dup_stmt)).scalar_one_or_none()

                if existing_att:
                    if duplicate_strategy == "skip":
                        skipped_duplicates += 1
                        sheet_breakdown["attendance"]["skipped"] += 1
                        continue
                    elif duplicate_strategy == "replace":
                        await db.delete(existing_att)
                        await db.flush()
                        existing_att = None
                    elif duplicate_strategy == "update":
                        existing_att.status = status
                        existing_att.marked_at = datetime.combine(marked_date, time(12, 0))
                        existing_att.session_id = session.id
                        existing_att.remarks = f"Roster submission - Imported: {remarks}" if remarks else "Roster submission - Imported"
                        await db.flush()
                        successfully_imported += 1
                        sheet_breakdown["attendance"]["imported"] += 1

                if not existing_att:
                    new_att = AttendanceRecord(
                        employee_id=employee.id,
                        training_id=training.id,
                        session_id=session.id,
                        status=status,
                        marked_at=datetime.combine(marked_date, time(12, 0)),
                        attendance_open_time=session.opens_at,
                        attendance_close_time=session.closes_at,
                        remarks=f"Roster submission - Imported: {remarks}" if remarks else "Roster submission - Imported"
                    )
                    db.add(new_att)
                    await db.flush()
                    successfully_imported += 1
                    sheet_breakdown["attendance"]["imported"] += 1

                # Check / Create Enrollment
                enr_stmt = select(Enrollment).where(
                    Enrollment.employee_id == employee.id,
                    Enrollment.training_id == training.id
                )
                enrollment = (await db.execute(enr_stmt)).scalar_one_or_none()
                if not enrollment:
                    if status in (AttendanceStatus.PRESENT, AttendanceStatus.LATE):
                        enrollment = Enrollment(
                            employee_id=employee.id,
                            training_id=training.id,
                            status=EnrollmentStatus.COMPLETED,
                            progress=100.0
                        )
                    else:
                        enrollment = Enrollment(
                            employee_id=employee.id,
                            training_id=training.id,
                            status=EnrollmentStatus.APPROVED,
                            progress=0.0
                        )
                    db.add(enrollment)
                    await db.flush()
                else:
                    if status in (AttendanceStatus.PRESENT, AttendanceStatus.LATE):
                        if enrollment.status != EnrollmentStatus.COMPLETED:
                            enrollment.status = EnrollmentStatus.COMPLETED
                            enrollment.progress = 100.0
                            await db.flush()

            # ─────────────────────────────────────────────────────────────────
            # STEP 5: FEEDBACK
            # ─────────────────────────────────────────────────────────────────
            feed_records = sheets.get("feedback", [])
            for r in feed_records:
                if not r.get("is_valid", True):
                    failed_records += 1
                    sheet_breakdown["feedback"]["failed"] += 1
                    continue

                d = r["data"]
                emp_code = d["employee_id"]
                title = d["training_name"]
                t_date_str = d["training_date"]
                overall_rating = d["overall_rating"]
                trainer_rating = d["trainer_rating"]
                content_rating = d["content_rating"]
                comments = d["feedback_comments"]

                t_date = datetime.strptime(t_date_str, "%d-%m-%Y").date()

                emp_stmt = select(Employee).where(Employee.employee_code == emp_code)
                employee = (await db.execute(emp_stmt)).scalar_one_or_none()

                train_stmt = select(Training).where(Training.title == title, Training.start_date == t_date)
                training = (await db.execute(train_stmt)).scalar_one_or_none()

                if not employee or not training:
                    failed_records += 1
                    sheet_breakdown["feedback"]["failed"] += 1
                    continue

                enr_stmt = select(Enrollment).where(
                    Enrollment.employee_id == employee.id,
                    Enrollment.training_id == training.id
                )
                enrollment = (await db.execute(enr_stmt)).scalar_one_or_none()

                if not enrollment:
                    enrollment = Enrollment(
                        employee_id=employee.id,
                        training_id=training.id,
                        status=EnrollmentStatus.COMPLETED,
                        progress=100.0
                    )
                    db.add(enrollment)
                    await db.flush()

                eff_stmt = select(Effectiveness).where(
                    Effectiveness.enrollment_id == enrollment.id,
                    Effectiveness.training_id == training.id
                )
                existing_eff = (await db.execute(eff_stmt)).scalar_one_or_none()

                full_comments = comments
                if trainer_rating or content_rating:
                    addon = []
                    if trainer_rating:
                        addon.append(f"Trainer Rating: {trainer_rating}/5")
                    if content_rating:
                        addon.append(f"Content Rating: {content_rating}/5")
                    full_comments = f"{comments} ({', '.join(addon)})" if comments else f"({', '.join(addon)})"

                if existing_eff:
                    if duplicate_strategy == "skip":
                        skipped_duplicates += 1
                        sheet_breakdown["feedback"]["skipped"] += 1
                        continue
                    elif duplicate_strategy == "replace":
                        await db.delete(existing_eff)
                        await db.flush()
                        existing_eff = None
                    elif duplicate_strategy == "update":
                        existing_eff.rating = overall_rating
                        existing_eff.comments = full_comments
                        existing_eff.status = EffectivenessStatus.SUBMITTED
                        existing_eff.completion_datetime = datetime.now()
                        await db.flush()
                        successfully_imported += 1
                        sheet_breakdown["feedback"]["imported"] += 1

                if not existing_eff:
                    new_eff = Effectiveness(
                        enrollment_id=enrollment.id,
                        training_id=training.id,
                        level=EffectivenessLevel.REACTION,
                        rating=overall_rating,
                        comments=full_comments,
                        status=EffectivenessStatus.SUBMITTED,
                        completion_datetime=datetime.now()
                    )
                    db.add(new_eff)
                    await db.flush()
                    successfully_imported += 1
                    sheet_breakdown["feedback"]["imported"] += 1

            # ─────────────────────────────────────────────────────────────────
            # STEP 6: LEARNING HOURS
            # ─────────────────────────────────────────────────────────────────
            hours_records = sheets.get("learning_hours", [])
            for r in hours_records:
                if not r.get("is_valid", True):
                    failed_records += 1
                    sheet_breakdown["learning_hours"]["failed"] += 1
                    continue

                d = r["data"]
                emp_code = d["employee_id"]
                title = d["training_name"]
                t_date_str = d["training_date"]
                learning_hours = d["learning_hours_earned"]

                t_date = datetime.strptime(t_date_str, "%d-%m-%Y").date()

                emp_stmt = select(Employee).where(Employee.employee_code == emp_code)
                employee = (await db.execute(emp_stmt)).scalar_one_or_none()

                train_stmt = select(Training).where(Training.title == title, Training.start_date == t_date)
                training = (await db.execute(train_stmt)).scalar_one_or_none()

                if not employee or not training:
                    failed_records += 1
                    sheet_breakdown["learning_hours"]["failed"] += 1
                    continue

                training.duration_hours = learning_hours
                training.max_hours_allowed = learning_hours
                await db.flush()

                enr_stmt = select(Enrollment).where(
                    Enrollment.employee_id == employee.id,
                    Enrollment.training_id == training.id
                )
                enrollment = (await db.execute(enr_stmt)).scalar_one_or_none()

                if not enrollment:
                    enrollment = Enrollment(
                        employee_id=employee.id,
                        training_id=training.id,
                        status=EnrollmentStatus.COMPLETED,
                        progress=100.0
                    )
                    db.add(enrollment)
                    await db.flush()
                else:
                    enrollment.status = EnrollmentStatus.COMPLETED
                    enrollment.progress = 100.0
                    await db.flush()

                att_stmt = select(AttendanceRecord).where(
                    AttendanceRecord.employee_id == employee.id,
                    AttendanceRecord.training_id == training.id
                )
                attendance = (await db.execute(att_stmt)).scalar_one_or_none()

                if not attendance:
                    session = await AttendanceService.ensure_session(db, training.id)
                    attendance = AttendanceRecord(
                        employee_id=employee.id,
                        training_id=training.id,
                        session_id=session.id,
                        status=AttendanceStatus.PRESENT,
                        marked_at=datetime.combine(t_date, time(12, 0)),
                        attendance_open_time=session.opens_at,
                        attendance_close_time=session.closes_at,
                        remarks="Roster submission - Imported"
                    )
                    db.add(attendance)
                    await db.flush()
                else:
                    attendance.status = AttendanceStatus.PRESENT
                    if not attendance.remarks or not attendance.remarks.startswith("Roster submission"):
                        attendance.remarks = "Roster submission - Imported"
                    await db.flush()

                successfully_imported += 1
                sheet_breakdown["learning_hours"]["imported"] += 1

            # Log import audit trail record
            history_record = TrainingImportHistory(
                created_by=created_by,
                records_imported=successfully_imported,
                records_failed=failed_records,
                records_skipped=skipped_duplicates,
                source_file=source_file or "Master Import Workbook"
            )
            db.add(history_record)
            await db.commit()

            return {
                "successfully_imported": successfully_imported,
                "failed_records": failed_records,
                "skipped_duplicates": skipped_duplicates,
                "sheet_breakdown": sheet_breakdown
            }

        except Exception as e:
            await db.rollback()
            raise e
