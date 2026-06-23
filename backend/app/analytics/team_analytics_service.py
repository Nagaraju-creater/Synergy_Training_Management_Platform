from datetime import date, datetime, timedelta
from typing import List, Optional
import io
import uuid
from sqlalchemy import func, select, and_, or_, cast, String, case
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.analytics.team_analytics_schemas import (
    TeamKPIs, DeptLearningHours, MonthlyLearningTrend,
    CourseParticipation, SkillGapItem, EffectivenessScore,
    DepartmentSummary, EmployeeAnalyticsRow, TopLearner,
    TeamAnalyticsDashboard,
)
from app.employees.models import Employee
from app.enrollments.models import Enrollment, EnrollmentStatus
from app.trainings.models import Training, TrainingStatus
from app.trainings.categories import TrainingCategory
from app.departments.models import Department
from app.effectiveness.models import Effectiveness, EffectivenessStatus
from app.nominations.models import Nomination
from app.attendance.models import AttendanceRecord, AttendanceStatus
from app.attendance.service import AttendanceService


class TeamAnalyticsService:
    LEARNING_TARGET_HOURS = 16.0

    @staticmethod
    def financial_year_bounds(financial_year: Optional[str] = None) -> tuple[date, date, str]:
        if financial_year:
            try:
                start_year = int(financial_year.split("-")[0])
            except (ValueError, TypeError):
                start_year = date.today().year if date.today().month >= 4 else date.today().year - 1
        else:
            start_year = date.today().year if date.today().month >= 4 else date.today().year - 1
        return date(start_year, 4, 1), date(start_year + 1, 3, 31), f"{start_year}-{start_year + 1}"

    @staticmethod
    def compliance_status(kpi_pct: float) -> str:
        if kpi_pct >= 100:
            return "Achieved"
        if kpi_pct >= 75:
            return "Good"
        if kpi_pct >= 50:
            return "Moderate"
        if kpi_pct >= 25:
            return "Low"
        return "Critical"

    @staticmethod
    def fy_months(fy_start: date) -> list[tuple[int, int, str]]:
        months = []
        for offset in range(12):
            month = ((fy_start.month + offset - 1) % 12) + 1
            year = fy_start.year + ((fy_start.month + offset - 1) // 12)
            months.append((year, month, date(year, month, 1).strftime("%B")))
        return months

    @staticmethod
    def parse_uuid(value: Optional[str]):
        if not value:
            return None
        try:
            return uuid.UUID(str(value))
        except (ValueError, TypeError):
            return None

    @staticmethod
    async def generate_kpi_excel(
        db: AsyncSession,
        financial_year: Optional[str] = None,
        month: Optional[int] = None,
        department_id: Optional[str] = None,
        employee_id: Optional[str] = None,
        manager_id: Optional[str] = None,
        training_id: Optional[str] = None,
        training_category_id: Optional[str] = None,
        attendance_status: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        # We will mimic the inputs of generate_kpi_excel

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
        from openpyxl.formatting.rule import CellIsRule


        # 1. Determine the bounds
        from app.analytics.team_analytics_service import TeamAnalyticsService
        fy_start, fy_end, fy_label = TeamAnalyticsService.financial_year_bounds(financial_year)
        range_start = start_date or fy_start
        range_end = end_date or fy_end
        months = TeamAnalyticsService.fy_months(fy_start)
        if month:
            months = [m for m in months if m[1] == month]

        dept_uuid = TeamAnalyticsService.parse_uuid(department_id)
        emp_uuid = TeamAnalyticsService.parse_uuid(employee_id)
        mgr_uuid = TeamAnalyticsService.parse_uuid(manager_id)
        training_uuid = TeamAnalyticsService.parse_uuid(training_id)
        category_uuid = TeamAnalyticsService.parse_uuid(training_category_id)

        # 2. Query metadata
        depts_stmt = select(Department).where(Department.deleted_at == None).order_by(Department.name.asc())
        depts_res = (await db.execute(depts_stmt)).scalars().all()
        dept_names = [d.name for d in depts_res]

        cats_stmt = select(TrainingCategory).order_by(TrainingCategory.name.asc())
        cats_res = (await db.execute(cats_stmt)).scalars().all()
        cat_names = [c.name for c in cats_res]

        mgrs_stmt = select(Employee).where(Employee.id.in_(select(Employee.manager_id).where(Employee.manager_id != None))).order_by(Employee.first_name.asc())
        mgrs_res = (await db.execute(mgrs_stmt)).scalars().all()
        mgr_names = [f"{m.first_name} {m.last_name}" for m in mgrs_res]

        training_types = ["INTERNAL", "EXTERNAL", "ONLINE", "WORKSHOP", "CERTIFICATION"]

        selected_fy = "All"
        selected_dept_name = "All"
        selected_cat_name = "All"
        selected_month_name = "All"

        if financial_year:
            selected_fy = financial_year
        if month and 1 <= month <= 12:
            month_map = {
                4: "April", 5: "May", 6: "June", 7: "July", 8: "August", 9: "September",
                10: "October", 11: "November", 12: "December", 1: "January", 2: "February", 3: "March"
            }
            if month in month_map:
                selected_month_name = month_map[month]

        if department_id:
            dept_obj = next((d for d in depts_res if d.id == dept_uuid), None)
            if dept_obj:
                selected_dept_name = dept_obj.name

        if training_category_id:
            cat_obj = next((c for c in cats_res if c.id == category_uuid), None)
            if cat_obj:
                selected_cat_name = cat_obj.name

        # manager_id filter selection resolved at database query layer

        # 3. Query Set A (unfiltered active employees and all attendance records for selected FY)
        all_employees_stmt = (
            select(Employee)
            .options(selectinload(Employee.department), selectinload(Employee.manager))
            .where(Employee.deleted_at == None)
            .order_by(Employee.employee_code.asc())
        )
        all_employees = (await db.execute(all_employees_stmt)).scalars().all()

        all_attendance_stmt = (
            select(AttendanceRecord, Training, Employee, Enrollment)
            .join(Training, AttendanceRecord.training_id == Training.id)
            .join(Employee, AttendanceRecord.employee_id == Employee.id)
            .join(Enrollment, and_(
                Enrollment.employee_id == AttendanceRecord.employee_id,
                Enrollment.training_id == AttendanceRecord.training_id,
                Enrollment.deleted_at == None,
            ))
            .options(
                selectinload(Employee.department),
                selectinload(Employee.manager),
                selectinload(Training.category)
            )
            .where(and_(
                Employee.deleted_at == None,
                AttendanceService.roster_record_filter(),
                Training.status != TrainingStatus.CANCELLED,
                Training.start_date >= fy_start,
                Training.start_date <= fy_end,
            ))
            .order_by(Training.start_date.asc())
        )
        all_attendance_results = (await db.execute(all_attendance_stmt)).all()

        # 4. Query Set B (filtered employees and attendance records matching query parameters)
        filtered_employees_stmt = (
            select(Employee)
            .options(selectinload(Employee.department), selectinload(Employee.manager))
            .where(Employee.deleted_at == None)
            .order_by(Employee.employee_code.asc())
        )
        if dept_uuid:
            filtered_employees_stmt = filtered_employees_stmt.where(Employee.department_id == dept_uuid)
        if emp_uuid:
            filtered_employees_stmt = filtered_employees_stmt.where(Employee.id == emp_uuid)
        if mgr_uuid:
            filtered_employees_stmt = filtered_employees_stmt.where(Employee.manager_id == mgr_uuid)
        
        filtered_employees = (await db.execute(filtered_employees_stmt)).scalars().all()
        filtered_employee_ids = [emp.id for emp in filtered_employees]

        filtered_attendance_results = []
        if filtered_employee_ids:
            stmt = (
                select(AttendanceRecord, Training, Employee, Enrollment)
                .join(Training, AttendanceRecord.training_id == Training.id)
                .join(Employee, AttendanceRecord.employee_id == Employee.id)
                .join(Enrollment, and_(
                    Enrollment.employee_id == AttendanceRecord.employee_id,
                    Enrollment.training_id == AttendanceRecord.training_id,
                    Enrollment.deleted_at == None,
                ))
                .options(
                    selectinload(Employee.department),
                    selectinload(Employee.manager),
                    selectinload(Training.category)
                )
                .where(and_(
                    AttendanceRecord.employee_id.in_(filtered_employee_ids),
                    AttendanceService.roster_record_filter(),
                    Training.status != TrainingStatus.CANCELLED,
                    Training.start_date >= range_start,
                    Training.start_date <= range_end,
                ))
            )
            if training_uuid:
                stmt = stmt.where(Training.id == training_uuid)
            if category_uuid:
                stmt = stmt.where(Training.category_id == category_uuid)
            if attendance_status:
                stmt = stmt.where(AttendanceRecord.status == attendance_status)
            
            filtered_attendance_results = (await db.execute(stmt.order_by(Training.start_date.asc(), Training.title.asc()))).all()

        # Build contribution map for Sheet 1
        contribution_map = {emp.id: {} for emp in filtered_employees}
        attendance_rows = []
        for record, training, employee, enrollment in filtered_attendance_results:
            key = (training.start_date.year, training.start_date.month) if training.start_date else None
            
            att_status_val = record.status.value if hasattr(record.status, "value") else str(record.status)
            
            contribution = {
                "employee_id": employee.id,
                "employee_name": f"{employee.first_name} {employee.last_name}",
                "employee_code": employee.employee_code,
                "department": employee.department.name if employee.department else "Unassigned",
                "manager": f"{employee.manager.first_name} {employee.manager.last_name}" if employee.manager else "",
                "training_id": training.id,
                "training_title": training.title,
                "training_date": training.start_date,
                "hours": float(training.duration_hours or 0),
                "max_hours": float(training.max_hours_allowed or training.duration_hours or 0),
                "status": att_status_val,
                "marked_at": record.marked_at,
                "category": training.category.name if training.category else "",
                "trainer": training.trainer_name or "",
                "duration": float(training.duration_hours or 0),
            }
            
            if att_status_val in ("PRESENT", "LATE") and enrollment.status == EnrollmentStatus.COMPLETED:
                contribution_map.setdefault(employee.id, {}).setdefault(key, []).append(contribution)
            
            attendance_rows.append(contribution)

        # 5. Initialize openpyxl workbook
        wb = openpyxl.Workbook()
        
        emp_report_ws = wb.active
        emp_report_ws.title = "Employee KPI Report"
        emp_report_ws.views.sheetView[0].showGridLines = True
        
        dept_ws = wb.create_sheet("Department KPI Analytics")
        dept_ws.views.sheetView[0].showGridLines = True
        
        month_ws = wb.create_sheet("Monthly Learning Summary")
        month_ws.views.sheetView[0].showGridLines = True
        
        att_ws = wb.create_sheet("Training Attendance Summary")
        att_ws.views.sheetView[0].showGridLines = True
        
        raw_ws = wb.create_sheet("Raw_Attendance_Data")
        raw_ws.views.sheetView[0].showGridLines = True
        raw_ws.sheet_state = "hidden"
        
        emp_summary_ws = wb.create_sheet("Employee_KPI_Summary")
        emp_summary_ws.views.sheetView[0].showGridLines = True
        emp_summary_ws.sheet_state = "hidden"
        
        list_ws = wb.create_sheet("Filter_Lists")
        list_ws.sheet_state = "hidden"

        # 6. Styling Helper
        blue = "4F81BD"
        deep_blue = "1F4E79"
        light_blue = "D9EAF7"
        green = "E2F0D9"
        yellow = "FFF2CC"
        red = "FCE4D6"
        white = "FFFFFF"
        thin = Side(style="thin", color="9EADCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        double_bottom_side = Side(style="double", color="1F4E79")
        double_bottom_border = Border(left=thin, right=thin, top=thin, bottom=double_bottom_side)

        def style_cell(cell, fill=white, font_color="1F2937", bold=False, align="center", size=9):
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Calibri", size=size, bold=bold, color=font_color)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = border

        # 7. Populate Sheet 1: Employee KPI Report
        total_cols = 4 + len(months) * 3 + 8
        emp_report_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        emp_report_ws.cell(1, 1, "Department KPI Dashboard - Employee Learning KPI Report")
        style_cell(emp_report_ws.cell(1, 1), fill=deep_blue, font_color=white, bold=True, size=14)
        emp_report_ws.row_dimensions[1].height = 30
        
        emp_report_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        filter_summary = f"FY: Apr {fy_start.year} - Mar {fy_end.year} | Date Range: {range_start:%d-%b-%Y} to {range_end:%d-%b-%Y}"
        if month:
            filter_summary += f" | Month: {date(fy_start.year if month >= 4 else fy_end.year, month, 1):%B}"
        emp_report_ws.cell(2, 1, filter_summary)
        style_cell(emp_report_ws.cell(2, 1), fill=light_blue, bold=True, align="left", size=10)
        emp_report_ws.row_dimensions[2].height = 24

        base_headers = ["S.No", "Employee ID", "Employee Name", "Date of Joining"]
        for idx, header in enumerate(base_headers, 1):
            emp_report_ws.merge_cells(start_row=4, start_column=idx, end_row=5, end_column=idx)
            emp_report_ws.cell(4, idx, header)
            style_cell(emp_report_ws.cell(4, idx), fill=blue, font_color=white, bold=True, size=9)
            
        col_idx = 5
        for _, _, month_name in months:
            emp_report_ws.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx + 2)
            emp_report_ws.cell(4, col_idx, month_name.upper())
            style_cell(emp_report_ws.cell(4, col_idx), fill=blue, font_color=white, bold=True, size=9)
            for sub in ["Date", "Actual", "Max Hours"]:
                emp_report_ws.cell(5, col_idx, sub)
                style_cell(emp_report_ws.cell(5, col_idx), fill=blue, font_color=white, bold=True, size=9)
                col_idx += 1
                
        final_headers = ["Total Hours", "Target Hours", "Balance", "KPI %", "Learning Compliance", "Department", "Manager", "Attendance %"]
        for header in final_headers:
            emp_report_ws.merge_cells(start_row=4, start_column=col_idx, end_row=5, end_column=col_idx)
            emp_report_ws.cell(4, col_idx, header)
            style_cell(emp_report_ws.cell(4, col_idx), fill=blue, font_color=white, bold=True, size=9)
            col_idx += 1
            
        emp_report_ws.row_dimensions[4].height = 20
        emp_report_ws.row_dimensions[5].height = 20

        sorted_employees = sorted(filtered_employees, key=lambda e: e.department.name if e.department else "Unassigned")
        from itertools import groupby
        def get_dept_name(e):
            return e.department.name if e.department else "Unassigned"
            
        current_row = 6
        subtotal_rows = []
        s_no_counter = 1
        
        if len(sorted_employees) == 0:
            emp_report_ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=total_cols)
            emp_report_ws.cell(6, 1, "No employee learning data available")
            style_cell(emp_report_ws.cell(6, 1), align="center", bold=True)
            r_master = 7
            emp_report_ws.cell(r_master, 1, "")
            emp_report_ws.cell(r_master, 2, "")
            emp_report_ws.cell(r_master, 3, "Master Total")
            emp_report_ws.cell(r_master, 4, "")
            
            col_c = 5
            for _ in months:
                emp_report_ws.cell(r_master, col_c, "")
                col_c += 1
                emp_report_ws.cell(r_master, col_c, 0)
                col_c += 1
                emp_report_ws.cell(r_master, col_c, 0)
                col_c += 1
                
            emp_report_ws.cell(r_master, col_c, 0) # Tot actual
            col_c += 1
            emp_report_ws.cell(r_master, col_c, 0) # Tot target
            col_c += 1
            emp_report_ws.cell(r_master, col_c, 0) # Tot balance
            col_c += 1
            emp_report_ws.cell(r_master, col_c, 0.0) # Tot KPI%
            col_c += 1
            emp_report_ws.cell(r_master, col_c, "Critical") # Compliance
            col_c += 1
            emp_report_ws.cell(r_master, col_c, "")
            col_c += 1
            emp_report_ws.cell(r_master, col_c, "")
            col_c += 1
            emp_report_ws.cell(r_master, col_c, 0.0) # Avg attendance
            
            for c in range(1, total_cols + 1):
                cell = emp_report_ws.cell(r_master, c)
                style_cell(cell, fill=light_blue, bold=True, align="left" if c in (3, 5) or c > 4 and (c - 5) % 3 == 0 else "center")
                cell.border = double_bottom_border
                if c == col_c - 4:
                    cell.number_format = "0.0%"
                elif c == col_c - 8 or c == col_c - 7 or c == col_c - 6:
                    cell.number_format = "#,##0.0"
                elif c == col_c:
                    cell.number_format = "0.0%"
                elif c > 4 and (c - 5) % 3 != 0:
                    cell.number_format = "#,##0.0"
            emp_report_ws.row_dimensions[r_master].height = 26
            current_row = 8
        else:
            for dept_name, dept_group in groupby(sorted_employees, key=get_dept_name):
                dept_employees = list(dept_group)
                d_start_row = current_row
                
                for emp in dept_employees:
                    r = current_row
                    row_values = [
                        s_no_counter,
                        emp.employee_code,
                        f"{emp.first_name} {emp.last_name}",
                        emp.date_of_joining.strftime("%d %m %Y") if emp.date_of_joining else "",
                    ]
                    s_no_counter += 1
                    
                    total_hours = 0.0
                    total_attendance_rows = 0
                    for y, m, _ in months:
                        entries = contribution_map.get(emp.id, {}).get((y, m), [])
                        total_attendance_rows += len(entries)
                        month_hours = sum(item["hours"] for item in entries)
                        max_hours = sum(item["max_hours"] for item in entries)
                        total_hours += month_hours
                        
                        dates = "\n".join(f"{item['training_date']:%d.%m.%Y} - {item['training_title']}" for item in entries)
                        row_values.extend([dates, month_hours if month_hours else "", max_hours if max_hours else ""])
                        
                    target = TeamAnalyticsService.LEARNING_TARGET_HOURS
                    balance = max(target - total_hours, 0)
                    kpi_pct = round((total_hours / target) * 100, 1) if target else 0
                    department = emp.department.name if emp.department else "Unassigned"
                    manager = f"{emp.manager.first_name} {emp.manager.last_name}" if emp.manager else ""
                    attendance_pct = 100.0 if total_attendance_rows else 0.0
                    
                    row_values.extend([
                        round(total_hours, 2),
                        target,
                        f"{round(balance, 2)} Hours Remaining" if balance else "0",
                        f"{kpi_pct}%",
                        TeamAnalyticsService.compliance_status(kpi_pct),
                        department,
                        manager,
                        f"{attendance_pct:.1f}%",
                    ])
                    
                    emp_report_ws.append(row_values)
                    compliance_fill = green if kpi_pct >= 75 else yellow if kpi_pct >= 25 else red
                    for c in range(1, total_cols + 1):
                        fill = compliance_fill if c in (total_cols - 3, total_cols - 4) else white
                        style_cell(emp_report_ws.cell(r, c), fill=fill, align="left" if c in (3, 5) or c > 4 and (c - 5) % 3 == 0 else "center")
                        
                    emp_report_ws.row_dimensions[r].height = 36
                    current_row += 1
                    
                d_end_row = current_row - 1
                
                # Subtotal row
                r_sub = current_row
                subtotal_rows.append(r_sub)
                
                emp_report_ws.cell(r_sub, 1, "")
                emp_report_ws.cell(r_sub, 2, "")
                emp_report_ws.cell(r_sub, 3, f"{dept_name} Subtotal")
                emp_report_ws.cell(r_sub, 4, "")
                
                col_c = 5
                for _ in months:
                    emp_report_ws.cell(r_sub, col_c, "")
                    col_c += 1
                    
                    act_letter = get_column_letter(col_c)
                    emp_report_ws.cell(r_sub, col_c, f"=SUM({act_letter}{d_start_row}:{act_letter}{d_end_row})")
                    col_c += 1
                    
                    max_letter = get_column_letter(col_c)
                    emp_report_ws.cell(r_sub, col_c, f"=SUM({max_letter}{d_start_row}:{max_letter}{d_end_row})")
                    col_c += 1
                    
                tot_letter = get_column_letter(col_c)
                emp_report_ws.cell(r_sub, col_c, f"=SUM({tot_letter}{d_start_row}:{tot_letter}{d_end_row})")
                col_c += 1
                
                targ_letter = get_column_letter(col_c)
                emp_report_ws.cell(r_sub, col_c, f"=SUM({targ_letter}{d_start_row}:{targ_letter}{d_end_row})")
                col_c += 1
                
                bal_letter = get_column_letter(col_c)
                emp_report_ws.cell(r_sub, col_c, f"=SUM({bal_letter}{d_start_row}:{bal_letter}{d_end_row})")
                col_c += 1
                
                kpi_letter = get_column_letter(col_c)
                emp_report_ws.cell(r_sub, col_c, f"=IF({targ_letter}{r_sub}>0, {tot_letter}{r_sub}/{targ_letter}{r_sub}, 0)")
                col_c += 1
                
                emp_report_ws.cell(r_sub, col_c, f"=IF({kpi_letter}{r_sub}>=1.0, \"Achieved\", IF({kpi_letter}{r_sub}>=0.75, \"Good\", IF({kpi_letter}{r_sub}>=0.5, \"Moderate\", IF({kpi_letter}{r_sub}>=0.25, \"Low\", \"Critical\"))))")
                col_c += 1
                
                emp_report_ws.cell(r_sub, col_c, dept_name)
                col_c += 1
                
                emp_report_ws.cell(r_sub, col_c, "")
                col_c += 1
                
                att_letter = get_column_letter(col_c)
                emp_report_ws.cell(r_sub, col_c, f"=AVERAGE({att_letter}{d_start_row}:{att_letter}{d_end_row})")
                
                for c in range(1, total_cols + 1):
                    cell = emp_report_ws.cell(r_sub, c)
                    style_cell(cell, fill=light_blue, bold=True, align="left" if c in (3, 5) or c > 4 and (c - 5) % 3 == 0 else "center")
                    if c == col_c - 4:
                        cell.number_format = "0.0%"
                    elif c == col_c - 8 or c == col_c - 7 or c == col_c - 6:
                        cell.number_format = "#,##0.0"
                    elif c == col_c:
                        cell.number_format = "0.0%"
                    elif c > 4 and (c - 5) % 3 != 0:
                        cell.number_format = "#,##0.0"
                        
                emp_report_ws.row_dimensions[r_sub].height = 24
                current_row += 1

            # Master Total row
            r_master = current_row
            emp_report_ws.cell(r_master, 1, "")
            emp_report_ws.cell(r_master, 2, "")
            emp_report_ws.cell(r_master, 3, "Master Total")
            emp_report_ws.cell(r_master, 4, "")
            
            col_c = 5
            for _ in months:
                emp_report_ws.cell(r_master, col_c, "")
                col_c += 1
                
                act_letter = get_column_letter(col_c)
                sub_sum_str = "+".join(f"{act_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
                emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
                col_c += 1
                
                max_letter = get_column_letter(col_c)
                sub_sum_str = "+".join(f"{max_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
                emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
                col_c += 1
                
            tot_letter = get_column_letter(col_c)
            sub_sum_str = "+".join(f"{tot_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
            emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
            col_c += 1
            
            targ_letter = get_column_letter(col_c)
            sub_sum_str = "+".join(f"{targ_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
            emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
            col_c += 1
            
            bal_letter = get_column_letter(col_c)
            sub_sum_str = "+".join(f"{bal_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
            emp_report_ws.cell(r_master, col_c, f"={sub_sum_str}")
            col_c += 1
            
            kpi_letter = get_column_letter(col_c)
            emp_report_ws.cell(r_master, col_c, f"=IF({targ_letter}{r_master}>0, {tot_letter}{r_master}/{targ_letter}{r_master}, 0)")
            col_c += 1
            
            emp_report_ws.cell(r_master, col_c, f"=IF({kpi_letter}{r_master}>=1.0, \"Achieved\", IF({kpi_letter}{r_master}>=0.75, \"Good\", IF({kpi_letter}{r_master}>=0.5, \"Moderate\", IF({kpi_letter}{r_master}>=0.25, \"Low\", \"Critical\"))))")
            col_c += 1
            
            emp_report_ws.cell(r_master, col_c, "")
            col_c += 1
            
            emp_report_ws.cell(r_master, col_c, "")
            col_c += 1
            
            att_letter = get_column_letter(col_c)
            sub_avg_str = ",".join(f"{att_letter}{sub_r}" for sub_r in subtotal_rows) if subtotal_rows else "0"
            emp_report_ws.cell(r_master, col_c, f"=AVERAGE({sub_avg_str})")
            
            for c in range(1, total_cols + 1):
                cell = emp_report_ws.cell(r_master, c)
                style_cell(cell, fill=light_blue, bold=True, align="left" if c in (3, 5) or c > 4 and (c - 5) % 3 == 0 else "center")
                cell.border = double_bottom_border
                if c == col_c - 4:
                    cell.number_format = "0.0%"
                elif c == col_c - 8 or c == col_c - 7 or c == col_c - 6:
                    cell.number_format = "#,##0.0"
                elif c == col_c:
                    cell.number_format = "0.0%"
                elif c > 4 and (c - 5) % 3 != 0:
                    cell.number_format = "#,##0.0"
                    
            emp_report_ws.row_dimensions[r_master].height = 26

        
        widths = [8, 14, 24, 16]
        widths.extend([28, 12, 12] * len(months))
        widths.extend([12, 12, 18, 10, 20, 18, 22, 14])
        for i, w in enumerate(widths, 1):
            emp_report_ws.column_dimensions[get_column_letter(i)].width = w
            
        emp_report_ws.auto_filter.ref = f"A5:{get_column_letter(total_cols)}{r_master - 1}"

        # 8. Populate hidden list sheet Filter_Lists
        list_ws.cell(1, 1, "Financial Years")
        list_ws.cell(1, 2, "Quarters")
        list_ws.cell(1, 3, "Months")
        list_ws.cell(1, 4, "Departments")
        list_ws.cell(1, 5, "Categories")
        list_ws.cell(1, 6, "Managers")
        list_ws.cell(1, 7, "Training Types")
        
        quarters_list = ["All", "Q1", "Q2", "Q3", "Q4"]
        months_list = ["All", "April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February", "March"]
        
        unique_fys = set()
        for record, training, employee, enrollment in all_attendance_results:
            if training.start_date:
                y = training.start_date.year
                m = training.start_date.month
                start_yr = y if m >= 4 else y - 1
                unique_fys.add(f"{start_yr}-{start_yr + 1}")
        unique_fys.add(fy_label)
        sorted_fys = sorted(list(unique_fys), reverse=True)
        
        for i, val in enumerate(["All"] + sorted_fys, 2):
            list_ws.cell(i, 1, val)
        for i, val in enumerate(quarters_list, 2):
            list_ws.cell(i, 2, val)
        for i, val in enumerate(months_list, 2):
            list_ws.cell(i, 3, val)
        for i, val in enumerate(["All"] + dept_names, 2):
            list_ws.cell(i, 4, val)
        for i, val in enumerate(["All"] + cat_names, 2):
            list_ws.cell(i, 5, val)
        for i, val in enumerate(["All"] + mgr_names, 2):
            list_ws.cell(i, 6, val)
        for i, val in enumerate(["All"] + training_types, 2):
            list_ws.cell(i, 7, val)

        # 9. Populate hidden Raw_Attendance_Data
        raw_row_count = len(all_attendance_results) + 1
        if raw_row_count < 2:
            raw_row_count = 2
            
        raw_headers = [
            "Matches_Filter", "Employee Code", "Employee Name", "Department", "Manager",
            "Training Date", "Financial Year", "Quarter", "Month Name", "Training Title",
            "Training Category", "Training Type", "Attendance Status", "Actual Hours",
            "Max Hours", "Is_Unique_Training", "Is_Attended_Present", "Is_Completed", "Actual Hours Credit"
        ]
        for c_idx, h in enumerate(raw_headers, 1):
            raw_ws.cell(1, c_idx, h)
            
        for idx, (record, training, employee, enrollment) in enumerate(all_attendance_results):
            r = idx + 2
            emp_code = employee.employee_code
            emp_name = f"{employee.first_name} {employee.last_name}"
            dept_name = employee.department.name if employee.department else "Unassigned"
            mgr_name = f"{employee.manager.first_name} {employee.manager.last_name}" if employee.manager else ""
            
            t_date = training.start_date
            fy_str = ""
            q_str = ""
            m_str = ""
            if t_date:
                y = t_date.year
                m = t_date.month
                start_yr = y if m >= 4 else y - 1
                fy_str = f"{start_yr}-{start_yr + 1}"
                if m in (4, 5, 6):
                    q_str = "Q1"
                elif m in (7, 8, 9):
                    q_str = "Q2"
                elif m in (10, 11, 12):
                    q_str = "Q3"
                else:
                    q_str = "Q4"
                m_str = t_date.strftime("%B")
                
            t_title = training.title
            t_cat = training.category.name if training.category else ""
            t_type = training.training_type.value if hasattr(training.training_type, "value") else str(training.training_type)
            att_status = record.status.value if hasattr(record.status, "value") else str(record.status)
            act_hours = float(training.duration_hours or 0)
            max_hours = float(training.max_hours_allowed or training.duration_hours or 0)
            
            raw_ws.cell(r, 1, f"=IF(AND(OR('Department KPI Analytics'!$B$6=\"All\", G{r}='Department KPI Analytics'!$B$6), OR('Department KPI Analytics'!$D$6=\"All\", H{r}='Department KPI Analytics'!$D$6), OR('Department KPI Analytics'!$E$6=\"All\", I{r}='Department KPI Analytics'!$E$6), OR('Department KPI Analytics'!$G$6=\"All\", D{r}='Department KPI Analytics'!$G$6), OR('Department KPI Analytics'!$J$6=\"All\", K{r}='Department KPI Analytics'!$J$6)), 1, 0)")
            raw_ws.cell(r, 2, emp_code)
            raw_ws.cell(r, 3, emp_name)
            raw_ws.cell(r, 4, dept_name)
            raw_ws.cell(r, 5, mgr_name)
            raw_ws.cell(r, 6, t_date.strftime("%Y-%m-%d") if t_date else "")
            raw_ws.cell(r, 7, fy_str)
            raw_ws.cell(r, 8, q_str)
            raw_ws.cell(r, 9, m_str)
            raw_ws.cell(r, 10, t_title)
            raw_ws.cell(r, 11, t_cat)
            raw_ws.cell(r, 12, t_type)
            raw_ws.cell(r, 13, att_status)
            raw_ws.cell(r, 14, act_hours)
            raw_ws.cell(r, 15, max_hours)
            
            raw_ws.cell(r, 16, f"=IF(AND(A{r}=1, COUNTIFS($J$2:J{r}, J{r}, $A$2:A{r}, 1)=1), 1, 0)")
            raw_ws.cell(r, 17, f"=IF(AND(A{r}=1, OR(M{r}=\"PRESENT\", M{r}=\"LATE\")), 1, 0)")
            raw_ws.cell(r, 18, 1 if enrollment.status == EnrollmentStatus.COMPLETED else 0)
            raw_ws.cell(r, 19, f"=IF(AND(Q{r}=1, R{r}=1), N{r}, 0)")

        # 10. Populate hidden Employee_KPI_Summary
        emp_headers = [
            "S.No", "Employee Code", "Employee Name", "Department", "Manager",
            "Target Hours", "Actual Hours", "Remaining Hours", "Achievement %",
            "Attended Count", "Total Assigned", "Attendance %"
        ]
        for c_idx, h in enumerate(emp_headers, 1):
            emp_summary_ws.cell(1, c_idx, h)
            
        emp_start_row = 6
        emp_summary_ws.merge_cells("A1:L1")
        emp_summary_ws.cell(1, 1, "Detailed Employee KPI Metrics & Compliance Status").font = Font(name="Calibri", size=12, bold=True)
        emp_summary_ws.cell(2, 1, "Do not edit this sheet directly. It contains formulas linking Employee metrics to the Dashboard filters.")
        
        emp_row_count = max(len(all_employees) + emp_start_row - 1, emp_start_row)
        for idx, emp in enumerate(all_employees):
            r = emp_start_row + idx
            emp_code = emp.employee_code
            emp_name = f"{emp.first_name} {emp.last_name}"
            dept_name = emp.department.name if emp.department else "Unassigned"
            mgr_name = f"{emp.manager.first_name} {emp.manager.last_name}" if emp.manager else ""
            
            emp_summary_ws.cell(r, 1, idx + 1)
            emp_summary_ws.cell(r, 2, emp_code)
            emp_summary_ws.cell(r, 3, emp_name)
            emp_summary_ws.cell(r, 4, dept_name)
            emp_summary_ws.cell(r, 5, mgr_name)
            emp_summary_ws.cell(r, 6, f"=IF(OR('Department KPI Analytics'!$G$6=\"All\", D{r}='Department KPI Analytics'!$G$6), 16, 0)")
            emp_summary_ws.cell(r, 7, f"=SUMIFS('Raw_Attendance_Data'!$S$2:$S${raw_row_count}, 'Raw_Attendance_Data'!$B$2:$B${raw_row_count}, B{r}, 'Raw_Attendance_Data'!$A$2:$A${raw_row_count}, 1)")
            emp_summary_ws.cell(r, 8, f"=MAX(F{r}-G{r}, 0)")
            emp_summary_ws.cell(r, 9, f"=IF(F{r}>0, G{r}/F{r}, 0)")
            emp_summary_ws.cell(r, 10, f"=COUNTIFS('Raw_Attendance_Data'!$B$2:$B${raw_row_count}, B{r}, 'Raw_Attendance_Data'!$A$2:$A${raw_row_count}, 1, 'Raw_Attendance_Data'!$Q$2:$Q${raw_row_count}, 1)")
            emp_summary_ws.cell(r, 11, f"=COUNTIFS('Raw_Attendance_Data'!$B$2:$B${raw_row_count}, B{r}, 'Raw_Attendance_Data'!$A$2:$A${raw_row_count}, 1)")
            emp_summary_ws.cell(r, 12, f"=IF(K{r}>0, J{r}/K{r}, 1.0)")


        # 11. Populate Sheet 2: Department KPI Analytics (Dashboard)
        dash_blue = "1F4E79"
        dash_light_blue = "D9EAF7"
        dash_light_gray = "F2F4F7"
        dash_white = "FFFFFF"
        dash_yellow = "FFF2CC"
        dash_font = "Segoe UI"
        
        thin_side = Side(style="thin", color="D9D9D9")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        card_border = Border(left=Side(style="thin", color="B0C4DE"), right=Side(style="thin", color="B0C4DE"), top=Side(style="thin", color="B0C4DE"), bottom=Side(style="thin", color="B0C4DE"))

        
        def create_filter_field(label_range, value_range, label_text, default_val):
            dept_ws.merge_cells(label_range)
            dept_ws.merge_cells(value_range)
            
            l_cell = dept_ws[label_range.split(":")[0]]
            v_cell = dept_ws[value_range.split(":")[0]]
            
            l_cell.value = label_text
            l_cell.font = Font(name=dash_font, size=9, bold=True, color="5A5A5A")
            l_cell.alignment = Alignment(horizontal="center", vertical="center")
            l_cell.fill = PatternFill("solid", fgColor="E9ECEF")
            
            v_cell.value = default_val
            v_cell.font = Font(name=dash_font, size=10, bold=True, color=dash_blue)
            v_cell.alignment = Alignment(horizontal="center", vertical="center")
            v_cell.fill = PatternFill("solid", fgColor=dash_yellow)
            
            def apply_range_border(ws, r_str, border_style):
                from openpyxl.utils.cell import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(r_str)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        ws.cell(row, col).border = border_style

            apply_range_border(dept_ws, label_range, thin_border)
            apply_range_border(dept_ws, value_range, thin_border)

        def create_kpi_card(label_range, value_range, label_text, formula_val, num_format=None):
            dept_ws.merge_cells(label_range)
            dept_ws.merge_cells(value_range)
            
            l_cell = dept_ws[label_range.split(":")[0]]
            v_cell = dept_ws[value_range.split(":")[0]]
            
            l_cell.value = label_text
            l_cell.font = Font(name=dash_font, size=8, bold=True, color="5A5A5A")
            l_cell.alignment = Alignment(horizontal="center", vertical="center")
            l_cell.fill = PatternFill("solid", fgColor=dash_light_gray)
            
            v_cell.value = formula_val
            v_cell.font = Font(name=dash_font, size=13, bold=True, color=dash_blue)
            v_cell.alignment = Alignment(horizontal="center", vertical="center")
            v_cell.fill = PatternFill("solid", fgColor=dash_light_gray)
            if num_format:
                v_cell.number_format = num_format
                
            def apply_range_border(ws, r_str, border_style):
                from openpyxl.utils.cell import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(r_str)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        ws.cell(row, col).border = border_style
                        ws.cell(row, col).fill = PatternFill("solid", fgColor=dash_light_gray)

            apply_range_border(dept_ws, label_range, card_border)
            apply_range_border(dept_ws, value_range, card_border)

        
        dept_ws.merge_cells("A1:L1")
        dept_ws["A1"] = "ORGANIZATION LEARNING & KPI ANALYTICS DASHBOARD"
        dept_ws["A1"].font = Font(name=dash_font, size=16, bold=True, color=dash_white)
        dept_ws["A1"].fill = PatternFill("solid", fgColor=dash_blue)
        dept_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        dept_ws.row_dimensions[1].height = 36
        
        dept_ws.merge_cells("A2:L2")
        dept_ws["A2"] = "Executive Overview & Department Contribution Analysis"
        dept_ws["A2"].font = Font(name=dash_font, size=10, italic=True, color="D9EAF7")
        dept_ws["A2"].fill = PatternFill("solid", fgColor=dash_blue)
        dept_ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        dept_ws.row_dimensions[2].height = 24
        
        dept_ws.merge_cells("A4:L4")
        dept_ws["A4"] = "Interactive Filters (Select values to dynamically update the dashboard)"
        dept_ws["A4"].font = Font(name=dash_font, size=10, bold=True, color=dash_blue)
        dept_ws["A4"].fill = PatternFill("solid", fgColor=dash_light_blue)
        dept_ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
        dept_ws.row_dimensions[4].height = 24
        
        create_filter_field("B5:C5", "B6:C6", "Financial Year", selected_fy)
        create_filter_field("D5", "D6", "Quarter", "All")
        create_filter_field("E5:F5", "E6:F6", "Month", selected_month_name)
        create_filter_field("G5:I5", "G6:I6", "Department", selected_dept_name)
        create_filter_field("J5:L5", "J6:L6", "Training Category", selected_cat_name)
        dept_ws.row_dimensions[5].height = 20
        dept_ws.row_dimensions[6].height = 24
        
        dv_fy = DataValidation(type="list", formula1="Filter_Lists!$A$2:$A$20", allow_blank=True)
        dv_q = DataValidation(type="list", formula1="Filter_Lists!$B$2:$B$6", allow_blank=True)
        dv_m = DataValidation(type="list", formula1="Filter_Lists!$C$2:$C$14", allow_blank=True)
        dv_dept = DataValidation(type="list", formula1=f"Filter_Lists!$D$2:$D${len(dept_names)+2}", allow_blank=True)
        dv_cat = DataValidation(type="list", formula1=f"Filter_Lists!$E$2:$E${len(cat_names)+2}", allow_blank=True)

        dept_ws.add_data_validation(dv_fy)

        dept_ws.add_data_validation(dv_q)
        dept_ws.add_data_validation(dv_m)
        dept_ws.add_data_validation(dv_dept)
        dept_ws.add_data_validation(dv_cat)
        
        dv_fy.add(dept_ws["B6"])
        dv_q.add(dept_ws["D6"])
        dv_m.add(dept_ws["E6"])
        dv_dept.add(dept_ws["G6"])
        dv_cat.add(dept_ws["J6"])
        
        create_kpi_card("A10:B10", "A11:B11", "TOTAL EMPLOYEES", f"=COUNTIFS('Employee_KPI_Summary'!$F$6:$F${emp_row_count}, \">0\")", "#,##0")
        create_kpi_card("C10:D10", "C11:D11", "TOTAL TARGET HOURS", f"=SUM('Employee_KPI_Summary'!$F$6:$F${emp_row_count})", "#,##0.0")
        create_kpi_card("E10:F10", "E11:F11", "TOTAL ACTUAL HOURS", f"=SUM('Employee_KPI_Summary'!$G$6:$G${emp_row_count})", "#,##0.0")
        create_kpi_card("G10:H10", "G11:H11", "TOTAL REMAINING HOURS", f"=SUM('Employee_KPI_Summary'!$H$6:$H${emp_row_count})", "#,##0.0")
        create_kpi_card("I10:L10", "I11:L11", "OVERALL ACHIEVEMENT %", "=IF(C11>0, E11/C11, 0)", "0.0%")
        
        create_kpi_card("A12:B12", "A13:B13", "LEARNING COMPLIANCE %", f"=IF(COUNTIFS('Employee_KPI_Summary'!$F$6:$F${emp_row_count}, \">0\")>0, COUNTIFS('Employee_KPI_Summary'!$F$6:$F${emp_row_count}, \">0\", 'Employee_KPI_Summary'!$I$6:$I${emp_row_count}, \">=1.0\")/COUNTIFS('Employee_KPI_Summary'!$F$6:$F${emp_row_count}, \">0\"), 1.0)", "0.0%")
        create_kpi_card("C12:D12", "C13:D13", "TRAININGS CONDUCTED", f"=SUM('Raw_Attendance_Data'!$P$2:$P${raw_row_count})", "#,##0")
        create_kpi_card("E12:F12", "E13:F13", "TOTAL ATTENDANCE %", f"=IF(SUM('Employee_KPI_Summary'!$K$6:$K${emp_row_count})>0, SUM('Employee_KPI_Summary'!$J$6:$J${emp_row_count})/SUM('Employee_KPI_Summary'!$K$6:$K${emp_row_count}), 1.0)", "0.0%")

        
        depts_to_write = dept_names if dept_names else ["Unassigned"]
        N_DEPTS = len(depts_to_write)
        depts_start_row = 17
        depts_end_row = depts_start_row + N_DEPTS - 1
        totals_row = depts_end_row + 1
        
        create_kpi_card("G12:H12", "G13:H13", "TOP PERFORMING DEPT", f"=INDEX(B17:B{depts_end_row}, MATCH(MAX(G17:G{depts_end_row}), G17:G{depts_end_row}, 0))")
        create_kpi_card("I12:L12", "I13:L13", "LOWEST PERFORMING DEPT", f"=INDEX(B17:B{depts_end_row}, MATCH(MIN(K17:K{depts_end_row}), K17:K{depts_end_row}, 0))")
        
        dept_ws.merge_cells("A15:K15")
        dept_ws["A15"] = "DEPARTMENT CONTRIBUTION & KPI SUMMARY"
        dept_ws["A15"].font = Font(name=dash_font, size=11, bold=True, color=dash_white)
        dept_ws["A15"].fill = PatternFill("solid", fgColor=dash_blue)
        dept_ws["A15"].alignment = Alignment(horizontal="center", vertical="center")
        dept_ws.row_dimensions[15].height = 24
        
        headers = ["S.No", "Department", "Employees", "Target Hours", "Actual Hours", "Remaining Hours", "Achievement %", "Attendance %", "KPI Status", "Rank Score", "Filtered Achievement"]
        for c_idx, h in enumerate(headers, 1):
            cell = dept_ws.cell(16, c_idx, h)
            cell.font = Font(name=dash_font, size=9, bold=True, color=dash_blue)
            cell.fill = PatternFill("solid", fgColor=dash_light_blue)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        dept_ws.row_dimensions[16].height = 20
        
        for idx, dept in enumerate(depts_to_write):
            r = depts_start_row + idx
            dept_ws.cell(r, 1, f"=ROW() - 16")
            dept_ws.cell(r, 2, dept)
            dept_ws.cell(r, 3, f"=COUNTIFS('Employee_KPI_Summary'!$D$6:$D${emp_row_count}, B{r}, 'Employee_KPI_Summary'!$F$6:$F${emp_row_count}, \">0\")")
            dept_ws.cell(r, 4, f"=SUMIFS('Employee_KPI_Summary'!$F$6:$F${emp_row_count}, 'Employee_KPI_Summary'!$D$6:$D${emp_row_count}, B{r})")
            dept_ws.cell(r, 5, f"=SUMIFS('Employee_KPI_Summary'!$G$6:$G${emp_row_count}, 'Employee_KPI_Summary'!$D$6:$D${emp_row_count}, B{r})")
            dept_ws.cell(r, 6, f"=SUMIFS('Employee_KPI_Summary'!$H$6:$H${emp_row_count}, 'Employee_KPI_Summary'!$D$6:$D${emp_row_count}, B{r})")
            dept_ws.cell(r, 7, f"=IF(D{r}>0, E{r}/D{r}, 0)")
            dept_ws.cell(r, 8, f"=IF(SUMIFS('Employee_KPI_Summary'!$K$6:$K${emp_row_count}, 'Employee_KPI_Summary'!$D$6:$D${emp_row_count}, B{r})>0, SUMIFS('Employee_KPI_Summary'!$J$6:$J${emp_row_count}, 'Employee_KPI_Summary'!$D$6:$D${emp_row_count}, B{r})/SUMIFS('Employee_KPI_Summary'!$K$6:$K${emp_row_count}, 'Employee_KPI_Summary'!$D$6:$D${emp_row_count}, B{r}), 1.0)")
            dept_ws.cell(r, 9, f"=IF(G{r}>=1.0, \"Achieved\", IF(G{r}>=0.75, \"Good\", IF(G{r}>=0.5, \"Moderate\", IF(G{r}>=0.25, \"Low\", \"Critical\"))))")
            dept_ws.cell(r, 10, f"=G{r}+(1000-ROW())/1000000")
            dept_ws.cell(r, 11, f"=IF(C{r}>0, G{r}+ROW()/1000000, 9.99)")
            
            for c_idx in range(1, 12):
                cell = dept_ws.cell(r, c_idx)
                cell.font = Font(name=dash_font, size=9, color="1F2937")
                cell.border = thin_border
                if c_idx == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 9:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if c_idx in (3, 10, 11):
                    cell.number_format = "#,##0"
                elif c_idx in (4, 5, 6):
                    cell.number_format = "#,##0.0"
                elif c_idx in (7, 8):
                    cell.number_format = "0.0%"
            dept_ws.row_dimensions[r].height = 20
            
        dept_ws.cell(totals_row, 1, "")
        dept_ws.cell(totals_row, 2, "Total")
        dept_ws.cell(totals_row, 3, f"=SUM(C17:C{depts_end_row})")
        dept_ws.cell(totals_row, 4, f"=SUM(D17:D{depts_end_row})")
        dept_ws.cell(totals_row, 5, f"=SUM(E17:E{depts_end_row})")
        dept_ws.cell(totals_row, 6, f"=SUM(F17:F{depts_end_row})")
        dept_ws.cell(totals_row, 7, f"=IF(D{totals_row}>0, E{totals_row}/D{totals_row}, 0)")
        dept_ws.cell(totals_row, 8, f"=IF(SUM('Employee_KPI_Summary'!$K$6:$K${emp_row_count})>0, SUM('Employee_KPI_Summary'!$J$6:$J${emp_row_count})/SUM('Employee_KPI_Summary'!$K$6:$K${emp_row_count}), 1.0)")
        dept_ws.cell(totals_row, 9, f"=IF(G{totals_row}>=1.0, \"Achieved\", IF(G{totals_row}>=0.75, \"Good\", IF(G{totals_row}>=0.5, \"Moderate\", IF(G{totals_row}>=0.25, \"Low\", \"Critical\"))))")

        dept_ws.cell(totals_row, 10, "")
        dept_ws.cell(totals_row, 11, "")
        
        for c_idx in range(1, 12):
            cell = dept_ws.cell(totals_row, c_idx)
            cell.font = Font(name=dash_font, size=9, bold=True, color=dash_blue)
            cell.fill = PatternFill("solid", fgColor=dash_light_blue)
            cell.border = double_bottom_border
            if c_idx == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx == 9:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if c_idx == 3:
                cell.number_format = "#,##0"
            elif c_idx in (4, 5, 6):
                cell.number_format = "#,##0.0"
            elif c_idx in (7, 8):
                cell.number_format = "0.0%"
        dept_ws.row_dimensions[totals_row].height = 22
        
        green_fill = PatternFill("solid", fgColor="D1FAE5")
        green_font = Font(name=dash_font, size=9, bold=True, color="065F46")
        light_green_fill = PatternFill("solid", fgColor="E2F0D9")
        light_green_font = Font(name=dash_font, size=9, bold=True, color="385723")
        yellow_fill = PatternFill("solid", fgColor="FFF2CC")
        yellow_font = Font(name=dash_font, size=9, bold=True, color="7F6000")
        orange_fill = PatternFill("solid", fgColor="FCE4D6")
        orange_font = Font(name=dash_font, size=9, bold=True, color="C65911")
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        red_font = Font(name=dash_font, size=9, bold=True, color="991B1B")
        
        rule_achieved = CellIsRule(operator="equal", formula=['"Achieved"'], fill=green_fill, font=green_font)
        rule_good = CellIsRule(operator="equal", formula=['"Good"'], fill=light_green_fill, font=light_green_font)
        rule_moderate = CellIsRule(operator="equal", formula=['"Moderate"'], fill=yellow_fill, font=yellow_font)
        rule_low = CellIsRule(operator="equal", formula=['"Low"'], fill=orange_fill, font=orange_font)
        rule_critical = CellIsRule(operator="equal", formula=['"Critical"'], fill=red_fill, font=red_font)
        
        dept_ws.conditional_formatting.add(f"I17:I{depts_end_row}", rule_achieved)
        dept_ws.conditional_formatting.add(f"I17:I{depts_end_row}", rule_good)
        dept_ws.conditional_formatting.add(f"I17:I{depts_end_row}", rule_moderate)
        dept_ws.conditional_formatting.add(f"I17:I{depts_end_row}", rule_low)
        dept_ws.conditional_formatting.add(f"I17:I{depts_end_row}", rule_critical)
        
        # Rankings
        rank_start_row = totals_row + 3
        dept_ws.merge_cells(f"A{rank_start_row}:E{rank_start_row}")
        dept_ws[f"A{rank_start_row}"] = "Top Performing Departments"
        dept_ws[f"A{rank_start_row}"].font = Font(name=dash_font, size=10, bold=True, color=dash_white)
        dept_ws[f"A{rank_start_row}"].fill = PatternFill("solid", fgColor="10B981")
        dept_ws[f"A{rank_start_row}"].alignment = Alignment(horizontal="center", vertical="center")
        
        dept_ws.merge_cells(f"G{rank_start_row}:J{rank_start_row}")
        dept_ws[f"G{rank_start_row}"] = "Departments Requiring Attention"
        dept_ws[f"G{rank_start_row}"].font = Font(name=dash_font, size=10, bold=True, color=dash_white)
        dept_ws[f"G{rank_start_row}"].fill = PatternFill("solid", fgColor="EF4444")
        dept_ws[f"G{rank_start_row}"].alignment = Alignment(horizontal="center", vertical="center")
        
        dept_ws.row_dimensions[rank_start_row].height = 22
        
        top_headers = ["Rank", "Department", "Achievement %", "Actual Hours", "Attendance %"]
        for c_idx, h in enumerate(top_headers, 1):
            cell = dept_ws.cell(rank_start_row + 1, c_idx, h)
            cell.font = Font(name=dash_font, size=9, bold=True, color="5A5A5A")
            cell.fill = PatternFill("solid", fgColor="F2F4F7")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        bot_headers = ["Department", "Remaining Hours", "Compliance %", "Risk Indicator"]
        for c_idx, h in enumerate(bot_headers, 7):
            cell = dept_ws.cell(rank_start_row + 1, c_idx, h)
            cell.font = Font(name=dash_font, size=9, bold=True, color="5A5A5A")
            cell.fill = PatternFill("solid", fgColor="F2F4F7")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        dept_ws.row_dimensions[rank_start_row + 1].height = 20
        
        num_rank_rows = min(5, N_DEPTS)
        for i in range(1, num_rank_rows + 1):
            r_idx = rank_start_row + 1 + i
            dept_ws.cell(r_idx, 1, i)
            dept_ws.cell(r_idx, 2, f"=INDEX($B$17:$B${depts_end_row}, MATCH(LARGE($J$17:$J${depts_end_row}, A{r_idx}), $J$17:$J${depts_end_row}, 0))")
            dept_ws.cell(r_idx, 3, f"=INDEX($G$17:$G${depts_end_row}, MATCH(LARGE($J$17:$J${depts_end_row}, A{r_idx}), $J$17:$J${depts_end_row}, 0))")
            dept_ws.cell(r_idx, 4, f"=INDEX($E$17:$E${depts_end_row}, MATCH(LARGE($J$17:$J${depts_end_row}, A{r_idx}), $J$17:$J${depts_end_row}, 0))")
            dept_ws.cell(r_idx, 5, f"=INDEX($H$17:$H${depts_end_row}, MATCH(LARGE($J$17:$J${depts_end_row}, A{r_idx}), $J$17:$J${depts_end_row}, 0))")
            
            dept_ws.cell(r_idx, 7, f"=INDEX($B$17:$B${depts_end_row}, MATCH(SMALL($K$17:$K${depts_end_row}, {i}), $K$17:$K${depts_end_row}, 0))")
            dept_ws.cell(r_idx, 8, f"=INDEX($F$17:$F${depts_end_row}, MATCH(SMALL($K$17:$K${depts_end_row}, {i}), $K$17:$K${depts_end_row}, 0))")
            dept_ws.cell(r_idx, 9, f"=INDEX($G$17:$G${depts_end_row}, MATCH(SMALL($K$17:$K${depts_end_row}, {i}), $K$17:$K${depts_end_row}, 0))")
            dept_ws.cell(r_idx, 10, f"=IF(I{r_idx}<0.5, \"🔴 Critical Risk\", IF(I{r_idx}<0.75, \"🟡 Moderate Risk\", \"🟢 Low Risk\"))")
            
            for c in range(1, 6):
                cell = dept_ws.cell(r_idx, c)
                cell.font = Font(name=dash_font, size=9, color="1F2937")
                cell.border = thin_border
                if c == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if c in (3, 5):
                    cell.number_format = "0.0%"
                elif c == 4:
                    cell.number_format = "#,##0.0"
                    
            for c in range(7, 11):
                cell = dept_ws.cell(r_idx, c)
                cell.font = Font(name=dash_font, size=9, color="1F2937")
                cell.border = thin_border
                if c == 7:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c == 10:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if c == 8:
                    cell.number_format = "#,##0.0"
                elif c == 9:
                    cell.number_format = "0.0%"
            dept_ws.row_dimensions[r_idx].height = 20
            
        rank_end_row = rank_start_row + 1 + num_rank_rows
        
        rule_crit_risk = CellIsRule(operator="equal", formula=['"🔴 Critical Risk"'], fill=red_fill, font=red_font)
        rule_mod_risk = CellIsRule(operator="equal", formula=['"🟡 Moderate Risk"'], fill=yellow_fill, font=yellow_font)
        rule_low_risk = CellIsRule(operator="equal", formula=['"🟢 Low Risk"'], fill=green_fill, font=green_font)
        
        risk_range = f"J{rank_start_row+2}:J{rank_start_row+1+num_rank_rows}"
        dept_ws.conditional_formatting.add(risk_range, rule_crit_risk)
        dept_ws.conditional_formatting.add(risk_range, rule_mod_risk)
        dept_ws.conditional_formatting.add(risk_range, rule_low_risk)
        
        # Monthly run rate table + Insights
        insights_start_row = rank_end_row + 3
        dept_ws.merge_cells(f"A{insights_start_row}:E{insights_start_row}")
        dept_ws[f"A{insights_start_row}"] = "Monthly Learning Trends (FY)"
        dept_ws[f"A{insights_start_row}"].font = Font(name=dash_font, size=10, bold=True, color=dash_white)
        dept_ws[f"A{insights_start_row}"].fill = PatternFill("solid", fgColor="8B5CF6")
        dept_ws[f"A{insights_start_row}"].alignment = Alignment(horizontal="center", vertical="center")
        
        dept_ws.merge_cells(f"G{insights_start_row}:K{insights_start_row}")
        dept_ws[f"G{insights_start_row}"] = "Executive KPI Insights"
        dept_ws[f"G{insights_start_row}"].font = Font(name=dash_font, size=10, bold=True, color=dash_white)
        dept_ws[f"G{insights_start_row}"].fill = PatternFill("solid", fgColor="1F4E79")
        dept_ws[f"G{insights_start_row}"].alignment = Alignment(horizontal="center", vertical="center")
        
        dept_ws.row_dimensions[insights_start_row].height = 22
        
        monthly_headers = ["Month", "Target Hours", "Actual Hours", "Achievement %"]
        for c_idx, h in enumerate(monthly_headers, 1):
            cell = dept_ws.cell(insights_start_row + 1, c_idx, h)
            cell.font = Font(name=dash_font, size=9, bold=True, color="5A5A5A")
            cell.fill = PatternFill("solid", fgColor="F2F4F7")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        dept_ws.row_dimensions[insights_start_row + 1].height = 20
        
        months_names = ["April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February", "March"]
        for idx, month_name in enumerate(months_names):
            curr_row = insights_start_row + 2 + idx
            dept_ws.cell(curr_row, 1, month_name)
            dept_ws.cell(curr_row, 2, f"=$A$11 * 1.33")
            dept_ws.cell(curr_row, 3, f"=SUMIFS('Raw_Attendance_Data'!$S$2:$S${raw_row_count}, 'Raw_Attendance_Data'!$I$2:$I${raw_row_count}, A{curr_row}, 'Raw_Attendance_Data'!$A$2:$A${raw_row_count}, 1)")
            dept_ws.cell(curr_row, 4, f"=IF(B{curr_row}>0, C{curr_row}/B{curr_row}, 0)")
            
            for c in range(1, 5):
                cell = dept_ws.cell(curr_row, c)
                cell.font = Font(name=dash_font, size=9, color="1F2937")
                cell.border = thin_border
                if c == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if c in (2, 3):
                    cell.number_format = "#,##0.0"
                elif c == 4:
                    cell.number_format = "0.0%"
            dept_ws.row_dimensions[curr_row].height = 18
            
        insights_formulas = [
            f"=\"• \" & INDEX(B17:B{depts_end_row}, MATCH(MAX(E17:E{depts_end_row}), E17:E{depts_end_row}, 0)) & \" contributed the highest learning hours (\" & TEXT(MAX(E17:E{depts_end_row}), \"#,##0\") & \" hours).\"",
            f"=\"• \" & INDEX(B17:B{depts_end_row}, MATCH(MIN(K17:K{depts_end_row}), K17:K{depts_end_row}, 0)) & \" is \" & TEXT(MAX(0, 1-MIN(K17:K{depts_end_row})), \"0%\") & \" below annual learning target.\"",
            f"=\"• \" & INDEX(B17:B{depts_end_row}, MATCH(MAX(G17:G{depts_end_row}), G17:G{depts_end_row}, 0)) & \" achieved \" & TEXT(MAX(G17:G{depts_end_row}), \"0.0%\") & \" learning compliance.\"",
            "=\"• Organization-wide learning compliance stands at \" & TEXT(A13, \"0.0%\") & \".\""
        ]
        for idx, formula in enumerate(insights_formulas):
            curr_row = insights_start_row + 2 + idx
            dept_ws.merge_cells(start_row=curr_row, start_column=7, end_row=curr_row, end_column=11)
            cell = dept_ws.cell(curr_row, 7, formula)
            cell.font = Font(name=dash_font, size=9.5, italic=True, color="1F4E79")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for col in range(7, 12):
                c_cell = dept_ws.cell(curr_row, col)
                c_cell.fill = PatternFill("solid", fgColor="F8F9FA")
                c_cell.border = thin_border
            dept_ws.row_dimensions[curr_row].height = 24
            
        column_widths = {
            "A": 8, "B": 24, "C": 14, "D": 16, "E": 16, "F": 16, "G": 18, "H": 18, "I": 16, "J": 14, "K": 18,
            "L": 4
        }
        for col_letter, w in column_widths.items():
            dept_ws.column_dimensions[col_letter].width = w
            
        # Embedded Charts on Sheet 2
        try:
            chart1 = DoughnutChart()
            chart1.title = "Organization Learning Contribution by Department"
            chart1.style = 10
            chart1.holeSize = 50
            chart1.width = 16
            chart1.height = 10
            data1 = Reference(dept_ws, min_col=5, min_row=16, max_row=depts_end_row)
            cats1 = Reference(dept_ws, min_col=2, min_row=17, max_row=depts_end_row)
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(cats1)
            dept_ws.add_chart(chart1, "N15")
            
            chart2 = BarChart()
            chart2.type = "col"
            chart2.style = 11
            chart2.title = "Target Hours vs Actual Hours by Department"
            chart2.y_axis.title = "Hours"
            chart2.x_axis.title = "Department"
            chart2.width = 16
            chart2.height = 10
            data2 = Reference(dept_ws, min_col=4, max_col=5, min_row=16, max_row=depts_end_row)
            cats2 = Reference(dept_ws, min_col=2, min_row=17, max_row=depts_end_row)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)
            dept_ws.add_chart(chart2, "N30")
            
            chart3 = BarChart()
            chart3.type = "col"
            chart3.style = 12
            chart3.title = "Department Compliance %"
            chart3.y_axis.title = "Compliance %"
            chart3.x_axis.title = "Department"
            chart3.width = 16
            chart3.height = 10
            data3 = Reference(dept_ws, min_col=3, min_row=rank_start_row + 1, max_row=rank_start_row + 1 + num_rank_rows)
            cats3 = Reference(dept_ws, min_col=2, min_row=rank_start_row + 2, max_row=rank_start_row + 1 + num_rank_rows)
            chart3.add_data(data3, titles_from_data=True)
            chart3.set_categories(cats3)
            dept_ws.add_chart(chart3, "N45")
        except Exception as chart_err:
            import logging
            logging.getLogger(__name__).warning("Chart generation failed: %s", chart_err)


        # 12. Populate Sheet 3: Monthly Learning Summary
        month_ws.merge_cells("A1:G1")
        month_ws["A1"] = "Monthly Learning Performance Summary"
        month_ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=white)
        month_ws["A1"].fill = PatternFill("solid", fgColor=deep_blue)
        month_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        month_ws.row_dimensions[1].height = 30
        
        headers = ["Month", "Target Hours", "Actual Hours", "Achievement %", "Attendance %", "Training Count", "Monthly Trends"]
        for c_idx, h in enumerate(headers, 1):
            cell = month_ws.cell(3, c_idx, h)
            style_cell(cell, fill=blue, font_color=white, bold=True)
        month_ws.row_dimensions[3].height = 24
        
        for idx, (y, m, month_name) in enumerate(months):
            r = 4 + idx
            month_records = [item for item in attendance_rows if item["training_date"].year == y and item["training_date"].month == m]
            attended_present = sum(1 for item in month_records if item["status"] in ("PRESENT", "LATE"))
            total_enrolled = len(month_records)
            attendance_pct = (attended_present / total_enrolled) if total_enrolled > 0 else 1.0
            unique_trainings = len(set(item["training_id"] for item in month_records))
            month_hours = sum(item["hours"] for item in month_records)
            
            target_hours_val = len(filtered_employees) * 1.33
            
            month_ws.cell(r, 1, month_name)
            month_ws.cell(r, 2, round(target_hours_val, 1))
            month_ws.cell(r, 3, round(month_hours, 1))
            month_ws.cell(r, 4, f"=IF(B{r}>0, C{r}/B{r}, 0)")
            month_ws.cell(r, 5, attendance_pct)
            month_ws.cell(r, 6, unique_trainings)
            
            if idx == 0:
                month_ws.cell(r, 7, "—")
            else:
                month_ws.cell(r, 7, f"=IF(C{r-1}>0, (C{r}-C{r-1})/C{r-1}, 0)")
                
            for col_i in range(1, 8):
                cell = month_ws.cell(r, col_i)
                style_cell(cell, align="left" if col_i == 1 else "center")
                if col_i in (2, 3):
                    cell.number_format = "#,##0.0"
                elif col_i in (4, 5, 7):
                    cell.number_format = "0.0%"
                elif col_i == 6:
                    cell.number_format = "#,##0"
            month_ws.row_dimensions[r].height = 20
            
        r_tot = 4 + len(months)
        month_ws.cell(r_tot, 1, "Total")
        month_ws.cell(r_tot, 2, f"=SUM(B4:B{r_tot-1})")
        month_ws.cell(r_tot, 3, f"=SUM(C4:C{r_tot-1})")
        month_ws.cell(r_tot, 4, f"=IF(B{r_tot}>0, C{r_tot}/B{r_tot}, 0)")
        month_ws.cell(r_tot, 5, f"=AVERAGE(E4:E{r_tot-1})")
        month_ws.cell(r_tot, 6, f"=SUM(F4:F{r_tot-1})")
        month_ws.cell(r_tot, 7, "")
        
        for col_i in range(1, 8):
            cell = month_ws.cell(r_tot, col_i)
            style_cell(cell, fill=light_blue, bold=True)
            if col_i in (2, 3):
                cell.number_format = "#,##0.0"
            elif col_i in (4, 5):
                cell.number_format = "0.0%"
            elif col_i == 6:
                cell.number_format = "#,##0"
        month_ws.row_dimensions[r_tot].height = 22
        
        for i, w in enumerate([16, 16, 22, 16, 18, 16, 18], 1):
            month_ws.column_dimensions[get_column_letter(i)].width = w

        # 13. Populate Sheet 4: Training Attendance Summary
        att_ws.merge_cells("A1:J1")
        att_ws["A1"] = "Training Session Attendance Summary"
        att_ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=white)
        att_ws["A1"].fill = PatternFill("solid", fgColor=deep_blue)
        att_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        att_ws.row_dimensions[1].height = 30
        
        att_headers = [
            "Training Name", "Department", "Training Date", "Enrolled Count",
            "Present Count", "Absent Count", "Late Count", "Attendance %",
            "Trainer", "Duration (Hours)"
        ]
        for c_idx, h in enumerate(att_headers, 1):
            cell = att_ws.cell(3, c_idx, h)
            style_cell(cell, fill=blue, font_color=white, bold=True)
        att_ws.row_dimensions[3].height = 24
        
        training_groups = {}
        for item in attendance_rows:
            t_id = item["training_id"]
            training_groups.setdefault(t_id, []).append(item)
            
        sorted_trainings = sorted(training_groups.values(), key=lambda items: items[0]["training_date"] or date.min)
        if len(sorted_trainings) == 0:
            att_ws.merge_cells("A4:J4")
            att_ws.cell(4, 1, "No training attendance data available")
            style_cell(att_ws.cell(4, 1), align="center", bold=True)
            r_tot = 5
            att_ws.cell(r_tot, 1, "Total / Average")
            att_ws.cell(r_tot, 2, "")
            att_ws.cell(r_tot, 3, "")
            att_ws.cell(r_tot, 4, 0)
            att_ws.cell(r_tot, 5, 0)
            att_ws.cell(r_tot, 6, 0)
            att_ws.cell(r_tot, 7, 0)
            att_ws.cell(r_tot, 8, 1.0)
            att_ws.cell(r_tot, 9, "")
            att_ws.cell(r_tot, 10, 0.0)
        else:
            for idx, items in enumerate(sorted_trainings):
                r = 4 + idx
                first_item = items[0]
                t_name = first_item["training_title"]
                t_date = first_item["training_date"]
                trainer = first_item.get("trainer", "")
                duration = first_item.get("duration", 0.0)
                
                depts = sorted(list(set(item["department"] for item in items if item["department"])))
                dept_str = ", ".join(depts) if depts else "Unassigned"
                
                enrolled = len(items)
                present = sum(1 for item in items if item["status"] == "PRESENT")
                absent = sum(1 for item in items if item["status"] == "ABSENT")
                late = sum(1 for item in items if item["status"] == "LATE")
                att_pct = ((present + late) / enrolled) if enrolled > 0 else 1.0
                
                att_ws.cell(r, 1, t_name)
                att_ws.cell(r, 2, dept_str)
                att_ws.cell(r, 3, t_date.strftime("%d-%b-%Y") if t_date else "")
                att_ws.cell(r, 4, enrolled)
                att_ws.cell(r, 5, present)
                att_ws.cell(r, 6, absent)
                att_ws.cell(r, 7, late)
                att_ws.cell(r, 8, att_pct)
                att_ws.cell(r, 9, trainer)
                att_ws.cell(r, 10, duration)
                
                for col_i in range(1, 11):
                    cell = att_ws.cell(r, col_i)
                    style_cell(cell, align="left" if col_i in (1, 2, 9) else "center")
                    if col_i in (4, 5, 6, 7):
                        cell.number_format = "#,##0"
                    elif col_i == 8:
                        cell.number_format = "0.0%"
                    elif col_i == 10:
                        cell.number_format = "#,##0.0"
                att_ws.row_dimensions[r].height = 20
                
            r_tot = 4 + len(sorted_trainings)
            att_ws.cell(r_tot, 1, "Total / Average")
            att_ws.cell(r_tot, 2, "")
            att_ws.cell(r_tot, 3, "")
            att_ws.cell(r_tot, 4, f"=SUM(D4:D{r_tot-1})")
            att_ws.cell(r_tot, 5, f"=SUM(E4:E{r_tot-1})")
            att_ws.cell(r_tot, 6, f"=SUM(F4:F{r_tot-1})")
            att_ws.cell(r_tot, 7, f"=SUM(G4:G{r_tot-1})")
            att_ws.cell(r_tot, 8, f"=IF(D{r_tot}>0, (E{r_tot}+G{r_tot})/D{r_tot}, 1.0)")
            att_ws.cell(r_tot, 9, "")
            att_ws.cell(r_tot, 10, f"=SUM(J4:J{r_tot-1})")
            
        for col_i in range(1, 11):
            cell = att_ws.cell(r_tot, col_i)
            style_cell(cell, fill=light_blue, bold=True)
            if col_i in (4, 5, 6, 7):
                cell.number_format = "#,##0"
            elif col_i == 8:
                cell.number_format = "0.0%"
            elif col_i == 10:
                cell.number_format = "#,##0.0"
        att_ws.row_dimensions[r_tot].height = 22

        
        for i, w in enumerate([32, 24, 16, 14, 14, 14, 14, 14, 20, 16], 1):
            att_ws.column_dimensions[get_column_letter(i)].width = w

        # Auto-fit hidden worksheets column widths
        for ws in [emp_summary_ws, raw_ws]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    async def get_team_dashboard(
        db: AsyncSession,
        financial_year: Optional[str] = None,
        quarter: Optional[str] = None,
        month: Optional[int] = None,
        department_id: Optional[str] = None,
        employee_id: Optional[str] = None,
        manager_id: Optional[str] = None,
        training_category_id: Optional[str] = None,
        training_type: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> TeamAnalyticsDashboard:
        # 1. Determine the bounds
        fy_start, fy_end, fy_label = TeamAnalyticsService.financial_year_bounds(financial_year)
        date_start = start_date or fy_start
        date_end = end_date or fy_end

        if not start_date and not end_date:
            if quarter and quarter != "All":
                if quarter == "Q1":
                    date_start = date(fy_start.year, 4, 1)
                    date_end = date(fy_start.year, 6, 30)
                elif quarter == "Q2":
                    date_start = date(fy_start.year, 7, 1)
                    date_end = date(fy_start.year, 9, 30)
                elif quarter == "Q3":
                    date_start = date(fy_start.year, 10, 1)
                    date_end = date(fy_start.year, 12, 31)
                elif quarter == "Q4":
                    date_start = date(fy_start.year + 1, 1, 1)
                    date_end = date(fy_start.year + 1, 3, 31)
            elif month and 1 <= month <= 12:
                year = fy_start.year if month >= 4 else fy_start.year + 1
                date_start = date(year, month, 1)
                if month == 12:
                    date_end = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    date_end = date(year, month + 1, 1) - timedelta(days=1)

        dept_uuid = TeamAnalyticsService.parse_uuid(department_id)
        emp_uuid = TeamAnalyticsService.parse_uuid(employee_id)
        mgr_uuid = TeamAnalyticsService.parse_uuid(manager_id)
        category_uuid = TeamAnalyticsService.parse_uuid(training_category_id)

        # ── Departments ────────────────────────────────────────────────────────
        depts_res = (await db.execute(
            select(Department).where(Department.deleted_at == None)
        )).scalars().all()
        dept_map = {d.id: d.name for d in depts_res}

        # ── Filter Employees ───────────────────────────────────────────────────
        stmt_emp = select(Employee).options(selectinload(Employee.department), selectinload(Employee.manager)).where(Employee.deleted_at == None)
        if dept_uuid:
            stmt_emp = stmt_emp.where(Employee.department_id == dept_uuid)
        if emp_uuid:
            stmt_emp = stmt_emp.where(Employee.id == emp_uuid)
        if mgr_uuid:
            stmt_emp = stmt_emp.where(Employee.manager_id == mgr_uuid)

        employees = (await db.execute(stmt_emp)).scalars().all()
        employee_ids = [emp.id for emp in employees]

        if not employee_ids:
            return TeamAnalyticsDashboard(
                kpis=TeamKPIs(
                    total_learning_hours=0.0,
                    completion_rate=0.0,
                    active_learners=0,
                    top_performing_department="N/A",
                    avg_hours_per_employee=0.0,
                    total_employees=0,
                    total_enrollments=0,
                    total_completed=0,
                    total_target_hours=0.0,
                    total_actual_hours=0.0,
                    remaining_hours=0.0,
                    learning_compliance_pct=100.0,
                    employees_achieved_goal=0,
                    employees_below_target=0,
                    total_attendance_pct=100.0,
                    learning_hours_generated=0.0,
                    missed_learning_hours=0.0,
                    training_participation_pct=0.0,
                ),
                dept_learning_hours=[],
                monthly_trends=[],
                course_participation=[],
                skill_gaps=[],
                effectiveness_scores=[],
                department_summaries=[],
                employee_table=[],
                top_learners_company=[],
                executive_insights=["No employees match the active filters."],
            )

        # ── Query Enrollments ───────────────────────────────────────────────────
        enrollment_stmt = (
            select(Enrollment, Training)
            .join(Training, Enrollment.training_id == Training.id)
            .where(and_(
                Enrollment.employee_id.in_(employee_ids),
                Enrollment.deleted_at == None,
                Training.status != TrainingStatus.CANCELLED,
                Training.start_date >= date_start,
                Training.start_date <= date_end,
            ))
        )
        if category_uuid:
            enrollment_stmt = enrollment_stmt.where(Training.category_id == category_uuid)
        if training_type:
            enrollment_stmt = enrollment_stmt.where(Training.training_type == training_type)

        enrollment_results = (await db.execute(enrollment_stmt)).all()

        # ── Query Attendance & Roster ─────────────────────────────────────────
        attendance_stmt = (
            select(AttendanceRecord, Training, Employee, Enrollment)
            .join(Training, AttendanceRecord.training_id == Training.id)
            .join(Employee, AttendanceRecord.employee_id == Employee.id)
            .join(Enrollment, and_(
                Enrollment.employee_id == AttendanceRecord.employee_id,
                Enrollment.training_id == AttendanceRecord.training_id,
                Enrollment.deleted_at == None,
            ))
            .where(and_(
                AttendanceRecord.employee_id.in_(employee_ids),
                AttendanceService.roster_record_filter(),
                Training.status != TrainingStatus.CANCELLED,
                Training.start_date >= date_start,
                Training.start_date <= date_end,
            ))
        )
        if category_uuid:
            attendance_stmt = attendance_stmt.where(Training.category_id == category_uuid)
        if training_type:
            attendance_stmt = attendance_stmt.where(Training.training_type == training_type)

        attendance_results = (await db.execute(attendance_stmt)).all()

        # ── Process Calculations ──────────────────────────────────────────────
        target_per_employee = TeamAnalyticsService.LEARNING_TARGET_HOURS
        if month is not None:
            target_per_employee = 1.33
        elif quarter and quarter != "All":
            target_per_employee = 4.0

        emp_actual_hours = {emp_id: 0.0 for emp_id in employee_ids}
        emp_attendance_present = {emp_id: 0 for emp_id in employee_ids}
        emp_attendance_total = {emp_id: 0 for emp_id in employee_ids}
        emp_missed_hours = {emp_id: 0.0 for emp_id in employee_ids}

        fy_m_list = TeamAnalyticsService.fy_months(fy_start)
        monthly_data = {(y, m): {"hours": 0.0, "enrollments": 0, "completions": 0} for y, m, _ in fy_m_list}
        course_data = {}
        skills_covered_progress = {}

        for enrollment, training in enrollment_results:
            emp_id = enrollment.employee_id
            t_title = training.title
            
            c_info = course_data.setdefault(t_title, {"participants": set(), "completions": 0, "total": 0})
            c_info["participants"].add(emp_id)
            c_info["total"] += 1
            if enrollment.status == EnrollmentStatus.COMPLETED:
                c_info["completions"] += 1
            
            t_date = training.start_date
            if t_date:
                key = (t_date.year, t_date.month)
                if key in monthly_data:
                    monthly_data[key]["enrollments"] += 1
                    if enrollment.status == EnrollmentStatus.COMPLETED:
                        monthly_data[key]["completions"] += 1
            
            if training.skills_covered:
                prog = float(enrollment.progress or 0.0)
                for s in training.skills_covered.split(","):
                    s = s.strip()
                    if s:
                        skills_covered_progress.setdefault(s, []).append(prog)

        for record, training, employee, enrollment in attendance_results:
            emp_id = employee.id
            att_status = record.status.value if hasattr(record.status, "value") else str(record.status)
            hours = float(training.duration_hours or 0.0)
            
            emp_attendance_total[emp_id] += 1
            
            if att_status in ("PRESENT", "LATE"):
                emp_attendance_present[emp_id] += 1
                if enrollment.status == EnrollmentStatus.COMPLETED:
                    emp_actual_hours[emp_id] += hours
                    t_date = training.start_date
                    if t_date:
                        key = (t_date.year, t_date.month)
                        if key in monthly_data:
                            monthly_data[key]["hours"] += hours
            elif att_status == "ABSENT":
                emp_missed_hours[emp_id] += hours

        # KPIs
        total_employees = len(employee_ids)
        total_learning_hours = sum(emp_actual_hours.values())
        total_enrollments = len(enrollment_results)
        total_completed = sum(1 for e, t in enrollment_results if e.status == EnrollmentStatus.COMPLETED)
        completion_rate = round((total_completed / total_enrollments * 100.0) if total_enrollments else 0.0, 1)
        active_learners = len(set(e.employee_id for e, t in enrollment_results if e.status == EnrollmentStatus.APPROVED))
        avg_hours = round(total_learning_hours / total_employees, 1) if total_employees else 0.0

        total_target_hours = total_employees * target_per_employee
        total_actual_hours = total_learning_hours
        remaining_hours = sum(max(0.0, target_per_employee - emp_actual_hours[emp_id]) for emp_id in employee_ids)
        employees_achieved_goal = sum(1 for emp_id in employee_ids if emp_actual_hours[emp_id] >= target_per_employee)
        employees_below_target = total_employees - employees_achieved_goal
        learning_compliance_pct = round((employees_achieved_goal / total_employees * 100.0) if total_employees else 100.0, 1)

        total_present_late = sum(emp_attendance_present.values())
        total_attendance_records = sum(emp_attendance_total.values())
        total_attendance_pct = round((total_present_late / total_attendance_records * 100.0) if total_attendance_records else 100.0, 1)

        learning_hours_generated = total_actual_hours
        missed_learning_hours = sum(emp_missed_hours.values())
        training_participation_pct = round((sum(1 for emp_id in employee_ids if emp_actual_hours[emp_id] > 0.0) / total_employees * 100.0) if total_employees else 0.0, 1)

        # Department-level contribution
        dept_employees = {}
        for emp in employees:
            dname = emp.department.name if emp.department else "Unassigned"
            dept_employees.setdefault(dname, []).append(emp.id)

        dept_learning_hours = []
        top_dept_name = "N/A"
        top_dept_hours = -1.0
        lowest_dept_name = "N/A"
        lowest_dept_compliance = 101.0

        for dname, d_emp_ids in dept_employees.items():
            d_hours = sum(emp_actual_hours[emp_id] for emp_id in d_emp_ids)
            d_learners = len(set(e.employee_id for e, t in enrollment_results if e.employee_id in d_emp_ids))
            d_enrollments = [e for e, t in enrollment_results if e.employee_id in d_emp_ids]
            d_completions = sum(1 for e in d_enrollments if e.status == EnrollmentStatus.COMPLETED)
            d_rate = round((d_completions / len(d_enrollments) * 100.0) if d_enrollments else 0.0, 1)
            
            dept_learning_hours.append(DeptLearningHours(
                department=dname,
                hours=round(d_hours, 1),
                employees=d_learners,
                completion_rate=d_rate
            ))
            
            if d_hours > top_dept_hours:
                top_dept_hours = d_hours
                top_dept_name = dname

            # Calculate compliance for this department to identify lowest performing
            d_achieved = sum(1 for emp_id in d_emp_ids if emp_actual_hours[emp_id] >= target_per_employee)
            d_compliance = (d_achieved / len(d_emp_ids) * 100.0) if d_emp_ids else 100.0
            if d_compliance < lowest_dept_compliance:
                lowest_dept_compliance = d_compliance
                lowest_dept_name = dname

        dept_learning_hours = sorted(dept_learning_hours, key=lambda x: x.department)

        # Monthly trends
        monthly_trends = []
        for y, m, m_name in fy_m_list:
            trend_data = monthly_data[(y, m)]
            monthly_trends.append(MonthlyLearningTrend(
                month=m_name[:3],
                hours=round(trend_data["hours"], 1),
                enrollments=trend_data["enrollments"],
                completions=trend_data["completions"]
            ))

        # Course participation
        course_participation = []
        for c_title, c_info in course_data.items():
            course_participation.append(CourseParticipation(
                course=c_title,
                participants=len(c_info["participants"]),
                completion_rate=round((c_info["completions"] / c_info["total"] * 100.0) if c_info["total"] else 0.0, 1)
            ))
        course_participation = sorted(course_participation, key=lambda x: x.participants, reverse=True)[:8]

        # Skills gap
        skill_gaps = []
        for skill, progs in skills_covered_progress.items():
            current = round(sum(progs) / len(progs), 1) if progs else 0.0
            target = min(100.0, current + 20.0)
            skill_gaps.append(SkillGapItem(skill=skill, current=current, target=target))
        skill_gaps = sorted(skill_gaps, key=lambda x: x.current, reverse=True)[:7]
        if not skill_gaps:
            defaults = [
                ("Leadership", 60.0, 85.0), ("Communication", 72.0, 90.0),
                ("Technical Skills", 55.0, 80.0), ("Data Analysis", 40.0, 75.0),
                ("Project Management", 65.0, 85.0),
            ]
            skill_gaps = [SkillGapItem(skill=s, current=c, target=t) for s, c, t in defaults]

        # Effectiveness Scores
        eff_stmt = (
            select(
                Employee.department_id,
                func.avg(Effectiveness.score).label("avg_score"),
            )
            .join(Enrollment, Effectiveness.enrollment_id == Enrollment.id)
            .join(Employee, Enrollment.employee_id == Employee.id)
            .join(Training, Enrollment.training_id == Training.id)
            .where(and_(
                Employee.id.in_(employee_ids),
                Effectiveness.deleted_at == None,
                Effectiveness.score != None,
                Training.status != TrainingStatus.CANCELLED,
                Training.start_date >= date_start,
                Training.start_date <= date_end,
            ))
            .group_by(Employee.department_id)
        )
        if category_uuid:
            eff_stmt = eff_stmt.where(Training.category_id == category_uuid)
        if training_type:
            eff_stmt = eff_stmt.where(Training.training_type == training_type)

        eff_raw = (await db.execute(eff_stmt)).all()
        effectiveness_scores = [
            EffectivenessScore(
                department=dept_map.get(r.department_id, "Unknown"),
                score=round(float(r.avg_score or 0) * 20, 1),
            )
            for r in eff_raw
        ]

        # Department summaries
        department_summaries = []
        for dept in depts_res:
            dept_emps = [emp for emp in employees if emp.department_id == dept.id]
            dept_emp_count = len(dept_emps)
            if dept_emp_count == 0:
                continue
            dept_emp_ids = [emp.id for emp in dept_emps]
            
            d_hours = sum(emp_actual_hours[emp_id] for emp_id in dept_emp_ids)
            d_enrs = [e for e, t in enrollment_results if e.employee_id in dept_emp_ids]
            d_total_enr = len(d_enrs)
            d_comp = sum(1 for e in d_enrs if e.status == EnrollmentStatus.COMPLETED)
            d_rate = round((d_comp / d_total_enr * 100.0) if d_total_enr else 0.0, 1)
            d_active = sum(1 for e in d_enrs if e.status == EnrollmentStatus.APPROVED)
            
            d_pending_eff = (await db.execute(
                select(func.count(Effectiveness.id))
                .join(Enrollment, Effectiveness.enrollment_id == Enrollment.id)
                .join(Training, Enrollment.training_id == Training.id)
                .where(and_(
                    Enrollment.employee_id.in_(dept_emp_ids),
                    Effectiveness.status == EffectivenessStatus.SUBMITTED,
                    Effectiveness.deleted_at == None,
                    Training.status != TrainingStatus.CANCELLED,
                    Training.start_date >= date_start,
                    Training.start_date <= date_end,
                ))
            )).scalar() or 0
            
            dept_learner_hours = {emp_id: emp_actual_hours[emp_id] for emp_id in dept_emp_ids}
            if dept_learner_hours:
                top_emp_id = max(dept_learner_hours, key=dept_learner_hours.get)
                top_emp = next(emp for emp in dept_emps if emp.id == top_emp_id)
                top_learner_name = f"{top_emp.first_name} {top_emp.last_name}" if dept_learner_hours[top_emp_id] > 0.0 else "—"
            else:
                top_learner_name = "—"
                
            department_summaries.append(DepartmentSummary(
                department=dept.name,
                top_learner=top_learner_name,
                total_hours=round(d_hours, 1),
                completion_pct=d_rate,
                active_enrollments=d_active,
                pending_evaluations=d_pending_eff,
                employee_count=dept_emp_count
            ))

        # Employee analytics table
        emp_eff_stmt = (
            select(
                Enrollment.employee_id,
                func.avg(Effectiveness.score).label("avg_score")
            )
            .join(Effectiveness, Effectiveness.enrollment_id == Enrollment.id)
            .join(Training, Enrollment.training_id == Training.id)
            .where(and_(
                Enrollment.employee_id.in_(employee_ids),
                Effectiveness.deleted_at == None,
                Effectiveness.score != None,
                Training.status != TrainingStatus.CANCELLED,
                Training.start_date >= date_start,
                Training.start_date <= date_end,
            ))
            .group_by(Enrollment.employee_id)
        )
        if category_uuid:
            emp_eff_stmt = emp_eff_stmt.where(Training.category_id == category_uuid)
        if training_type:
            emp_eff_stmt = emp_eff_stmt.where(Training.training_type == training_type)

        emp_eff_res = (await db.execute(emp_eff_stmt)).all()
        emp_eff_map = {r.employee_id: float(r.avg_score or 0.0) * 20.0 for r in emp_eff_res}

        emp_last_active_stmt = (
            select(
                Enrollment.employee_id,
                func.max(Enrollment.updated_at).label("last_active")
            )
            .join(Training, Enrollment.training_id == Training.id)
            .where(and_(
                Enrollment.employee_id.in_(employee_ids),
                Enrollment.deleted_at == None,
                Training.status != TrainingStatus.CANCELLED,
                Training.start_date >= date_start,
                Training.start_date <= date_end,
            ))
            .group_by(Enrollment.employee_id)
        )
        if category_uuid:
            emp_last_active_stmt = emp_last_active_stmt.where(Training.category_id == category_uuid)
        if training_type:
            emp_last_active_stmt = emp_last_active_stmt.where(Training.training_type == training_type)

        emp_last_active_res = (await db.execute(emp_last_active_stmt)).all()
        emp_last_active_map = {r.employee_id: r.last_active for r in emp_last_active_res}

        employee_table = []
        for emp in employees:
            emp_id = emp.id
            emp_enrs = [e for e, t in enrollment_results if e.employee_id == emp_id]
            emp_total_enr = len(emp_enrs)
            emp_comp = sum(1 for e in emp_enrs if e.status == EnrollmentStatus.COMPLETED)
            emp_comp_pct = round((emp_comp / emp_total_enr * 100.0) if emp_total_enr else 0.0, 1)
            
            eff = round(emp_eff_map.get(emp_id, 0.0), 1)
            last_act = emp_last_active_map.get(emp_id)
            last_active_str = last_act.strftime("%Y-%m-%d") if last_act else None
            
            employee_table.append(EmployeeAnalyticsRow(
                id=str(emp_id),
                name=f"{emp.first_name} {emp.last_name}",
                department=emp.department.name if emp.department else "N/A",
                total_hours=round(emp_actual_hours[emp_id], 1),
                trainings_completed=emp_comp,
                completion_pct=emp_comp_pct,
                effectiveness_score=eff,
                last_active=last_active_str
            ))
        employee_table = sorted(employee_table, key=lambda x: x.total_hours, reverse=True)[:50]

        # Top learners
        top_learners_company = []
        for idx, row in enumerate(sorted(employee_table, key=lambda x: x.total_hours, reverse=True)[:10]):
            initials = ""
            if row.name:
                parts = row.name.split()
                if len(parts) >= 2:
                    initials = parts[0][0].upper() + parts[1][0].upper()
                elif len(parts) == 1:
                    initials = parts[0][0].upper()
            if not initials:
                initials = "??"
                
            top_learners_company.append(TopLearner(
                rank=idx + 1,
                id=row.id,
                name=row.name,
                department=row.department,
                hours=row.total_hours,
                completions=row.trainings_completed,
                avatar_initials=initials
            ))

        # Executive insights
        executive_insights = []
        if top_dept_name != "N/A" and top_dept_hours > 0.0:
            executive_insights.append(f"• {top_dept_name} contributed the highest learning hours ({top_dept_hours:.1f} hours).")
        
        if lowest_dept_name != "N/A" and lowest_dept_compliance < 100.0:
            executive_insights.append(f"• {lowest_dept_name} is below annual learning target with a {lowest_dept_compliance:.1f}% compliance rate.")
        
        executive_insights.append(f"• Organization-wide learning compliance stands at {learning_compliance_pct:.1f}%.")
        
        if missed_learning_hours > 0:
            executive_insights.append(f"• Total missed learning hours due to absences is {missed_learning_hours:.1f} hours.")
        
        if total_attendance_pct < 90.0:
            executive_insights.append(f"• Attendance rate is {total_attendance_pct:.1f}%, which has a direct correlation to overall completion performance.")

        kpis = TeamKPIs(
            total_learning_hours=round(float(total_learning_hours), 1),
            completion_rate=completion_rate,
            active_learners=active_learners,
            top_performing_department=top_dept_name,
            avg_hours_per_employee=avg_hours,
            total_employees=total_employees,
            total_enrollments=total_enrollments,
            total_completed=total_completed,
            total_target_hours=round(float(total_target_hours), 1),
            total_actual_hours=round(float(total_actual_hours), 1),
            remaining_hours=round(float(remaining_hours), 1),
            learning_compliance_pct=learning_compliance_pct,
            employees_achieved_goal=employees_achieved_goal,
            employees_below_target=employees_below_target,
            total_attendance_pct=total_attendance_pct,
            learning_hours_generated=round(float(learning_hours_generated), 1),
            missed_learning_hours=round(float(missed_learning_hours), 1),
            training_participation_pct=training_participation_pct,
        )

        return TeamAnalyticsDashboard(
            kpis=kpis,
            dept_learning_hours=dept_learning_hours,
            monthly_trends=monthly_trends,
            course_participation=course_participation,
            skill_gaps=skill_gaps,
            effectiveness_scores=effectiveness_scores,
            department_summaries=department_summaries,
            employee_table=employee_table,
            top_learners_company=top_learners_company,
            executive_insights=executive_insights,
        )
