    async def generate_kpi_excel(
        db: AsyncSession,
        financial_year: Optional[str] = None,
        month: Optional[int] = None,
        department_id: Optional[str] = None,
        employee_id: Optional[str] = None,
        manager_id: Optional[str] = None,
        training_id: Optional[str] = None,
        training_category_id: Optional[str] = None,
        attendance_status: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        fy_start, fy_end, fy_label = TeamAnalyticsService.financial_year_bounds(financial_year)
        range_start = start_date or fy_start
        range_end = end_date or fy_end
        months = TeamAnalyticsService.fy_months(fy_start)
        if month:
            months = [m for m in months if m[1] == month]

        dept_uuid = TeamAnalyticsService.parse_uuid(department_id)
        emp_uuid = TeamAnalyticsService.parse_uuid(employee_id)
        mgr_uuid = TeamAnalyticsService.parse_uuid(manager_id)
        training_uuid = TeamAnalyticsService.parse_uuid(training_id)
        category_uuid = TeamAnalyticsService.parse_uuid(training_category_id)

        # Learning KPI credit is intentionally strict: only completed trainings
        # with roster attendance marked PRESENT contribute hours.
        status_enum = AttendanceStatus.PRESENT

        employees_stmt = (
            select(Employee)
            .options(selectinload(Employee.department), selectinload(Employee.manager))
            .where(Employee.deleted_at == None)
            .order_by(Employee.employee_code.asc())
        )
        if dept_uuid:
            employees_stmt = employees_stmt.where(Employee.department_id == dept_uuid)
        if emp_uuid:
            employees_stmt = employees_stmt.where(Employee.id == emp_uuid)
        if mgr_uuid:
            employees_stmt = employees_stmt.where(Employee.manager_id == mgr_uuid)
        employees = (await db.execute(employees_stmt)).scalars().all()
        employee_ids = [emp.id for emp in employees]

        contribution_map: dict = {emp.id: {} for emp in employees}
        attendance_rows = []
        if employee_ids:
            stmt = (
                select(AttendanceRecord, Training, Employee)
                .join(Training, AttendanceRecord.training_id == Training.id)
                .join(Employee, AttendanceRecord.employee_id == Employee.id)
                .join(Enrollment, and_(
                    Enrollment.employee_id == AttendanceRecord.employee_id,
                    Enrollment.training_id == AttendanceRecord.training_id,
                    Enrollment.deleted_at == None,
                ))
                .options(selectinload(Employee.department), selectinload(Employee.manager), selectinload(Training.category))
                .where(and_(
                    AttendanceRecord.employee_id.in_(employee_ids),
                    AttendanceRecord.status == status_enum,
                    AttendanceService.roster_record_filter(),
                    Enrollment.status == EnrollmentStatus.COMPLETED,
                    Training.status != TrainingStatus.CANCELLED,
                    Training.start_date >= range_start,
                    Training.start_date <= range_end,
                ))
            )
            if training_uuid:
                stmt = stmt.where(Training.id == training_uuid)
            if category_uuid:
                stmt = stmt.where(Training.category_id == category_uuid)
            raw_rows = (await db.execute(stmt.order_by(Training.start_date.asc(), Training.title.asc()))).all()
            for record, training, employee in raw_rows:
                if not training.start_date:
                    continue
                key = (training.start_date.year, training.start_date.month)
                contribution = {
                    "employee_id": employee.id,
                    "employee_name": f"{employee.first_name} {employee.last_name}",
                    "employee_code": employee.employee_code,
                    "department": employee.department.name if employee.department else "Unassigned",
                    "manager": f"{employee.manager.first_name} {employee.manager.last_name}" if employee.manager else "",
                    "training_id": training.id,
                    "training_title": training.title,
                    "training_date": training.start_date,
                    "hours": float(training.duration_hours or 0),
                    "max_hours": float(training.max_hours_allowed or training.duration_hours or 0),
                    "status": record.status.value,
                    "marked_at": record.marked_at,
                    "category": training.category.name if training.category else "",
                }
                contribution_map.setdefault(employee.id, {}).setdefault(key, []).append(contribution)
                attendance_rows.append(contribution)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employee Learning KPI Report"
        ws.freeze_panes = "E6"

        blue = "4F81BD"
        deep_blue = "1F4E79"
        light_blue = "D9EAF7"
        green = "E2F0D9"
        yellow = "FFF2CC"
        red = "FCE4D6"
        white = "FFFFFF"
        thin = Side(style="thin", color="9EADCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_cell(cell, fill=white, font_color="1F2937", bold=False, align="center"):
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Calibri", size=9, bold=bold, color=font_color)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = border

        total_cols = 4 + len(months) * 3 + 8
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(1, 1, "Department KPI Dashboard - Employee Learning KPI Report")
        style_cell(ws.cell(1, 1), fill=deep_blue, font_color=white, bold=True)
        ws.cell(1, 1).font = Font(name="Calibri", size=14, bold=True, color=white)

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        filter_summary = f"FY: Apr {fy_start.year} - Mar {fy_end.year} | Date Range: {range_start:%d-%b-%Y} to {range_end:%d-%b-%Y}"
        if month:
            filter_summary += f" | Month: {date(fy_start.year if month >= 4 else fy_end.year, month, 1):%B}"
        ws.cell(2, 1, filter_summary)
        style_cell(ws.cell(2, 1), fill=light_blue, bold=True, align="left")

        base_headers = ["S.No", "Employee ID", "Employee Name", "Date of Joining"]
        for idx, header in enumerate(base_headers, 1):
            ws.merge_cells(start_row=4, start_column=idx, end_row=5, end_column=idx)
            ws.cell(4, idx, header)
            style_cell(ws.cell(4, idx), fill=blue, font_color=white, bold=True)

        col = 5
        for _, _, month_name in months:
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 2)
            ws.cell(4, col, month_name.upper())
            style_cell(ws.cell(4, col), fill=blue, font_color=white, bold=True)
            for sub in ["Date", "Actual", "Max Hours"]:
                ws.cell(5, col, sub)
                style_cell(ws.cell(5, col), fill=blue, font_color=white, bold=True)
                col += 1

        final_headers = ["Total Hours", "Target Hours", "Balance", "KPI %", "Learning Compliance", "Department", "Manager", "Attendance %"]
        for header in final_headers:
            ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
            ws.cell(4, col, header)
            style_cell(ws.cell(4, col), fill=blue, font_color=white, bold=True)
            col += 1

        dept_stats: dict = {}
        monthly_stats: dict = {(y, m): {"hours": 0.0, "employees": set(), "trainings": set()} for y, m, _ in months}

        for row_idx, emp in enumerate(employees, 6):
            row_values = [
                row_idx - 5,
                emp.employee_code,
                f"{emp.first_name} {emp.last_name}",
                emp.date_of_joining.strftime("%d %m %Y") if emp.date_of_joining else "",
            ]
            total_hours = 0.0
            total_attendance_rows = 0
            for y, m, _ in months:
                entries = contribution_map.get(emp.id, {}).get((y, m), [])
                total_attendance_rows += len(entries)
                month_hours = round(sum(item["hours"] for item in entries), 2)
                max_hours = round(sum(item["max_hours"] for item in entries), 2)
                total_hours += month_hours
                dates = "\n".join(f"{item['training_date']:%d.%m.%Y} - {item['training_title']}" for item in entries)
                row_values.extend([dates, month_hours if month_hours else "", max_hours if max_hours else ""])
                if (y, m) in monthly_stats:
                    monthly_stats[(y, m)]["hours"] += month_hours
                    if month_hours:
                        monthly_stats[(y, m)]["employees"].add(emp.id)
                    for item in entries:
                        monthly_stats[(y, m)]["trainings"].add(item["training_id"])

            target = TeamAnalyticsService.LEARNING_TARGET_HOURS
            balance = max(target - total_hours, 0)
            kpi_pct = round((total_hours / target) * 100, 1) if target else 0
            department = emp.department.name if emp.department else "Unassigned"
            manager = f"{emp.manager.first_name} {emp.manager.last_name}" if emp.manager else ""
            attendance_pct = 100.0 if total_attendance_rows else 0.0
            row_values.extend([
                round(total_hours, 2),
                target,
                f"{round(balance, 2)} Hours Remaining" if balance else "0",
                f"{kpi_pct}%",
                TeamAnalyticsService.compliance_status(kpi_pct),
                department,
                manager,
                f"{attendance_pct:.1f}%",
            ])
            ws.append(row_values)
            compliance_fill = green if kpi_pct >= 75 else yellow if kpi_pct >= 25 else red
            for c in range(1, total_cols + 1):
                fill = compliance_fill if c in (total_cols - 3, total_cols - 4) else white
                style_cell(ws.cell(row_idx, c), fill=fill, align="left" if c in (3, 5) or c > 4 and (c - 5) % 3 == 0 else "center")

            stats = dept_stats.setdefault(department, {
                "employees": 0, "hours": 0.0, "below": 0, "achieved": 0,
                "top_name": "", "top_hours": -1.0, "attendance_rows": 0,
            })
            stats["employees"] += 1
            stats["hours"] += total_hours
            stats["attendance_rows"] += total_attendance_rows
            if total_hours < target:
                stats["below"] += 1
            if total_hours >= target:
                stats["achieved"] += 1
            if total_hours > stats["top_hours"]:
                stats["top_hours"] = total_hours
                stats["top_name"] = f"{emp.first_name} {emp.last_name}"

        widths = [8, 14, 24, 16]
        widths.extend([28, 12, 12] * len(months))
        widths.extend([12, 12, 18, 10, 20, 18, 22, 14])
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 36 if row >= 6 else 24
        ws.auto_filter.ref = f"A5:{get_column_letter(total_cols)}{ws.max_row}"

        dept_ws = wb.create_sheet("Department KPI Analytics")
        dept_headers = ["Department", "Total Employees", "Total Learning Hours", "Avg Learning Hours", "Completion %", "Employees Below Target", "Top Learners", "Attendance %", "KPI Score"]
        dept_ws.append(dept_headers)
        for c in range(1, len(dept_headers) + 1):
            style_cell(dept_ws.cell(1, c), fill=blue, font_color=white, bold=True)
        for department, stats in sorted(dept_stats.items()):
            completion = round((stats["achieved"] / stats["employees"]) * 100, 1) if stats["employees"] else 0
            avg_hours = round(stats["hours"] / stats["employees"], 2) if stats["employees"] else 0
            kpi_score = round((avg_hours / TeamAnalyticsService.LEARNING_TARGET_HOURS) * 100, 1) if TeamAnalyticsService.LEARNING_TARGET_HOURS else 0
            dept_ws.append([
                department,
                stats["employees"],
                round(stats["hours"], 2),
                avg_hours,
                f"{completion}%",
                stats["below"],
                stats["top_name"],
                "100.0%" if stats["attendance_rows"] else "0.0%",
                f"{kpi_score}%",
            ])
        for row in dept_ws.iter_rows():
            for cell in row:
                style_cell(cell, fill=cell.fill.fgColor.rgb[-6:] if cell.row == 1 else white, font_color=white if cell.row == 1 else "1F2937", bold=cell.row == 1)
        for i, width in enumerate([24, 16, 20, 18, 16, 22, 24, 14, 14], 1):
            dept_ws.column_dimensions[get_column_letter(i)].width = width

        month_ws = wb.create_sheet("Monthly Learning Summary")
        month_ws.append(["Month", "Financial Year", "Total Learning Hours", "Active Learners", "Training Programs", "Monthly KPI Contribution"])
        for c in range(1, 7):
            style_cell(month_ws.cell(1, c), fill=blue, font_color=white, bold=True)
        for y, m, name in months:
            stats = monthly_stats[(y, m)]
            month_ws.append([name, fy_label, round(stats["hours"], 2), len(stats["employees"]), len(stats["trainings"]), f"{round((stats['hours'] / max(1, len(employee_ids) * TeamAnalyticsService.LEARNING_TARGET_HOURS)) * 100, 1)}%"])
        for row in month_ws.iter_rows(min_row=2):
            for cell in row:
                style_cell(cell)
        for i, width in enumerate([16, 16, 22, 16, 18, 24], 1):
            month_ws.column_dimensions[get_column_letter(i)].width = width

        att_ws = wb.create_sheet("Training Attendance Summary")
        att_headers = ["Employee ID", "Employee Name", "Department", "Manager", "Training Date", "Training Name", "Category", "Attendance Status", "Actual Hours", "Max Hours"]
        att_ws.append(att_headers)
        for c in range(1, len(att_headers) + 1):
            style_cell(att_ws.cell(1, c), fill=blue, font_color=white, bold=True)
        for item in attendance_rows:
            att_ws.append([
                item["employee_code"],
                item["employee_name"],
                item["department"],
                item["manager"],
                item["training_date"].strftime("%d-%b-%Y"),
                item["training_title"],
                item["category"],
                item["status"],
                item["hours"],
                item["max_hours"],
            ])
        for row in att_ws.iter_rows(min_row=2):
            for cell in row:
                style_cell(cell, align="left")
        for i, width in enumerate([14, 24, 18, 22, 16, 32, 18, 18, 14, 12], 1):
            att_ws.column_dimensions[get_column_letter(i)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

