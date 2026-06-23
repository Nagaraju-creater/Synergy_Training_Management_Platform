import asyncio
import sys
import os
import io

# Add backend directory to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "backend")))

# Import models to prevent mapper errors
from app.users.models import User
from app.roles.models import Role
from app.employees.models import Employee
from app.departments.models import Department, DepartmentHead
from app.trainings.models import Training, TrainingImportHistory
from app.trainings.categories import TrainingCategory
from app.enrollments.models import Enrollment, EnrollmentStatus
from app.nominations.models import Nomination
from app.effectiveness.models import Effectiveness
from app.effectiveness.reviews import DepartmentReview
from app.attendance.models import AttendanceSession, AttendanceRecord, AttendanceStatus
from app.notifications.models import Notification
from app.signatures.models import DigitalSignature
from app.gamification.models import Achievement, LeaderboardPoint
from app.audit.models import AuditLog
from app.analytics.models import AnalyticsSnapshot
from app.training_plans.models import TrainingPlan

from app.database import AsyncSessionLocal
from app.trainings.service import TrainingService
from sqlalchemy import select, delete
from openpyxl import load_workbook

async def run_tests():
    db = AsyncSessionLocal()
    print("Starting import tests...")
    
    try:
        # Clean up any previous test trainings
        await db.execute(delete(Enrollment).where(Enrollment.training_id.in_(
            select(Training.id).where(Training.title.in_(["Test Historical Excel Import", "Test Upcoming Mandatory Excel Import"]))
        )))
        await db.execute(delete(AttendanceRecord).where(AttendanceRecord.training_id.in_(
            select(Training.id).where(Training.title.in_(["Test Historical Excel Import", "Test Upcoming Mandatory Excel Import"]))
        )))
        await db.execute(delete(Effectiveness).where(Effectiveness.training_id.in_(
            select(Training.id).where(Training.title.in_(["Test Historical Excel Import", "Test Upcoming Mandatory Excel Import"]))
        )))
        await db.execute(delete(Training).where(Training.title.in_(["Test Historical Excel Import", "Test Upcoming Mandatory Excel Import"])))
        await db.commit()
        
        # 1. Generate Import Template
        print("\n1. Generating Excel Template...")
        template_io = await TrainingService.generate_import_template(db)
        wb = load_workbook(template_io)
        ws = wb["Training Import Template"]
        print(f"Sheet Name: {ws.title}")
        print(f"Header Row: {[ws.cell(row=1, column=c).value for c in range(1, 17)]}")
        
        # 2. Add Test Records to Workbook
        # Row 3: Historical completed training (Date in the past)
        ws.cell(row=3, column=1).value = "Test Historical Excel Import"
        ws.cell(row=3, column=2).value = "Test historical training for testing backfill."
        ws.cell(row=3, column=3).value = "Accounts" # Valid department
        ws.cell(row=3, column=4).value = "Technical" # Category
        ws.cell(row=3, column=5).value = "John Doe"
        ws.cell(row=3, column=6).value = "10-06-2026" # Historical date
        ws.cell(row=3, column=7).value = "09:00"
        ws.cell(row=3, column=8).value = "11:00"
        ws.cell(row=3, column=9).value = 2.0
        ws.cell(row=3, column=10).value = "Online"
        ws.cell(row=3, column=11).value = "Zoom link"
        ws.cell(row=3, column=12).value = 30
        ws.cell(row=3, column=13).value = "09-06-2026"
        ws.cell(row=3, column=14).value = "18:00"
        ws.cell(row=3, column=15).value = "Optional"
        ws.cell(row=3, column=16).value = "Scheduled"
        
        # Row 4: Upcoming mandatory training (Date in the future)
        ws.cell(row=4, column=1).value = "Test Upcoming Mandatory Excel Import"
        ws.cell(row=4, column=2).value = "Test upcoming mandatory training for auto-enrollment."
        ws.cell(row=4, column=3).value = "Accounts" # Valid department
        ws.cell(row=4, column=4).value = "Technical"
        ws.cell(row=4, column=5).value = "Jane Doe"
        ws.cell(row=4, column=6).value = "25-06-2026" # Future date
        ws.cell(row=4, column=7).value = "14:00"
        ws.cell(row=4, column=8).value = "16:00"
        ws.cell(row=4, column=9).value = 2.0
        ws.cell(row=4, column=10).value = "In-Person"
        ws.cell(row=4, column=11).value = "Room 303"
        ws.cell(row=4, column=12).value = 40
        ws.cell(row=4, column=13).value = "24-06-2026"
        ws.cell(row=4, column=14).value = "18:00"
        ws.cell(row=4, column=15).value = "Mandatory"
        ws.cell(row=4, column=16).value = "Scheduled"
        
        # Save sheet to bytes
        test_file_io = io.BytesIO()
        wb.save(test_file_io)
        test_file_bytes = test_file_io.getvalue()
        
        # 3. Parse Import File
        print("\n2. Parsing Excel File...")
        parsed_res = await TrainingService.parse_import_file(db, test_file_bytes)
        summary = parsed_res["summary"]
        records = parsed_res["records"]
        print(f"Summary: {summary}")
        print(f"Total Records Parsed: {len(records)}")
        
        # Verify warnings/errors on test rows
        for r in records:
            print(f"Index {r['index']} - Title: '{r['title']}', Valid: {r['is_valid']}, Errors: {r['errors']}, Warnings: {r['warnings']}")
            
        # We expect index 1 (sample row), 2 (test historical), and 3 (test upcoming) to be valid
        assert summary["failed_records"] == 0, f"Expected 0 failed records, got {summary['failed_records']}"
        
        # 4. Confirm Import
        print("\n3. Confirming Import (applying strategies)...")
        # Fetch an admin user id to set as creator
        user_stmt = select(User).where(User.is_active == True).limit(1)
        user_res = await db.execute(user_stmt)
        admin_user = user_res.scalar_one_or_none()
        if not admin_user:
            raise Exception("No active admin user found in database to run import test.")
        admin_id = admin_user.id
        print(f"Using Admin User: {admin_user.email} ({admin_id})")
        
        confirm_res = await TrainingService.confirm_import(
            db,
            records,
            duplicate_strategy="replace", # replace existing matching
            created_by=admin_id,
            source_file="test_import.xlsx"
        )
        print(f"Confirm Results: {confirm_res}")
        
        # 5. Verify Database Records
        print("\n4. Verifying imported trainings in database...")
        # Check historical training
        hist_stmt = select(Training).where(Training.title == "Test Historical Excel Import")
        hist_res = await db.execute(hist_stmt)
        hist_training = hist_res.scalar_one_or_none()
        assert hist_training is not None, "Historical training was not created!"
        print(f"Historical Training: ID={hist_training.id}, Status={hist_training.status}")
        
        # Check upcoming mandatory training
        up_stmt = select(Training).where(Training.title == "Test Upcoming Mandatory Excel Import")
        up_res = await db.execute(up_stmt)
        up_training = up_res.scalar_one_or_none()
        assert up_training is not None, "Upcoming training was not created!"
        print(f"Upcoming Training: ID={up_training.id}, Status={up_training.status}")
        
        # Check enrollments backfilled for historical training (must be COMPLETED with 100% progress)
        enr_stmt = select(Enrollment).where(Enrollment.training_id == hist_training.id)
        enr_res = await db.execute(enr_stmt)
        hist_enrollments = enr_res.scalars().all()
        print(f"Historical enrollments created: {len(hist_enrollments)}")
        for e in hist_enrollments:
            print(f"- Employee ID={e.employee_id}, Status={e.status}, Progress={e.progress}%")
            assert e.status == EnrollmentStatus.COMPLETED, f"Expected COMPLETED status, got {e.status}"
            assert e.progress == 100.0, f"Expected progress=100.0, got {e.progress}"
            
        # Check attendance records backfilled for historical training (must be PRESENT)
        att_stmt = select(AttendanceRecord).where(AttendanceRecord.training_id == hist_training.id)
        att_res = await db.execute(att_stmt)
        hist_attendance = att_res.scalars().all()
        print(f"Historical attendance records created: {len(hist_attendance)}")
        for a in hist_attendance:
            print(f"- Employee ID={a.employee_id}, Status={a.status}")
            assert a.status == AttendanceStatus.PRESENT, f"Expected PRESENT status, got {a.status}"
            
        # Check effectiveness records backfilled for historical training (must be PENDING)
        eff_stmt = select(Effectiveness).where(Effectiveness.training_id == hist_training.id)
        eff_res = await db.execute(eff_stmt)
        hist_effectiveness = eff_res.scalars().all()
        print(f"Historical effectiveness reviews created: {len(hist_effectiveness)}")
        for ef in hist_effectiveness:
            print(f"- Enrollment ID={ef.enrollment_id}, Status={ef.status}")
            
        # Check upcoming mandatory enrollments auto-created (status APPROVED)
        up_enr_stmt = select(Enrollment).where(Enrollment.training_id == up_training.id)
        up_enr_res = await db.execute(up_enr_stmt)
        up_enrollments = up_enr_res.scalars().all()
        print(f"Upcoming mandatory enrollments created: {len(up_enrollments)}")
        for e in up_enrollments:
            print(f"- Employee ID={e.employee_id}, Status={e.status}")
            assert e.status == EnrollmentStatus.APPROVED, f"Expected APPROVED status, got {e.status}"
            
        # Check Import History log
        history_stmt = select(TrainingImportHistory).order_by(TrainingImportHistory.created_at.desc()).limit(1)
        history_res = await db.execute(history_stmt)
        hist_log = history_res.scalar_one_or_none()
        assert hist_log is not None, "Import history log was not created!"
        print(f"Import History log: Imported={hist_log.records_imported}, Failed={hist_log.records_failed}, Skipped={hist_log.records_skipped}, File={hist_log.source_file}")
        
        print("\nALL TESTS PASSED SUCCESSFULLY!")
        
    except Exception as e:
        print(f"\nTESTS FAILED: {e}")
        import traceback
        traceback.print_exc()
    finally:
        await db.close()

if __name__ == "__main__":
    asyncio.run(run_tests())
